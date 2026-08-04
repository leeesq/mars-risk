from __future__ import annotations

from pathlib import Path

import matplotlib.pyplot as plt
import polars as pl
import pytest
from openpyxl import load_workbook

from mars.analysis.missing_shift import (
    MarsMissingShiftConfig,
    MarsMissingShiftResult,
    MarsMissingShiftScanner,
    _apply_benjamini_hochberg,
    _Candidate,
    _two_proportion_test,
    _WindowStats,
)


def _daily_missing_frame(
    *,
    rates: list[float],
    rows_per_day: int | list[int] = 100,
    feature: str = "x",
) -> pl.DataFrame:
    daily_counts = (
        [rows_per_day] * len(rates) if isinstance(rows_per_day, int) else rows_per_day
    )
    rows: list[dict[str, object]] = []
    for day_idx, (rate, daily_count) in enumerate(zip(rates, daily_counts), start=1):
        missing_count = int(round(daily_count * rate))
        for row_idx in range(daily_count):
            rows.append(
                {
                    "apply_dt": f"2026-01-{day_idx:02d}",
                    feature: None if row_idx < missing_count else float(row_idx),
                    "stable": float(row_idx),
                }
            )
    return pl.DataFrame(rows)


def _config(
    *detectors: str,
    **overrides: object,
) -> MarsMissingShiftConfig:
    params: dict[str, object] = {
        "enabled_detectors": detectors
        or ("segment_shift", "boundary", "point", "high_level"),
        "min_period_samples": 30,
        "min_segment_size": 3,
        "reference_window": 3,
        "max_boundary_periods": 3,
        "max_segment_candidates": 3,
        "min_abs_delta": 0.03,
        "min_effect_delta": 0.005,
        "min_relative_delta": 0.30,
        "fdr_q_threshold": 0.05,
        "high_missing_rate_threshold": 0.90,
        "feature_high_missing_rate_thresholds": {},
        "min_high_periods": 2,
    }
    params.update(overrides)
    return MarsMissingShiftConfig(**params)  # type: ignore[arg-type]


def test_missing_shift_scanner_detects_segment_increase_and_decrease() -> None:
    increasing = _daily_missing_frame(rates=[0.05] * 6 + [0.35] * 6)
    decreasing = _daily_missing_frame(rates=[0.80] * 6 + [0.20] * 6)
    scanner = MarsMissingShiftScanner()

    increase_result = scanner.scan(
        increasing,
        date_col="apply_dt",
        features=["x"],
        feature_data_source={"x": "bureau"},
        config=_config("segment_shift"),
    )
    decrease_result = scanner.scan(
        decreasing,
        date_col="apply_dt",
        features=["x"],
        config=_config("segment_shift"),
    )

    assert isinstance(increase_result, MarsMissingShiftResult)
    increase_row = increase_result.detail_table.row(0, named=True)
    assert increase_row["anomaly_type"] == "segment_shift"
    assert increase_row["direction"] == "increase"
    assert increase_row["abs_delta"] >= 0.25
    assert increase_row["q_value"] <= 0.05
    assert increase_result.summary_table.get_column("feature").to_list() == ["x"]
    assert increase_result.source_table.get_column("data_source").to_list() == ["bureau"]
    assert "decrease" in decrease_result.detail_table.get_column("direction").to_list()


def test_boundary_detector_locates_first_day_spike_exactly() -> None:
    df = _daily_missing_frame(rates=[0.80] + [0.10] * 8)

    result = MarsMissingShiftScanner().scan(
        df,
        date_col="apply_dt",
        features=["x"],
        config=_config("boundary"),
    )

    row = result.detail_table.filter(pl.col("direction") == "increase").row(0, named=True)
    assert row["event_start_period"] == "2026-01-01"
    assert row["event_end_period"] == "2026-01-01"
    assert row["reference_type"] == "next_window"


def test_boundary_detector_merges_three_day_start_run() -> None:
    df = _daily_missing_frame(rates=[0.80] * 3 + [0.10] * 6)

    result = MarsMissingShiftScanner().scan(
        df,
        date_col="apply_dt",
        features=["x"],
        config=_config("boundary"),
    )

    row = result.detail_table.filter(pl.col("direction") == "increase").row(0, named=True)
    assert row["event_start_period"] == "2026-01-01"
    assert row["event_end_period"] == "2026-01-03"


def test_boundary_detector_uses_explicit_benchmark_for_start_and_end() -> None:
    current = _daily_missing_frame(rates=[0.70, 0.10, 0.10, 0.10, 0.70])
    benchmark = _daily_missing_frame(rates=[0.10] * 3).drop("apply_dt")

    result = MarsMissingShiftScanner().scan(
        current,
        date_col="apply_dt",
        features=["x"],
        benchmark_df=benchmark,
        config=_config("boundary", max_boundary_periods=1),
    )

    assert result.detail_table.height == 2
    assert set(result.detail_table.get_column("reference_type")) == {"benchmark"}
    assert set(result.detail_table.get_column("event_start_period")) == {
        "2026-01-01",
        "2026-01-05",
    }


def test_point_detector_only_flags_isolated_internal_day() -> None:
    df = _daily_missing_frame(rates=[0.10] * 6 + [0.80] + [0.10] * 6)

    result = MarsMissingShiftScanner().scan(
        df,
        date_col="apply_dt",
        features=["x"],
        config=_config("point"),
    )

    assert result.detail_table.height == 1
    row = result.detail_table.row(0, named=True)
    assert row["anomaly_type"] == "point"
    assert row["event_start_period"] == row["event_end_period"] == "2026-01-07"
    assert row["reference_type"] == "surrounding_window"


def test_high_level_detector_flags_stable_high_series_and_feature_override() -> None:
    high_df = _daily_missing_frame(rates=[0.95] * 5)
    overridden_df = _daily_missing_frame(rates=[0.85] * 5)

    high_result = MarsMissingShiftScanner().scan(
        high_df,
        date_col="apply_dt",
        features=["x"],
        config=_config("high_level"),
    )
    override_result = MarsMissingShiftScanner().scan(
        overridden_df,
        date_col="apply_dt",
        features=["x"],
        config=_config(
            "high_level",
            feature_high_missing_rate_thresholds={"x": 0.80},
        ),
    )

    for result in (high_result, override_result):
        row = result.detail_table.row(0, named=True)
        assert row["anomaly_type"] == "high_level"
        assert row["p_value"] is None
        assert row["q_value"] is None
        assert row["reference_missing_rate"] is None


def test_overlapping_segment_and_high_level_candidates_merge_once() -> None:
    df = _daily_missing_frame(rates=[0.10] * 4 + [0.95] * 6)

    result = MarsMissingShiftScanner().scan(
        df,
        date_col="apply_dt",
        features=["x"],
        config=_config("segment_shift", "high_level"),
    )

    assert result.detail_table.height == 1
    row = result.detail_table.row(0, named=True)
    assert row["anomaly_type"] == "high_level"
    assert set(row["detected_by"].split(",")) == {"high_level", "segment_shift"}
    assert result.summary_table.get_column("anomaly_count").to_list() == [1]


def test_low_sample_day_is_visible_but_splits_detection_windows() -> None:
    df = _daily_missing_frame(
        rates=[0.10, 0.10, 0.10, 0.80, 0.80, 0.80, 0.80],
        rows_per_day=[100, 100, 100, 10, 100, 100, 100],
    )

    result = MarsMissingShiftScanner().scan(
        df,
        date_col="apply_dt",
        features=["x"],
        config=_config("segment_shift"),
    )

    excluded = result.trend_table.filter(~pl.col("is_detection_eligible"))
    assert excluded.get_column("period").to_list() == ["2026-01-04"]
    assert excluded.get_column("exclusion_reason").to_list() == [
        "period_sample_count_below_minimum"
    ]
    assert result.detail_table.is_empty()


def test_all_low_sample_periods_return_empty_events_with_diagnostics() -> None:
    df = _daily_missing_frame(rates=[0.95] * 4, rows_per_day=10)

    result = MarsMissingShiftScanner().scan(
        df,
        date_col="apply_dt",
        features=["x"],
    )

    assert result.detail_table.is_empty()
    assert result.summary_table.is_empty()
    assert not result.trend_table.get_column("is_detection_eligible").any()


def test_statistical_test_switches_between_fisher_and_z_test() -> None:
    fisher = _two_proportion_test(
        _WindowStats(missing_count=0, total_count=30),
        _WindowStats(missing_count=5, total_count=30),
    )
    z_test = _two_proportion_test(
        _WindowStats(missing_count=30, total_count=100),
        _WindowStats(missing_count=10, total_count=100),
    )

    assert fisher[1] == "fisher_exact"
    assert z_test[1] == "z_test"


def test_benjamini_hochberg_adjusts_all_candidates_globally() -> None:
    base = {
        "feature": "x",
        "data_source": "UNMAPPED",
        "anomaly_type": "point",
        "event_start_idx": 0,
        "event_end_idx": 0,
        "event_stats": _WindowStats(20, 100),
        "reference_type": "surrounding_window",
        "reference_start_period": "2026-01-01",
        "reference_end_period": "2026-01-02",
        "reference_stats": _WindowStats(10, 100),
        "threshold": None,
        "delta": 0.1,
        "abs_delta": 0.1,
        "relative_delta": 1.0,
        "q_value": None,
        "test_method": "z_test",
    }
    candidates = [
        _Candidate(**base, p_value=p_value)  # type: ignore[arg-type]
        for p_value in [0.01, 0.04, 0.20]
    ]

    adjusted = _apply_benjamini_hochberg(candidates)

    assert [candidate.q_value for candidate in adjusted] == pytest.approx([0.03, 0.06, 0.20])


def test_missing_shift_scanner_validates_config_and_benchmark_contracts() -> None:
    df = _daily_missing_frame(rates=[0.10] * 4)
    scanner = MarsMissingShiftScanner()

    with pytest.raises(ValueError, match="Date column"):
        scanner.scan(df, date_col="missing_dt", features=["x"])
    with pytest.raises(ValueError, match="Features not found"):
        scanner.scan(df, date_col="apply_dt", features=["missing_feature"])
    with pytest.raises(ValueError, match="min_segment_size"):
        scanner.scan(
            df,
            date_col="apply_dt",
            features=["x"],
            config=_config(min_segment_size=1),
        )
    with pytest.raises(ValueError, match="threshold features not found"):
        scanner.scan(
            df,
            date_col="apply_dt",
            features=["x"],
            config=_config(feature_high_missing_rate_thresholds={"unknown": 0.5}),
        )
    with pytest.raises(ValueError, match="must not be empty"):
        scanner.scan(
            df,
            date_col="apply_dt",
            features=["x"],
            benchmark_df=pl.DataFrame(schema={"x": pl.Float64}),
        )
    with pytest.raises(ValueError, match="at least 30 rows"):
        scanner.scan(
            df,
            date_col="apply_dt",
            features=["x"],
            benchmark_df=pl.DataFrame({"x": [1.0] * 5}),
        )


def test_result_tables_notebook_views_plot_and_excel(tmp_path: Path) -> None:
    df = _daily_missing_frame(rates=[0.10] * 6 + [0.80] + [0.10] * 6)
    result = MarsMissingShiftScanner().scan(
        df,
        date_col="apply_dt",
        features=["x"],
        feature_data_source={"x": "application"},
        config=_config("point"),
    )

    assert result.trend_table.columns == [
        "feature",
        "dtype",
        "data_source",
        "period",
        "missing_count",
        "total_count",
        "missing_rate",
        "is_detection_eligible",
        "exclusion_reason",
        "is_anomaly",
        "anomaly_types",
    ]
    assert hasattr(result.show_summary(), "to_html")
    assert hasattr(result.show_detail(anomaly_types="point"), "to_html")
    assert hasattr(result.show_trend(), "to_html")
    figure = result.plot_trends()
    assert len(figure.axes) == 1
    plt.close(figure)

    output_path = tmp_path / "missing_shift.xlsx"
    result.write_excel(output_path)
    assert output_path.exists()
    workbook = load_workbook(output_path, read_only=True)
    assert set(workbook.sheetnames) == {"summary", "detail", "source", "trend"}
    workbook.close()


def test_wide_scan_uses_explicit_anomaly_annotation_schema() -> None:
    base = _daily_missing_frame(rates=[0.10] * 6 + [0.80] + [0.10] * 6)
    wide = base.with_columns(
        [
            pl.col("x").alias(f"feature_{feature_idx:03d}")
            for feature_idx in range(90)
        ]
    )
    target_feature = "feature_089"

    result = MarsMissingShiftScanner().scan(
        wide,
        date_col="apply_dt",
        features=[f"feature_{feature_idx:03d}" for feature_idx in range(90)],
        config=_config("point"),
    )

    target_trend = result.trend_table.filter(pl.col("feature") == target_feature)
    assert target_trend.height == 13
    assert target_trend.get_column("anomaly_types").drop_nulls().to_list() == ["point"]
