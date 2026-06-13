import warnings
from pathlib import Path
from typing import Any

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import polars as pl
import pytest

from mars.analysis import MarsBinEvaluator, MarsBinningReport, MarsRiskProfile, profile_risk
from mars.feature import MarsNativeBinner
from mars.reporting.plotter import MarsPlotter


def _as_pandas(df):
    return df.to_pandas() if isinstance(df, pl.DataFrame) else df.copy()


def _profile_risk_report(*args, **kwargs):
    run = profile_risk(*args, **kwargs)
    return run.report, run


def _make_exact_start_aware_monthly_df() -> pl.DataFrame:
    rows = []
    for day in pd.date_range("2018-01-01", periods=100, freq="D"):
        if day < pd.Timestamp("2018-02-15"):
            for _ in range(40):
                rows.append({"dt": day, "EXT_SOURCE_1": None, "target": 0})
        else:
            for _ in range(20):
                rows.append({"dt": day, "EXT_SOURCE_1": 0.2, "target": 0})
            for _ in range(20):
                rows.append({"dt": day, "EXT_SOURCE_1": 0.8, "target": 1})
    return pl.DataFrame(
        rows,
        schema={
            "dt": pl.Datetime,
            "EXT_SOURCE_1": pl.Float64,
            "target": pl.Int64,
        },
    )


def test_profile_risk_returns_structured_run(sample_credit_df):
    run = profile_risk(
        sample_credit_df,
        target="target",
        features=["income", "utilization"],
        group_col="month",
        binner_params={"method": "quantile", "n_bins": 3},
    )

    assert isinstance(run, MarsRiskProfile)
    assert isinstance(run.binner, MarsNativeBinner)
    assert run.targets == ["target"]
    assert run.report.summary_table.height == 2


def test_profile_risk_no_longer_accepts_plot_argument(
    sample_credit_df: pl.DataFrame,
) -> None:
    with pytest.raises(TypeError, match="plot"):
        profile_risk(  # type: ignore[call-arg]
            sample_credit_df,
            target="target",
            features=["income"],
            plot=False,
        )


def test_binning_report_plot_risk_trends_uses_report_as_public_entry(
    sample_credit_df: pl.DataFrame,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    run = profile_risk(
        sample_credit_df,
        target="target",
        features=["income", "utilization"],
        group_col="month",
        binner_params={"method": "quantile", "n_bins": 3},
    )

    captured: dict[str, Any] = {}

    def _fake_batch(
        df_detail: pd.DataFrame | pl.DataFrame,
        features: list[str],
        group_col: str = "month",
        target_name: str = "Target",
        dpi: int = 150,
        sort_by: str = "iv",
        ascending: bool = False,
    ) -> None:
        captured["df_detail"] = df_detail
        captured["features"] = features
        captured["group_col"] = group_col
        captured["target_name"] = target_name
        captured["dpi"] = dpi
        captured["sort_by"] = sort_by
        captured["ascending"] = ascending

    monkeypatch.setattr(
        MarsPlotter,
        "plot_feature_binning_risk_trend_batch",
        staticmethod(_fake_batch),
    )

    run.report.plot_risk_trends(features="income", dpi=90)

    assert run.report.group_col == "month"
    assert run.report.detail_group_col == "mars_group"
    assert captured["features"] == ["income"]
    assert captured["group_col"] == "mars_group"
    assert captured["target_name"] == "target"
    assert captured["dpi"] == 90
    assert captured["sort_by"] == ""


def test_binning_report_plot_risk_trends_supports_multi_target_filter(
    sample_credit_df: pl.DataFrame,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    df = sample_credit_df.with_columns(
        (pl.col("utilization") >= 0.45).cast(pl.Int8).alias("target_alt")
    )
    run = profile_risk(
        df,
        target=["target", "target_alt"],
        features=["income", "utilization"],
        group_col="month",
        binner_params={"method": "quantile", "n_bins": 3},
    )

    calls: list[dict[str, Any]] = []

    def _fake_batch(
        df_detail: pd.DataFrame | pl.DataFrame,
        features: list[str],
        group_col: str = "month",
        target_name: str = "Target",
        dpi: int = 150,
        sort_by: str = "iv",
        ascending: bool = False,
    ) -> None:
        calls.append(
            {
                "df_detail": df_detail,
                "features": features,
                "group_col": group_col,
                "target_name": target_name,
                "dpi": dpi,
                "sort_by": sort_by,
                "ascending": ascending,
            }
        )

    monkeypatch.setattr(
        MarsPlotter,
        "plot_feature_binning_risk_trend_batch",
        staticmethod(_fake_batch),
    )

    run.report.plot_risk_trends(target="target_alt", max_plots=1)

    assert run.report.group_col == "month"
    assert run.report.detail_group_col == "mars_group"
    assert len(calls) == 1
    assert calls[0]["target_name"] == "target_alt"
    assert calls[0]["group_col"] == "mars_group"
    assert calls[0]["sort_by"] == ""
    assert len(calls[0]["features"]) == 1
    assert set(calls[0]["df_detail"]["y"].astype(str).tolist()) == {"target_alt"}


def test_binning_plot_figure_keeps_total_panel_on_the_right(
    sample_credit_df: pl.DataFrame,
) -> None:
    run = profile_risk(
        sample_credit_df,
        target="target",
        features=["income"],
        group_col="month",
        binner_params={"method": "quantile", "n_bins": 3},
    )

    fig = MarsPlotter._build_feature_binning_risk_figure(
        df_detail=run.report.detail_table,
        feature="income",
        group_col="mars_group",
        target_name="target",
    )

    assert fig is not None
    titled_axes = [ax for ax in fig.axes if ax.get_title()]
    assert titled_axes
    assert titled_axes[-1].get_title().startswith("Total")
    plt.close(fig)


def test_binning_plot_summary_header_uses_total_panel_metric_scope(
    sample_credit_df: pl.DataFrame,
) -> None:
    run = profile_risk(
        sample_credit_df,
        target="target",
        features=["income"],
        group_col="month",
        binner_params={"method": "quantile", "n_bins": 3},
    )

    fig = MarsPlotter._build_feature_binning_risk_figure(
        df_detail=run.report.detail_table,
        feature="income",
        group_col="mars_group",
        target_name="target",
    )

    assert fig is not None
    detail_pd = _as_pandas(run.report.detail_table)
    total_panel_df = detail_pd[
        (detail_pd["feature"].astype(str) == "income")
        & (detail_pd["mars_group"].astype(str) == "Total")
        & (detail_pd["bin_label"].astype(str) != "Total")
        & (detail_pd["bin_type"].astype(str) != "汇总组")
    ].copy()
    expected_iv, expected_ks, expected_auc = MarsPlotter._summarize_binning_metrics(
        total_panel_df,
    )
    summary_text = fig.texts[0].get_text()

    assert f"IV: {expected_iv:.3f}" in summary_text
    assert f"KS: {expected_ks:.1f}" in summary_text
    assert f"AUC: {expected_auc:.2f}" in summary_text
    plt.close(fig)


def test_binning_type_opt_alias_is_removed(sample_credit_df):
    with pytest.raises(ValueError, match="optimal"):
        MarsBinEvaluator(binning_type="opt")  # type: ignore[arg-type]

    with pytest.raises(ValueError, match="optimal"):
        profile_risk(
            sample_credit_df,
            target="target",
            features=["income"],
            binning_type="opt",  # type: ignore[arg-type]
        )


@pytest.mark.parametrize("time_grain", ["1d", "1w", "2w", "1m", "2m"])
def test_profile_risk_time_grain_aliases_do_not_warn(time_grain, caplog):
    dates = pd.date_range("2024-01-01", periods=90, freq="D")
    x = np.linspace(0.0, 1.0, len(dates))
    df = pl.DataFrame(
        {
            "apply_dt": dates.strftime("%Y-%m-%d").to_list(),
            "x": x,
            "target": (x > 0.5).astype(int),
        }
    )

    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        with caplog.at_level("WARNING", logger="mars"):
            run = profile_risk(
                df,
                target="target",
                features=["x"],
                time_col="apply_dt",
                time_grain=time_grain,
                binning_type="native",
                binner_params={"method": "quantile", "n_bins": 3},
            )

    assert run.report.report_meta["profile_by_input"] == time_grain
    assert run.report.trend_tables["psi"].height == 1
    assert not caplog.messages
    assert caught == []


def test_profile_risk_returns_report_and_evaluator_for_pandas_input(sample_credit_pd):
    report, evaluator = _profile_risk_report(
        sample_credit_pd,
        target="target",
        features=["income", "utilization"],
        group_col="month",
        binning_type="native",
        binner_params={"method": "quantile", "n_bins": 3},
    )

    assert isinstance(report.summary_table, pd.DataFrame)
    assert isinstance(evaluator.binner, MarsNativeBinner)
    assert set(report.summary_table["feature"]) == {"income", "utilization"}


def test_profile_risk_summary_is_consistent_between_polars_and_pandas(sample_credit_df, sample_credit_pd):
    report_pl, _ = _profile_risk_report(
        sample_credit_df,
        target="target",
        features=["income", "utilization"],
        group_col="month",
        binning_type="native",
        binner_params={"method": "quantile", "n_bins": 3},
    )
    report_pd, _ = _profile_risk_report(
        sample_credit_pd,
        target="target",
        features=["income", "utilization"],
        group_col="month",
        binning_type="native",
        binner_params={"method": "quantile", "n_bins": 3},
    )

    summary_pl = report_pl.summary_table.to_pandas().sort_values("feature").reset_index(drop=True)
    summary_pd = report_pd.summary_table.sort_values("feature").reset_index(drop=True)

    pd.testing.assert_frame_equal(summary_pl, summary_pd, check_dtype=False, check_like=False)


def test_profile_risk_multi_target_keeps_pandas_return_type(sample_credit_pd):
    df = sample_credit_pd.copy()
    df["target_alt"] = (df["utilization"] >= 0.45).astype(int)

    report, evaluator = _profile_risk_report(
        df,
        target=["target", "target_alt"],
        features=["income", "utilization"],
        group_col="month",
        binning_type="native",
        binner_params={"method": "quantile", "n_bins": 3},
    )

    assert isinstance(report.summary_table, pd.DataFrame)
    assert isinstance(report.detail_table, pd.DataFrame)
    assert set(report.summary_table["target"]) == {"target", "target_alt"}
    assert isinstance(evaluator.binner, MarsNativeBinner)


def test_profile_risk_without_target_returns_distribution_only_metrics(sample_credit_df):
    monitor_df = sample_credit_df.select(["month", "income", "utilization", "segment"])

    report, run = _profile_risk_report(
        monitor_df,
        target=None,
        features=["income", "utilization", "segment"],
        group_col="month",
        binning_type="native",
        binner_params={"method": "quantile", "n_bins": 3},
    )

    summary = report.summary_table

    assert run.targets == []
    assert report.report_meta["targets"] == []
    assert "psi" in report.trend_tables
    assert summary.select(pl.col("iv").is_null().all()).item()


def test_evaluator_label_free_mode_does_not_mutate_target_state(sample_credit_df):
    monitor_df = sample_credit_df.select(["month", "income", "utilization", "segment"])
    evaluator = MarsBinEvaluator(binner_params={"method": "quantile", "n_bins": 3})

    run = evaluator.evaluate(
        monitor_df,
        features=["income", "utilization"],
        group_col="month",
    )
    report = run.report

    assert evaluator.target is None
    assert evaluator.has_target_ is False
    assert report.summary_table.select(pl.col("iv").is_null().all()).item()


def test_evaluator_does_not_reuse_fitted_binner_between_calls(sample_credit_df):
    evaluator = MarsBinEvaluator(binner_params={"method": "quantile", "n_bins": 3})

    first = evaluator.evaluate(
        sample_credit_df,
        target="target",
        features=["income"],
        group_col="month",
    )
    second = evaluator.evaluate(
        sample_credit_df,
        target="target",
        features=["utilization"],
        group_col="month",
    )

    assert first.binner is not second.binner
    assert first.binner.features == ["income"]
    assert second.binner.features == ["utilization"]
    assert evaluator.binner is None


def test_evaluator_reuses_only_explicit_binner(sample_credit_df):
    binner = MarsNativeBinner(method="quantile", n_bins=3)
    binner.fit(
        sample_credit_df.select(["income"]),
        sample_credit_df.get_column("target"),
        features=["income"],
    )
    evaluator = MarsBinEvaluator()

    run = evaluator.evaluate(
        sample_credit_df,
        target="target",
        features=["income"],
        binner=binner,
        group_col="month",
    )

    assert run.binner is binner


def test_profile_risk_rejects_binner_and_binner_params_together(sample_credit_df):
    binner = MarsNativeBinner(method="quantile", n_bins=3)

    with pytest.raises(ValueError, match="binner_params"):
        profile_risk(
            sample_credit_df,
            target="target",
            features=["income"],
            binner=binner,
            binner_params={"method": "quantile"},
        )


def test_evaluator_external_benchmark_skips_missing_benchmark_columns(sample_credit_df):
    benchmark_df = sample_credit_df.select(["month", "income", "target"])

    report, _ = _profile_risk_report(
        sample_credit_df,
        target="target",
        features=["income", "utilization"],
        group_col="month",
        benchmark_df=benchmark_df,
        binning_type="native",
        binner_params={"method": "quantile", "n_bins": 3},
    )

    assert set(report.summary_table["feature"].to_list()) == {"income", "utilization"}


def test_evaluation_report_can_write_excel(sample_credit_df, caplog):
    report, _ = _profile_risk_report(
        sample_credit_df,
        target="target",
        features=["income"],
        group_col="month",
        binning_type="native",
        binner_params={"method": "quantile", "n_bins": 3},
    )

    artifacts_dir = Path(__file__).resolve().parent / "_artifacts"
    artifacts_dir.mkdir(exist_ok=True)
    output_path = artifacts_dir / "evaluation_report.xlsx"
    if output_path.exists():
        output_path.unlink()

    try:
        with caplog.at_level("INFO", logger="mars"):
            report.write_excel(str(output_path), engine="openpyxl")
        assert output_path.exists()
        assert any("导出成功" in message or "Exported" in message for message in caplog.messages)
    finally:
        if output_path.exists():
            output_path.unlink()
        if artifacts_dir.exists() and not any(artifacts_dir.iterdir()):
            artifacts_dir.rmdir()


def test_evaluation_report_excel_contains_detail_sheet_and_data_source(sample_credit_df, tmp_path):
    from openpyxl import load_workbook
    from openpyxl.utils.cell import range_boundaries

    report, _ = _profile_risk_report(
        sample_credit_df,
        target="target",
        features=["income", "utilization"],
        feature_data_source={
            "EXT_SOURCE_1": ["income"],
            "EXT_SOURCE_2": ["utilization"],
        },
        group_col="month",
        binning_type="native",
        binner_params={"method": "quantile", "n_bins": 3},
    )
    output_path = tmp_path / "evaluation_report.xlsx"

    report.write_excel(str(output_path), engine="openpyxl")

    workbook = load_workbook(output_path)
    try:
        assert "分组明细" in workbook.sheetnames
        worksheet = workbook["分组明细"]
        headers = [
            worksheet.cell(row=1, column=col_idx).value
            for col_idx in range(1, worksheet.max_column + 1)
        ]
        header_index = {
            str(header): col_idx + 1
            for col_idx, header in enumerate(headers)
            if header is not None
        }
        assert {"feature", "data_source", "bin_label", "count"}.issubset(header_index)

        rows = [
            {
                column_name: worksheet.cell(row=row_idx, column=col_idx).value
                for column_name, col_idx in header_index.items()
            }
            for row_idx in range(4, worksheet.max_row + 1)
        ]
        data_rows = [row for row in rows if row.get("feature")]
        feature_values = {row["feature"] for row in data_rows}
        data_sources = {row["data_source"] for row in data_rows}

        assert {"income", "utilization"}.issubset(feature_values)
        assert {"EXT_SOURCE_1", "EXT_SOURCE_2"}.issubset(data_sources)
        assert worksheet.tables

        table_ref = next(iter(worksheet.tables.values())).ref
        _, min_row, max_col, max_row = range_boundaries(table_ref)
        assert min_row == 1
        assert max_col >= header_index["data_source"]
        assert max_row >= worksheet.max_row
    finally:
        workbook.close()


def test_evaluation_report_can_write_html(sample_credit_df, caplog):
    report, _ = _profile_risk_report(
        sample_credit_df,
        target="target",
        features=["income", "utilization"],
        group_col="month",
        binning_type="native",
        binner_params={"method": "quantile", "n_bins": 3},
    )

    artifacts_dir = Path(__file__).resolve().parent / "_artifacts"
    artifacts_dir.mkdir(exist_ok=True)
    output_path = artifacts_dir / "evaluation_report.html"
    if output_path.exists():
        output_path.unlink()

    try:
        with caplog.at_level("INFO", logger="mars"):
            report.write_html(str(output_path), report_name="中文风控监控报告", sort_by="not_a_real_metric", max_plots=1)
        assert output_path.exists()
        html_text = output_path.read_text(encoding="utf-8")
        assert "<title>中文风控监控报告</title>" in html_text
        assert "中文风控监控报告" in html_text
        assert "Summary" in html_text
        assert "Dataset Overview" in html_text
        assert "Trend Tables" in html_text
        assert "Grouped Pivot" in html_text
        assert "Charts" in html_text
        assert "Regex Mode" in html_text
        assert "Global search across tables and charts" in html_text
        assert "Export Feature List" in html_text
        assert "Jump to Feature" in html_text
        assert "Jumps to the matching row in Summary." in html_text
        assert "marsJumpToFeature()" in html_text
        assert "marsOpenAncestorSections" in html_text
        assert "marsClearJumpHighlight" in html_text
        assert "marsActivateJumpHighlight" in html_text
        assert 'Feature "${value}" does not exist in Summary.' in html_text
        assert 'Feature "${value}" is hidden by data source, global search, or summary filter.' in html_text
        assert 'id="mars-page-top"' in html_text
        assert 'id="mars-floating-header-host"' in html_text
        assert 'id="mars-floating-header-scroll"' in html_text
        assert 'id="mars-back-to-top"' in html_text
        assert 'onclick="marsBackToTop()"' in html_text
        assert "mars-table-ownership-sentinel" in html_text
        assert 'data-sentinel-role="start"' in html_text
        assert 'data-sentinel-role="end"' in html_text
        assert "mars-resize-handle" in html_text
        assert "--mars-feature-col-width: 220px" in html_text
        assert "mars-secondary-col" in html_text
        assert 'data-table-id="mars-summary-table"' in html_text
        assert "mars-summary-table-query" in html_text
        assert "Feature Filter Expression" in html_text
        assert "marsSetSummaryExpression(this.value)" in html_text
        assert "Supported operators: &gt;, &gt;=, &lt;, &lt;=, ==, !=, &amp;, |, ( ). Use &amp; for AND and | for OR." in html_text
        assert "MISSING" in html_text
        assert "LIFT" in html_text
        assert "Lift Min" in html_text
        assert "purple at or below 0.5" in html_text
        assert "<strong>Lift Max</strong> 1.2 / 1.3 / 1.4, purple above 1.5" in html_text
        assert "<strong>Lift</strong> 1.2 / 1.3 / 1.4, purple above 1.5" in html_text
        assert "Search chart features..." in html_text
        assert "Search grouped pivot..." in html_text
        assert 'class="mars-source-checkbox"' in html_text
        assert 'function marsQueueRefresh(scopeToken="all", delayMs=0)' in html_text
        assert 'function marsQueueTextRefresh(scopeToken="all")' in html_text
        assert "marsResolveLocalScope" in html_text
        assert 'marsQueueRefresh("table:mars-summary-table")' in html_text
        assert "marsRefreshFloatingHeader" in html_text
        assert "marsCloneFloatingHeader" in html_text
        assert "marsGetFirstVisibleDataRowTop" in html_text
        assert "marsAncestorsDetailsOpen" in html_text
        assert "marsTableIsActuallyVisible" in html_text
        assert "marsCollectLeafColumnWidths" in html_text
        assert "marsBuildFloatingHeaderColGroup" in html_text
        assert "marsRegisterTableScrollListeners" in html_text
        assert "marsResolveFloatingHeaderOwner" in html_text
        assert "marsUpdateBackToTopVisibility" in html_text
        assert 'window.addEventListener("scroll", marsScheduleViewportRefresh' in html_text
        assert 'document.addEventListener("toggle", () => { marsHideFloatingHeader(); marsQueueLayoutSync("all"); marsScheduleViewportRefresh(); }, true);' in html_text
        assert "jumpHighlightTimerId" in html_text
        assert "jumpHighlightArmTimerId" in html_text
        assert "3000" in html_text
        assert "marsState.jumpHighlightArmTimerId=window.setTimeout" in html_text
        assert "const readingLine=Math.max(1, hostHeight || owner.headerHeight) + 1;" in html_text
        assert "item.theadTop > 0 && item.firstDataRowTop <= readingLine" in html_text
        assert "if(!hasVisibleReadingTable) return null;" in html_text
        assert "marsGetTableOwnershipSentinels" not in html_text
        assert "startRect.top > 0 || endRect.top <= headerHeight" not in html_text
        assert "rect.top <= 0 && rect.bottom > headerHeight" not in html_text
        assert "mars-chart-cards-status" in html_text
        assert "mars-result-status" in html_text
        assert "marsSelectAllSources()" in html_text
        assert "window.requestAnimationFrame" in html_text
        assert 'window.addEventListener("resize"' in html_text
        assert "JSON.stringify(featureMap, null, 2)" in html_text
        assert "let cursor=0;" in html_text
        assert "tokenPattern.lastIndex = cursor" in html_text
        assert "cursor < text.length" in html_text
        assert "lastIndex!==text.length" not in html_text
        assert '["identifier", "compare", "and", "or"].includes(left.type)' in html_text
        assert "Export uses Summary expression + Data Source only." in html_text
        assert "Global and local searches only affect display." in html_text
        assert "marsCollectVisibleFeatures" not in html_text
        assert "mars_features.txt" in html_text
        assert "activeTableId" not in html_text
        assert "detail-section" not in html_text
        assert "mars-trend-bad-rate" not in html_text
        assert 'id="missing-day-section"' not in html_text
        assert "Bin Type" not in html_text
        assert "Threshold Filter (Total)" not in html_text
        assert ">mono<" not in html_text
        assert "Event Rate" in html_text
        assert "Binned distribution and risk comparison across groups." in html_text
        assert "Grouped pivot aligned with the Excel-style source + bin matrix." not in html_text
        assert html_text.index("mars-trend-missing") < html_text.index("mars-trend-psi")
        assert html_text.index("mars-trend-psi") < html_text.index("mars-trend-iv")
        assert html_text.index("mars-trend-iv") < html_text.index("mars-trend-ks")
        assert html_text.index("mars-trend-ks") < html_text.index("mars-trend-lift")
        assert html_text.index("mars-trend-lift") < html_text.index("mars-trend-auc")
        assert html_text.index("mars-trend-auc") < html_text.index("mars-trend-risk-corr")
        assert any("HTML" in message or "html" in message for message in caplog.messages)
    finally:
        if output_path.exists():
            output_path.unlink()
        if artifacts_dir.exists() and not any(artifacts_dir.iterdir()):
            artifacts_dir.rmdir()


def test_evaluation_report_produces_missing_and_lift_trend_tables(sample_credit_df):
    report, _ = _profile_risk_report(
        sample_credit_df,
        target="target",
        features=["income", "utilization"],
        group_col="month",
        binning_type="native",
        binner_params={"method": "quantile", "n_bins": 3},
    )

    assert "missing" in report.trend_tables
    assert "lift" in report.trend_tables

    missing_df = report.trend_tables["missing"]
    lift_df = report.trend_tables["lift"]
    if isinstance(missing_df, pl.DataFrame):
        missing_df = missing_df.to_pandas()
    if isinstance(lift_df, pl.DataFrame):
        lift_df = lift_df.to_pandas()

    assert "Total" in missing_df.columns
    assert "Total" in lift_df.columns
    assert missing_df["feature"].isin(["income", "utilization"]).all()


def test_multi_target_html_includes_target_switchers(sample_credit_df, caplog):
    df = sample_credit_df.with_columns(
        (pl.col("utilization") >= 0.45).cast(pl.Int8).alias("target_alt")
    )

    report, _ = _profile_risk_report(
        df,
        target=["target", "target_alt"],
        features=["income", "utilization"],
        group_col="month",
        binning_type="native",
        binner_params={"method": "quantile", "n_bins": 3},
    )

    artifacts_dir = Path(__file__).resolve().parent / "_artifacts"
    artifacts_dir.mkdir(exist_ok=True)
    output_path = artifacts_dir / "evaluation_report_multi.html"
    if output_path.exists():
        output_path.unlink()

    try:
        with caplog.at_level("INFO", logger="mars"):
            report.write_html(str(output_path), max_plots=1)
        html_text = output_path.read_text(encoding="utf-8")
        assert "Pivot Target" in html_text
        assert "Chart Target" in html_text
        assert "target_alt" in html_text
    finally:
        if output_path.exists():
            output_path.unlink()
        if artifacts_dir.exists() and not any(artifacts_dir.iterdir()):
            artifacts_dir.rmdir()


def test_show_summary_uses_pandas_view_without_mutating_polars_report(sample_credit_df):
    report, _ = _profile_risk_report(
        sample_credit_df,
        target="target",
        features=["income", "utilization"],
        group_col="month",
        binning_type="native",
        binner_params={"method": "quantile", "n_bins": 3},
    )

    styler = report.show_summary(features=["income"])

    assert isinstance(styler.data, pd.DataFrame)
    assert isinstance(report.summary_table, pl.DataFrame)
    assert set(report.summary_table["feature"].to_list()) == {"income", "utilization"}


def test_feature_data_source_is_attached_to_report_outputs(sample_credit_df):
    report, _ = _profile_risk_report(
        sample_credit_df,
        target="target",
        features=["income", "utilization"],
        feature_data_source={"APP": ["income"]},
        group_col="month",
        binning_type="native",
        binner_params={"method": "quantile", "n_bins": 3},
    )

    summary_map = {
        row["feature"]: row["data_source"]
        for row in report.summary_table.select(["feature", "data_source"]).to_dicts()
    }
    detail_map = {
        row["feature"]: row["data_source"]
        for row in report.detail_table.select(["feature", "data_source"]).unique().to_dicts()
    }

    assert summary_map["income"] == "APP"
    assert summary_map["utilization"] == "UNMAPPED"
    assert detail_map["income"] == "APP"
    assert detail_map["utilization"] == "UNMAPPED"


def test_feature_data_source_rejects_features_outside_active_feature_set(sample_credit_df):
    evaluator = MarsBinEvaluator(binner_params={"method": "quantile", "n_bins": 3})

    with pytest.raises(ValueError, match="feature_data_source"):
        evaluator.evaluate(
            sample_credit_df,
            target="target",
            features=["income", "utilization"],
            feature_data_source={"BAD": ["age"]},
            group_col="month",
        )


def test_evaluator_generates_missing_by_day_table_when_dt_col_is_provided(sample_credit_df):
    df = sample_credit_df.with_columns(
        pl.Series("biz_dt", pd.date_range("2024-01-01", periods=sample_credit_df.height, freq="D"))
    )

    report, _ = _profile_risk_report(
        df,
        target="target",
        features=["income", "utilization"],
        group_col="month",
        time_col="biz_dt",
        binning_type="native",
        binner_params={"method": "quantile", "n_bins": 3},
    )

    assert report.dt_col == "biz_dt"
    assert report.missing_by_day_table is not None
    missing_by_day = report.missing_by_day_table
    if isinstance(missing_by_day, pd.DataFrame):
        assert "feature" in missing_by_day.columns
    else:
        assert "feature" in missing_by_day.columns


def test_summary_table_includes_missing_and_lift_monitor_columns(sample_credit_df):
    report, _ = _profile_risk_report(
        sample_credit_df,
        target="target",
        features=["income", "utilization"],
        group_col="month",
        binning_type="native",
        binner_params={"method": "quantile", "n_bins": 3},
    )

    summary_pd = report.summary_table.to_pandas() if isinstance(report.summary_table, pl.DataFrame) else report.summary_table.copy()
    assert {"lift_min", "lift_max", "missing", "missing_min", "missing_max", "mono"}.issubset(summary_pd.columns)

    income_row = summary_pd.loc[summary_pd["feature"] == "income"].iloc[0]
    assert income_row["missing"] == pytest.approx(0.125, rel=1e-6)
    assert income_row["missing_min"] == pytest.approx(0.125, rel=1e-6)
    assert income_row["missing_max"] == pytest.approx(0.125, rel=1e-6)
    assert income_row["lift_min"] == pytest.approx(0.7272727, rel=1e-6)
    assert income_row["lift_max"] == pytest.approx(1.3636363, rel=1e-6)


def test_evaluation_report_html_includes_missing_by_day_and_data_source_filter(sample_credit_df, caplog):
    df = sample_credit_df.with_columns(
        pl.Series("biz_dt", pd.date_range("2024-01-01", periods=sample_credit_df.height, freq="D"))
    )

    report, _ = _profile_risk_report(
        df,
        target="target",
        features=["income", "utilization"],
        feature_data_source={"EXT_SOURCE_1": ["income"], "EXT_SOURCE_2": ["utilization"]},
        group_col="month",
        time_col="biz_dt",
        binning_type="native",
        binner_params={"method": "quantile", "n_bins": 3},
    )

    artifacts_dir = Path(__file__).resolve().parent / "_artifacts"
    artifacts_dir.mkdir(exist_ok=True)
    output_path = artifacts_dir / "evaluation_report_with_day.html"
    if output_path.exists():
        output_path.unlink()

    try:
        with caplog.at_level("INFO", logger="mars"):
            report.write_html(str(output_path), max_plots=1)
        html_text = output_path.read_text(encoding="utf-8")
        assert "Missing Trend By Day" in html_text
        assert "Data Source" in html_text
        assert "EXT_SOURCE_1" in html_text
        assert "EXT_SOURCE_2" in html_text
        assert "Export Feature List" in html_text
        assert "Jump to Feature" in html_text
        assert "Jumps to the matching row in Summary." in html_text
        assert "marsOpenAncestorSections" in html_text
        assert "mars-secondary-col" in html_text
        assert html_text.index('id="missing-day-section"') < html_text.index('id="trend-section"')
        assert "Search chart features..." in html_text
        assert 'class="mars-source-checkbox"' in html_text
        assert "Bin Type" not in html_text
    finally:
        if output_path.exists():
            output_path.unlink()
        if artifacts_dir.exists() and not any(artifacts_dir.iterdir()):
            artifacts_dir.rmdir()


def test_trend_threshold_style_rules_cover_purple_and_three_color_thresholds():
    iv_rule = MarsBinningReport._trend_style_rule("iv")
    psi_rule = MarsBinningReport._trend_style_rule("psi")
    lift_min_rule = MarsBinningReport._summary_style_rule("lift_min")
    lift_rule = MarsBinningReport._trend_style_rule("lift")
    assert iv_rule is not None
    assert psi_rule is not None
    assert lift_min_rule is not None
    assert lift_rule is not None

    iv_gradient = MarsBinningReport._cell_style(
        0.15,
        semantic="good_high",
        vmin=0.0,
        vmax=1.0,
        style_rule=iv_rule,
    )
    iv_purple = MarsBinningReport._cell_style(
        0.25,
        semantic="good_high",
        vmin=0.0,
        vmax=1.0,
        style_rule=iv_rule,
    )
    psi_green = MarsBinningReport._cell_style(
        0.0,
        semantic="risk_high",
        vmin=0.0,
        vmax=1.0,
        style_rule=psi_rule,
    )
    psi_red = MarsBinningReport._cell_style(
        0.25,
        semantic="risk_high",
        vmin=0.0,
        vmax=1.0,
        style_rule=psi_rule,
    )
    lift_min_green = MarsBinningReport._cell_style(
        0.5,
        semantic="risk_high",
        vmin=0.0,
        vmax=1.0,
        style_rule=lift_min_rule,
    )
    lift_min_yellow = MarsBinningReport._cell_style(
        0.7,
        semantic="risk_high",
        vmin=0.0,
        vmax=1.0,
        style_rule=lift_min_rule,
    )
    lift_purple = MarsBinningReport._cell_style(
        1.5,
        semantic="good_high",
        vmin=0.0,
        vmax=2.0,
        style_rule=lift_rule,
    )

    assert "130, 144, 160" in iv_gradient
    assert "160, 98, 196" in iv_purple
    assert "99, 190, 123" in psi_green
    assert "248, 105, 107" in psi_red
    assert "160, 98, 196" in lift_min_green
    assert "255, 235, 132" in lift_min_yellow
    assert "160, 98, 196" in lift_purple


def test_grouped_pivot_recomputes_pct_and_sorts_features_by_total_iv():
    detail_df = pd.DataFrame(
        [
            {"data_source": "SRC_A", "feature": "feature_high", "bin_label": "A", "bin_index": 0, "bin_type": "\u9996\u5c3e\u7ec4", "grp": "202401", "bad": 1, "count": 20, "lift": 1.30, "iv_bin": 0.10},
            {"data_source": "SRC_B", "feature": "feature_high", "bin_label": "A", "bin_index": 0, "bin_type": "\u9996\u5c3e\u7ec4", "grp": "202401", "bad": 0, "count": 10, "lift": 1.20, "iv_bin": 0.08},
            {"data_source": "SRC_A", "feature": "feature_high", "bin_label": "Missing", "bin_index": -1, "bin_type": "\u7a7a\u503c\u7ec4", "grp": "202401", "bad": 1, "count": 10, "lift": 1.10, "iv_bin": 0.07},
            {"data_source": "SRC_A", "feature": "feature_low", "bin_label": "A", "bin_index": 0, "bin_type": "\u9996\u5c3e\u7ec4", "grp": "202401", "bad": 1, "count": 20, "lift": 1.05, "iv_bin": 0.04},
            {"data_source": "SRC_A", "feature": "feature_low", "bin_label": "B", "bin_index": 1, "bin_type": "\u6b63\u5e38\u7ec4", "grp": "202401", "bad": 1, "count": 20, "lift": 1.01, "iv_bin": 0.03},
        ]
    )

    html_text = MarsBinningReport._build_grouped_pivot_section_html(
        detail_df,
        group_col="grp",
        feature_sources={"feature_high": "SRC", "feature_low": "SRC"},
    )

    assert "Search grouped pivot..." in html_text
    assert "25.00%" in html_text
    assert "75.00%" in html_text
    assert "Missing" in html_text
    assert html_text.index("feature_high") < html_text.index("feature_low")
    assert "SRC_A" not in html_text
    assert "SRC_B" not in html_text
    assert "marsStartColumnResize(event, 'mars-pivot-target', 'bin')" in html_text
    assert "--mars-bin-col-width: 140px;" in html_text
    assert 'data-table-kind="pivot"' in html_text
    assert "No grouped pivot rows match current filters." in html_text


def test_feature_start_aware_baseline_reanchors_monthly_psi_and_summary_metrics(feature_start_aware_df):
    default_report, _ = _profile_risk_report(
        feature_start_aware_df,
        target="target",
        features=["x"],
        group_col="month",
        time_col="biz_dt",
        binning_type="native",
        binner_params={"method": "quantile", "n_bins": 2},
    )
    aware_report, _ = _profile_risk_report(
        feature_start_aware_df,
        target="target",
        features=["x"],
        group_col="month",
        time_col="biz_dt",
        feature_start_aware_baseline=True,
        binning_type="native",
        binner_params={"method": "quantile", "n_bins": 2},
    )

    aware_summary = _as_pandas(aware_report.summary_table)
    default_summary = _as_pandas(default_report.summary_table)
    aware_psi = _as_pandas(aware_report.trend_tables["psi"])

    aware_row = aware_summary.loc[aware_summary["feature"] == "x"].iloc[0]
    default_row = default_summary.loc[default_summary["feature"] == "x"].iloc[0]
    psi_row = aware_psi.loc[aware_psi["feature"] == "x"].iloc[0]

    assert aware_report.report_meta["feature_start_aware_baseline"] is True
    assert aware_report.report_meta["feature_start_baseline_dates"] == {"x": "2024-02-15"}
    assert default_row["psi_max"] > 0.1
    assert "202401" not in aware_psi.columns
    assert psi_row["202402"] == pytest.approx(0.0, abs=1e-9)
    assert psi_row["202403"] == pytest.approx(0.0, abs=1e-9)
    assert psi_row["Total"] == pytest.approx(0.0, abs=1e-9)
    assert aware_row["psi_max"] == pytest.approx(0.0, abs=1e-9)
    assert aware_row["rc_min"] == pytest.approx(1.0, abs=1e-9)


@pytest.mark.parametrize("include_missing", [False, True])
def test_feature_start_aware_baseline_exact_monthly_cutover_keeps_feb_psi_zero(include_missing):
    df = _make_exact_start_aware_monthly_df()
    evaluator = MarsBinEvaluator(binner_params={"method": "quantile", "n_bins": 2})

    run = evaluator.evaluate(
        df,
        target="target",
        features=["EXT_SOURCE_1"],
        time_col="dt",
        time_grain="month",
        feature_start_aware_baseline=True,
        psi_include_missing=include_missing,
    )
    report = run.report

    psi_df = _as_pandas(report.trend_tables["psi"])
    psi_row = psi_df.loc[psi_df["feature"] == "EXT_SOURCE_1"].iloc[0]

    assert report.report_meta["feature_start_aware_baseline"] is True
    assert report.report_meta["feature_start_baseline_dates"] == {"EXT_SOURCE_1": "2018-02-15"}
    assert "201801" not in psi_df.columns
    assert psi_row["201802"] == pytest.approx(0.0, abs=1e-9)
    assert psi_row["201803"] == pytest.approx(0.0, abs=1e-9)
    assert psi_row["201804"] == pytest.approx(0.0, abs=1e-9)
    assert psi_row["Total"] == pytest.approx(0.0, abs=1e-9)


def test_profile_risk_exposes_psi_missing_and_special_scope() -> None:
    df = pl.DataFrame(
        {
            "month": ["2024-01"] * 8 + ["2024-02"] * 8,
            "score": [
                -999.0,
                0.10,
                0.20,
                0.30,
                0.70,
                0.80,
                0.90,
                1.00,
                -999.0,
                -999.0,
                -999.0,
                None,
                0.70,
                0.80,
                0.90,
                1.00,
            ],
            "target": [0, 0, 0, 0, 1, 1, 1, 1] * 2,
        }
    )

    base_run = profile_risk(
        df,
        target="target",
        features=["score"],
        group_col="month",
        binning_type="native",
        binner_params={
            "method": "quantile",
            "n_bins": 2,
            "missing_values": [-999],
            "special_values": [-999],
        },
        psi_include_missing=False,
        psi_include_special=False,
    )
    scoped_run = profile_risk(
        df,
        target="target",
        features=["score"],
        group_col="month",
        binning_type="native",
        binner_params={
            "method": "quantile",
            "n_bins": 2,
            "missing_values": [-999],
            "special_values": [-999],
        },
        psi_include_missing=True,
        psi_include_special=True,
    )

    base_psi = _as_pandas(base_run.report.trend_tables["psi"]).loc[0, "2024-02"]
    scoped_psi = _as_pandas(scoped_run.report.trend_tables["psi"]).loc[0, "2024-02"]

    assert scoped_psi != pytest.approx(base_psi)
    assert scoped_run.report.report_meta["psi_include_missing"] is True
    assert scoped_run.report.report_meta["psi_include_special"] is True


def test_feature_start_aware_baseline_supports_custom_profile_by_with_dt_col(feature_start_aware_df):
    report, _ = _profile_risk_report(
        feature_start_aware_df,
        target="target",
        features=["x"],
        group_col="segment",
        time_col="biz_dt",
        feature_start_aware_baseline=True,
        binning_type="native",
        binner_params={"method": "quantile", "n_bins": 2},
    )

    psi_df = _as_pandas(report.trend_tables["psi"])
    psi_row = psi_df.loc[psi_df["feature"] == "x"].iloc[0]

    assert report.report_meta["feature_start_aware_baseline"] is True
    assert report.report_meta["feature_start_baseline_dates"] == {"x": "2024-02-15"}
    assert "ACTIVE_A" in psi_df.columns
    assert "ACTIVE_B" in psi_df.columns
    assert "PRE" not in psi_df.columns
    assert psi_row["ACTIVE_A"] == pytest.approx(0.0, abs=1e-9)
    assert psi_row["ACTIVE_B"] == pytest.approx(0.0, abs=1e-9)


def test_feature_start_aware_baseline_is_ignored_when_benchmark_df_is_provided(feature_start_aware_df, caplog):
    benchmark_df = feature_start_aware_df.select(["biz_dt", "x", "target"])

    with caplog.at_level("WARNING", logger="mars"):
        report, _ = _profile_risk_report(
            feature_start_aware_df,
            target="target",
            features=["x"],
            group_col="month",
            time_col="biz_dt",
            benchmark_df=benchmark_df,
            feature_start_aware_baseline=True,
            binning_type="native",
            binner_params={"method": "quantile", "n_bins": 2},
        )

    assert report.report_meta["feature_start_aware_baseline"] is False
    assert any("ignored because `benchmark_df` was provided" in message for message in caplog.messages)


def test_profile_risk_handles_notebook_drift_missing_and_special_values() -> None:
    rng = np.random.default_rng(2027)
    rows = 180
    month_idx = np.arange(rows) // 60
    months = np.array(["2024-01", "2024-02", "2024-03"])[month_idx]
    stable = rng.normal(loc=0.0, scale=1.0, size=rows)
    drift = rng.normal(loc=month_idx * 0.6, scale=1.0, size=rows)
    utilization = rng.uniform(0.02, 0.95, size=rows)
    special_feature = rng.normal(loc=0.0, scale=1.0, size=rows).astype(object)
    special_feature[::19] = -999
    special_feature[::23] = None
    raw_score = 1.4 * drift + 0.7 * utilization + rng.normal(scale=0.4, size=rows)
    target = (raw_score > np.median(raw_score)).astype(int)

    df = pl.DataFrame(
        {
            "month": months.tolist(),
            "stable": stable,
            "drift": drift,
            "utilization": utilization,
            "special_feature": special_feature.tolist(),
            "target": target,
        }
    )

    risk_profile = profile_risk(
        df,
        target="target",
        features=["stable", "drift", "utilization", "special_feature"],
        group_col="month",
        binning_type="native",
        binner_params={
            "method": "quantile",
            "n_bins": 4,
            "special_values": [-999],
            "missing_values": [-999],
        },
    )

    report = risk_profile.report
    summary = _as_pandas(report.summary_table)
    psi_table = _as_pandas(report.trend_tables["psi"])

    assert set(summary["feature"]) == {"stable", "drift", "utilization", "special_feature"}
    assert "drift" in psi_table["feature"].to_list()
    assert risk_profile.metadata["feature_count"] == 4
    assert risk_profile.targets == ["target"]
