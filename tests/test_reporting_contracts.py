"""Reporting 层公开契约回归测试。"""

from __future__ import annotations

from importlib import resources

import polars as pl
import pytest
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from mars.reporting import MarsBinningReport, MarsHtmlRenderResult, MarsProfileReport
from mars.reporting._binning_excel import _BinningExcelWriter
from mars.reporting._binning_html import _BinningHtmlRenderer
from mars.reporting._binning_plot import _BinningPlotRenderer
from mars.reporting._matplotlib import ensure_matplotlib_environment, require_pyplot
from mars.reporting._profile_excel import _ProfileExcelWriter
from mars.reporting.plotter import MarsPlotter


def _sample_binning_report() -> MarsBinningReport:
    """构造覆盖 HTML、Excel 与绘图入口的最小分箱报告。"""
    summary = pl.DataFrame(
        {
            "feature": ["income"],
            "data_source": ["base"],
            "iv": [0.12],
            "ks": [18.0],
            "auc": [0.62],
            "psi_max": [0.03],
            "rc_min": [1.0],
            "lift_min": [0.8],
            "lift_max": [1.5],
            "missing": [0.1],
            "target": ["target"],
        },
    )
    detail = pl.DataFrame(
        {
            "y": ["target", "target", "target"],
            "feature": ["income", "income", "income"],
            "trend": ["ascending", "ascending", "ascending"],
            "mars_group": ["2024-01", "2024-01", "Total"],
            "bin_index": [0, 1, 9999],
            "bin_label": ["low", "high", "Total"],
            "count": [60.0, 40.0, 100.0],
            "observed_count": [60.0, 40.0, 100.0],
            "bad": [6.0, 8.0, 14.0],
            "good": [54.0, 32.0, 86.0],
            "pct": [0.6, 0.4, 1.0],
            "bad_rate": [0.1, 0.2, 0.14],
            "lift": [0.7, 1.4, 1.0],
            "cum_count": [60.0, 100.0, 100.0],
            "cum_observed_count": [60.0, 100.0, 100.0],
            "cum_bad": [6.0, 14.0, 14.0],
            "cum_bad_rate": [0.1, 0.14, 0.14],
            "psi_bin": [0.01, 0.02, 0.03],
            "ks_bin": [8.0, 18.0, 18.0],
            "auc_bin": [0.2, 0.42, 0.62],
            "iv_bin": [0.04, 0.08, 0.12],
            "total_count": [100.0, 100.0, 100.0],
            "bin_type": ["正常组", "正常组", "汇总组"],
            "data_source": ["base", "base", "base"],
        },
    )
    trend_tables = {
        "iv": pl.DataFrame({"feature": ["income"], "dtype": ["numeric"], "2024-01": [0.12], "Total": [0.12]}),
        "bad_rate": pl.DataFrame({"feature": ["income"], "dtype": ["numeric"], "2024-01": [0.14], "Total": [0.14]}),
    }
    return MarsBinningReport(
        summary_table=summary,
        trend_tables=trend_tables,
        detail_table=detail,
        feature_data_source={"income": "base"},
        report_meta={
            "targets": ["target"],
            "row_count": 100,
            "feature_count": 1,
            "profile_by_input": "mars_group",
            "group_count": 1,
            "dt_col": "biz_dt",
            "start_dt": "2024-01-01",
            "end_dt": "2024-01-31",
            "event_rate_by_target": {"target": 0.14},
        },
    )


def test_binning_report_html_exports_public_sections(tmp_path) -> None:
    """HTML 导出需要保留公开区块和运行时入口。"""
    report = _sample_binning_report()
    output_path = tmp_path / "report.html"

    report.write_html(str(output_path), max_plots=0)

    html_text = output_path.read_text(encoding="utf-8")
    assert "<title>MARS Evaluation Report</title>" in html_text
    assert "Dataset Overview" in html_text
    assert "Summary" in html_text
    assert "Trend Tables" in html_text
    assert "Grouped Pivot" in html_text
    assert "Charts" in html_text
    assert "Global search across tables and charts" in html_text
    assert "marsQueueRefresh" in html_text


def test_write_html_supports_spa_views_and_asset_chart_mode(tmp_path) -> None:
    report = _sample_binning_report()

    inline_path = tmp_path / "inline.html"
    report.write_html(
        str(inline_path),
        max_plots=1,
        chart_embed_mode="inline",
    )
    inline_html = inline_path.read_text(encoding="utf-8")
    assert 'data-mars-view="overview"' in inline_html
    assert 'data-mars-view="charts"' in inline_html
    assert 'href="#overview"' in inline_html
    assert 'href="#charts"' in inline_html
    assert "marsApplyPageFromHash" in inline_html
    assert "marsJumpToGlobalSearch" in inline_html
    assert "marsShowChartCandidates" in inline_html
    assert "marsChooseChartCandidate" in inline_html
    assert 'data-target="target"' in inline_html

    asset_path = tmp_path / "asset.html"
    report.write_html(
        str(asset_path),
        max_plots=1,
        chart_embed_mode="asset",
    )
    asset_html = asset_path.read_text(encoding="utf-8")
    asset_dir = tmp_path / "asset_assets"
    assert list(asset_dir.glob("*.png"))
    assert 'data-src="asset_assets/' in asset_html
    assert "data:image/png;base64" not in asset_html


def test_write_html_defaults_to_500_plots_and_preserves_trend_switch(
    tmp_path,
    monkeypatch,
) -> None:
    report = _sample_binning_report()
    captured: dict[str, object] = {}

    def fake_chart_section(renderer: object, **kwargs: object) -> str:
        _ = renderer
        captured.update(kwargs)
        return '<details id="chart-section" class="mars-section"><summary>Charts</summary></details>'

    monkeypatch.setattr(_BinningHtmlRenderer, "_build_chart_section_html", fake_chart_section)
    report.write_html(str(tmp_path / "default.html"))

    assert captured["max_plots"] == 500
    assert captured["chart_embed_mode"] == "auto"

    report.write_html(
        str(tmp_path / "without_trends.html"),
        include_trends=False,
        include_charts=False,
    )
    without_trends_html = (tmp_path / "without_trends.html").read_text(encoding="utf-8")
    assert 'id="missing-day-section"' not in without_trends_html


def test_chart_embed_mode_auto_uses_the_large_report_threshold() -> None:
    assert _BinningHtmlRenderer._resolve_chart_embed_mode("auto", 50) == "inline"
    assert _BinningHtmlRenderer._resolve_chart_embed_mode("auto", 51) == "asset"
    with pytest.raises(ValueError, match="chart_embed_mode"):
        _BinningHtmlRenderer._resolve_chart_embed_mode("invalid", 1)


def test_write_html_auto_assetizes_more_than_50_chart_candidates(
    tmp_path,
    monkeypatch,
) -> None:
    source = _sample_binning_report()
    summary = pl.concat(
        [
            source.summary_table.with_columns(pl.lit(f"feature_{index}").alias("feature"))
            for index in range(51)
        ]
    )
    report = MarsBinningReport(
        summary_table=summary,
        trend_tables=source.trend_tables,
        detail_table=source.detail_table,
        feature_data_source={f"feature_{index}": "base" for index in range(51)},
        report_meta=source.report_meta,
    )
    pyplot = require_pyplot(feature_name="test")

    def fake_build_feature_binning_risk_figure(**kwargs: object):
        _ = kwargs
        return pyplot.figure()

    monkeypatch.setattr(
        MarsPlotter,
        "_build_feature_binning_risk_figure",
        staticmethod(fake_build_feature_binning_risk_figure),
    )

    output_path = tmp_path / "large.html"
    report.write_html(str(output_path), max_plots=500)

    asset_dir = tmp_path / "large_assets"
    assert len(list(asset_dir.glob("*.png"))) == 51
    html_text = output_path.read_text(encoding="utf-8")
    assert 'data-src="large_assets/' in html_text
    assert "data:image/png;base64" not in html_text


def test_report_objects_delegate_to_internal_renderers() -> None:
    """报告对象保持数据容器身份，不再通过 mixin 继承导出能力。"""
    assert _BinningExcelWriter not in MarsBinningReport.__mro__
    assert _BinningHtmlRenderer not in MarsBinningReport.__mro__
    assert _BinningPlotRenderer not in MarsBinningReport.__mro__
    assert _ProfileExcelWriter not in MarsProfileReport.__mro__

    binning_report = _sample_binning_report()
    profile_report = MarsProfileReport(
        overview=pl.DataFrame({"feature": ["age"], "missing_rate": [0.0]}),
        dq_tables={},
        stats_tables={},
    )

    assert callable(binning_report.write_excel)
    assert callable(binning_report.write_html)
    assert callable(binning_report.build_risk_trend_figures)
    assert callable(binning_report.save_risk_trend_images)
    assert callable(binning_report.render_risk_trends_html)
    assert callable(binning_report.plot_risk_trends)
    assert callable(profile_report.write_excel)


def test_binning_report_excel_uses_template_columns_only(tmp_path) -> None:
    """Excel 明细导出只能写模板表头已有列。"""
    report = _sample_binning_report()
    output_path = tmp_path / "report.xlsx"

    report.write_excel(str(output_path), engine="openpyxl")

    template_path = resources.files("mars.reporting").joinpath(
        "template",
        "mars_bin_report_linux.xlsx",
    )
    template_workbook = load_workbook(template_path)
    output_workbook = load_workbook(output_path)
    try:
        template_sheet = template_workbook["分组明细"]
        output_sheet = output_workbook["分组明细"]
        template_ref = next(iter(template_sheet.tables.values())).ref
        output_ref = next(iter(output_sheet.tables.values())).ref
        _, _, template_max_col, _ = range_boundaries(template_ref)
        _, _, output_max_col, _ = range_boundaries(output_ref)
        template_headers = [
            template_sheet.cell(row=1, column=col_idx).value
            for col_idx in range(1, template_max_col + 1)
        ]
        output_headers = [
            output_sheet.cell(row=1, column=col_idx).value
            for col_idx in range(1, output_max_col + 1)
        ]
        assert output_headers == template_headers
        assert output_max_col == template_max_col
    finally:
        template_workbook.close()
        output_workbook.close()


def test_binning_report_plot_entry_uses_shared_figure_builder(monkeypatch) -> None:
    """绘图公开入口必须复用共享 figure builder。"""
    report = _sample_binning_report()
    captured: dict[str, object] = {}
    pyplot = require_pyplot(feature_name="test")
    fig = pyplot.figure()

    def fake_build_feature_binning_risk_figure(**kwargs: object):
        captured.update(kwargs)
        return fig

    def fake_display_figure(figure, *, dpi: int, close: bool) -> None:
        captured["displayed_figure"] = figure
        captured["dpi"] = dpi
        captured["close"] = close

    monkeypatch.setattr(
        "mars.reporting._binning_plot.MarsPlotter._build_feature_binning_risk_figure",
        staticmethod(fake_build_feature_binning_risk_figure),
    )
    monkeypatch.setattr(_BinningPlotRenderer, "_display_figure", staticmethod(fake_display_figure))

    result = report.plot_risk_trends(features="income", show_risk="both", max_plots=1, dpi=90)

    assert result is None
    assert captured["feature"] == "income"
    assert captured["group_col"] == "mars_group"
    assert captured["target_name"] == "target"
    assert captured["show_risk"] == "both"
    assert captured["time_range"] == ("2024-01-01", "2024-01-31")
    assert captured["displayed_figure"] is fig
    assert captured["dpi"] == 90
    assert captured["close"] is True
    pyplot.close(fig)


def test_risk_trend_report_entries_require_time_range(tmp_path) -> None:
    source = _sample_binning_report()
    report = MarsBinningReport(
        summary_table=source.summary_table,
        trend_tables=source.trend_tables,
        detail_table=source.detail_table,
        detail_group_col=source.detail_group_col,
    )

    with pytest.raises(ValueError, match="time_col"):
        report.build_risk_trend_figures(features="income")
    with pytest.raises(ValueError, match="time_col"):
        report.plot_risk_trends(features="income")
    with pytest.raises(ValueError, match="time_col"):
        report.render_risk_trends_html(features="income")
    with pytest.raises(ValueError, match="time_col"):
        report.save_risk_trend_images(tmp_path, features="income")
    with pytest.raises(ValueError, match="time_col"):
        report.write_html(tmp_path / "missing_time.html", max_plots=1)


def test_binning_report_build_save_and_render_risk_trends(tmp_path) -> None:
    """风险趋势图 public API 支持 build、save 和 HTML fragment 渲染。"""
    report = _sample_binning_report()
    pyplot = require_pyplot(feature_name="test")

    figures = report.build_risk_trend_figures(features="income", dpi=90)
    try:
        assert len(figures) == 1
        assert figures[0].dpi == 90
    finally:
        for figure in figures:
            pyplot.close(figure)

    asset_dir = tmp_path / "report" / "assets"
    svg_assets = report.save_risk_trend_images(
        asset_dir,
        features="income",
        image_format="svg",
        filename_prefix="risk",
    )
    png_assets = report.save_risk_trend_images(
        asset_dir,
        features="income",
        image_format="png",
        filename_prefix="risk_png",
    )

    assert svg_assets[0].name == "risk_001_target_income.svg"
    assert svg_assets[0].read_text(encoding="utf-8").lstrip().startswith("<?xml")
    assert png_assets[0].suffix == ".png"

    try:
        report.save_risk_trend_images(
            asset_dir,
            features="income",
            image_format="svg",
            filename_prefix="risk",
            overwrite=False,
        )
    except FileExistsError as exc:
        assert "risk_001_target_income.svg" in str(exc)
    else:
        raise AssertionError("overwrite=False should reject an existing asset")

    inline_svg = report.render_risk_trends_html(
        features="income",
        image_format="svg",
        embed_mode="inline",
    )
    assert isinstance(inline_svg, MarsHtmlRenderResult)
    assert '<div class="mars-risk-trends">' in inline_svg.html
    assert "<svg" in inline_svg.html
    assert "<html" not in inline_svg.html.lower()
    assert "<body" not in inline_svg.html.lower()
    assert inline_svg.assets == []
    assert inline_svg.figures is None

    inline_png = report.render_risk_trends_html(
        features="income",
        image_format="png",
        embed_mode="inline",
        include_title=False,
        include_caption=False,
    )
    assert "data:image/png;base64," in inline_png.html

    asset_result = report.render_risk_trends_html(
        features="income",
        image_format="svg",
        embed_mode="asset",
        output_dir=asset_dir,
        relative_to=tmp_path / "report",
        filename_prefix="asset",
    )
    assert asset_result.assets[0].name == "asset_001_target_income.svg"
    assert 'src="assets/asset_001_target_income.svg"' in asset_result.html


def test_binning_report_render_can_return_figures() -> None:
    """HTML 渲染可选返回未关闭的 figure，供调用方继续加工。"""
    report = _sample_binning_report()
    pyplot = require_pyplot(feature_name="test")

    result = report.render_risk_trends_html(
        features="income",
        image_format="png",
        embed_mode="inline",
        return_figures=True,
    )
    try:
        assert result.figures is not None
        assert len(result.figures) == 1
        assert result.figures[0].number in pyplot.get_fignums()
    finally:
        for figure in result.figures or []:
            pyplot.close(figure)


def test_binning_report_plot_can_return_figures(monkeypatch) -> None:
    """plot_risk_trends 兼容旧返回值，并可选返回未关闭 figure。"""
    report = _sample_binning_report()
    pyplot = require_pyplot(feature_name="test")
    captured: dict[str, object] = {}

    def fake_display_figure(figure, *, dpi: int, close: bool) -> None:
        captured["figure"] = figure
        captured["dpi"] = dpi
        captured["close"] = close

    monkeypatch.setattr(_BinningPlotRenderer, "_display_figure", staticmethod(fake_display_figure))

    figures = report.plot_risk_trends(features="income", return_figures=True, dpi=120)
    try:
        assert figures is not None
        assert len(figures) == 1
        assert captured["figure"] is figures[0]
        assert captured["dpi"] == 120
        assert captured["close"] is False
    finally:
        for figure in figures or []:
            pyplot.close(figure)


def test_matplotlib_environment_prepares_config_dir(monkeypatch, tmp_path) -> None:
    """Matplotlib 初始化 helper 会准备可写配置目录。"""
    monkeypatch.delenv("MPLCONFIGDIR", raising=False)
    monkeypatch.delenv("MPLBACKEND", raising=False)

    resolved = ensure_matplotlib_environment()
    probe = resolved / "probe.txt"
    probe.write_text("ok", encoding="utf-8")

    assert resolved.exists()
    assert resolved.name == "mars-risk-matplotlib"
    assert probe.read_text(encoding="utf-8") == "ok"
