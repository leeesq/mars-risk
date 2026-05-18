# mars/analysis/report.py

import os
import sys
import html
import json
from copy import copy
from importlib import resources
import polars as pl
import pandas as pd
import numpy as np
from typing import Dict, Tuple, Optional, Union, List, Any, NamedTuple
from mars.utils.logger import logger

try:
    from IPython.display import display, HTML
except ImportError:
    display = None


def _as_pandas_frame(df: Union[pl.DataFrame, pd.DataFrame]) -> pd.DataFrame:
    """
    将展示层输入统一转换为 Pandas DataFrame。

    Parameters
    ----------
    df : Union[pl.DataFrame, pd.DataFrame]
        需要用于样式渲染或 Excel 导出的数据表。

    Returns
    -------
    pd.DataFrame
        转换后的 Pandas DataFrame。若输入本身已为 Pandas，则直接返回原对象。
    """
    if isinstance(df, pl.DataFrame):
        return df.to_pandas()
    return df
    
class ProfileData(NamedTuple):
    """画像报告底层数据对象集合。"""

    overview: Union[pl.DataFrame, pd.DataFrame]
    dq_trends: Dict[str, Union[pl.DataFrame, pd.DataFrame]]
    stats_trends: Dict[str, Union[pl.DataFrame, pd.DataFrame]]

class MarsProfileReport:
    """
    数据特征画像与质量评估报告容器。

    作为数据探查（EDA）流水线的标准输出载体，该组件负责统一管理并呈现由 
    `MarsDataProfiler` 产出的高维特征统计指标与多维趋势矩阵。系统封装了对底层
    数据帧的只读访问接口、面向 Jupyter 环境的交互式富文本渲染，以及携带高保真
    条件格式的跨平台电子表格（Excel）持久化导出能力。

    Parameters
    ----------
    overview : DataFrame
        全量特征概览宽表。包含全体特征的全局数据质量（DQ）度量与统计分布特征计算结果。
        
    dq_tables : dict of str to DataFrame
        数据质量指标的分组趋势透视表字典。键为具体的度量名称（如 'missing', 'zeros', 'unique'），
        值为对应特征在各时间切片或客群维度下的交叉透视矩阵。
        
    stats_tables : dict of str to DataFrame
        统计分布指标的分组趋势透视表字典。键为具体的度量名称（如 'mean', 'max', 'p25'），
        值为对应特征在各时间切片或客群维度下的交叉透视矩阵。

    Attributes
    ----------
    overview_table : DataFrame
        内部持有的全量特征概览宽表上下文引用。
        
    dq_tables : dict of str to DataFrame
        内部持有的数据质量指标趋势字典上下文引用。
        
    stats_tables : dict of str to DataFrame
        内部持有的统计分布指标趋势字典上下文引用。

    Notes
    -----
    该容器提供了针对数据质量探查与时序追踪的统一交互层 API。
    其内部暴露的 `show_overview` 与 `show_trend` 方法通过动态构建级联的样式渲染器
    （Pandas Styler），支持在交互式计算环境中直接将统计梯度映射为色带（Colormaps）、
    数据条（Data Bars）及微缩分布字符图（Sparklines）。
    调用持久化导出时，底层引擎将在物理存储层面上严格还原交互式环境中的条件格式规则，
    以确保离线监控报表与线上分析环境的视觉一致性。

    Examples
    --------
    >>> from mars.analysis import MarsDataProfiler
    >>> profiler = MarsDataProfiler(df)
    >>> report = profiler.generate_profile(profile_by="month")
    >>> 
    >>> # 1. 触发交互式富文本视图渲染
    >>> report.show_overview(sort_by="missing_rate")
    >>> report.show_trend("missing", features=["age", "income"])
    >>> 
    >>> # 2. 剥离并获取底层物理数据帧以执行二次开发
    >>> overview_df, dq_dict, stat_dict = report.get_profile_data()
    >>> 
    >>> # 3. 执行携带条件格式映射的跨平台报表持久化导出
    >>> report.write_excel("mars_data_health_audit.xlsx")
    """

    def __init__(
        self, 
        overview: Union[pl.DataFrame, pd.DataFrame],
        dq_tables: Dict[str, Union[pl.DataFrame, pd.DataFrame]],
        stats_tables: Dict[str, Union[pl.DataFrame, pd.DataFrame]]
    ) -> None:
        """
        初始化画像报告容器。

        Parameters
        ----------
        overview : DataFrame
            特征概览宽表。
        dq_tables : dict of str to DataFrame
            数据质量趋势表字典。
        stats_tables : dict of str to DataFrame
            统计指标趋势表字典。
        """
        self.overview_table = overview
        self.dq_tables = dq_tables
        self.stats_tables = stats_tables
        
        # 建立索引：将所有指标名映射到对应的数据源类型 ('dq' 或 'stat')
        # 这允许我们在 show_trend 中快速定位
        self._metric_index: Dict[str, str] = {}
        for k in self.dq_tables.keys():
            self._metric_index[k] = "dq"
        for k in self.stats_tables.keys():
            self._metric_index[k] = "stat"

    def get_profile_data(self) -> ProfileData:
        """
        返回画像报告的原始数据对象。

        Returns
        -------
        ProfileData
            包含概览表、数据质量趋势表字典和统计指标趋势表字典的命名元组。
        """
        return ProfileData(
            overview=self.overview_table,
            dq_trends=self.dq_tables,
            stats_trends=self.stats_tables
        )

    def _repr_html_(self) -> str:
        """
        返回 Jupyter 环境下的 HTML 摘要面板。
        """
        df_ov = self.overview_table
        n_feats = len(df_ov) if hasattr(df_ov, "__len__") else df_ov.height
        
        dq_keys = list(self.dq_tables.keys())
        stat_keys = list(self.stats_tables.keys())

        # 样式定义 (Inline CSS for portability)
        # 胶囊样式，用于包裹指标名
        pill_style = (
            "background-color: #e8f4f8; color: #2980b9; border: 1px solid #bce0eb; "
            "padding: 2px 6px; border-radius: 4px; font-family: monospace; font-size: 0.9em; margin-right: 4px;"
        )
        # 代码块样式
        code_style = (
            "background-color: #f0f0f0; padding: 2px 4px; border-radius: 3px; "
            "font-family: monospace; color: #e74c3c; font-weight: bold;"
        )
        
        # 辅助函数：生成指标徽章列表
        def _fmt_pills(keys):
            """将指标名称渲染为 HTML 胶囊标签。"""
            if not keys: return "<span style='color:#ccc'>None</span>"
            # 为了防止指标太多撑爆屏幕，限制显示数量 (例如只显示前 20 个，后面加 ...)
            display_keys = keys[:30] 
            pills = "".join([f"<span style='{pill_style}'>'{k}'</span>" for k in display_keys])
            if len(keys) > 30:
                pills += f"<span style='color:#999; font-size:0.8em'> (+{len(keys)-30} more)</span>"
            return pills

        # 组装 HTML
        return f"""
        <div style="border: 1px solid #e0e0e0; border-left: 5px solid #2980b9; border-radius: 4px; background: white; max-width: 900px; font-family: 'Segoe UI', sans-serif;">
            
            <div style="padding: 12px 15px; background-color: #f8f9fa; border-bottom: 1px solid #e0e0e0; display: flex; justify-content: space-between; align-items: center;">
                <div style="font-weight: bold; color: #2c3e50; font-size: 1.1em;">
                    📊 Mars Data Profile
                </div>
                <div style="font-size: 0.85em; color: #7f8c8d;">
                    <span style="margin-left: 15px;">🏷️ Features: <b>{n_feats}</b></span>
                    <span style="margin-left: 15px;">🔍 DQ Metrics: <b>{len(dq_keys)}</b></span>
                    <span style="margin-left: 15px;">📉 Stat Metrics: <b>{len(stat_keys)}</b></span>
                </div>
            </div>

            <div style="padding: 15px;">
                
                <div style="margin-bottom: 15px;">
                    <div style="font-size: 0.8em; text-transform: uppercase; color: #95a5a6; font-weight: bold; margin-bottom: 5px;">Quick Actions</div>
                    <div style="display: flex; gap: 20px; font-size: 0.95em;">
                        <div>👉 <span style="{code_style}">.show_overview()</span> &nbsp;<span style="color:#555">View Full Report</span></div>
                        <div>💾 <span style="{code_style}">.write_excel()</span> &nbsp;<span style="color:#555">Export XLSX</span></div>
                        <div>📥 <span style="{code_style}">.get_profile_data()</span> &nbsp;<span style="color:#555">Get Raw Data</span></div>
                    </div>
                </div>

                <div style="border-top: 1px dashed #e0e0e0; padding-top: 12px;">
                    <div style="font-size: 0.8em; text-transform: uppercase; color: #95a5a6; font-weight: bold; margin-bottom: 8px;">
                        Trend Analysis <span style="font-weight:normal; text-transform:none; color:#bbb">(Use <code>.show_trend('metric_name')</code>)</span>
                    </div>
                    
                    <div style="display: flex; margin-bottom: 8px; align-items: baseline;">
                        <div style="width: 80px; font-weight: bold; color: #27ae60; font-size: 0.9em;">DQ:</div>
                        <div style="flex: 1; line-height: 1.6;">
                            {_fmt_pills(dq_keys)}
                        </div>
                    </div>
                    
                    <div style="display: flex; align-items: baseline;">
                        <div style="width: 80px; font-weight: bold; color: #2980b9; font-size: 0.9em;">Stats:</div>
                        <div style="flex: 1; line-height: 1.6;">
                            {_fmt_pills(stat_keys)}
                        </div>
                    </div>
                </div>

            </div>
            
            <div style="padding: 6px 15px; background-color: #fff8e1; border-top: 1px solid #fae5b0; font-size: 0.8em; color: #d35400;">
                💡 <b>Pro Tip:</b> Use <span style="{code_style}">.show_trend('psi')</span> to detect population stability drift.
            </div>
        </div>
        """

    def show_overview(self, 
                      features: Optional[Union[str, List[str]]] = None, 
                      sort_by: Optional[Union[str, List[str]]] = None, 
                      sort_ascending: bool = False) -> "pd.io.formats.style.Styler":
        """
        展示特征概览宽表。

        Parameters
        ----------
        features : str or List[str], optional
            需要展示的特征名称。若为 ``None``，展示全部特征。
        sort_by : str or List[str], optional
            排序依据列。若为 ``None``，默认先按 ``dtype`` 再按 ``missing_rate`` 排序。
        sort_ascending : bool, default False
            是否按 ``sort_by`` 升序排列。

        Returns
        -------
        pd.io.formats.style.Styler
            适合在 Jupyter 环境中直接渲染的样式化概览表。
        """
        # 转换为 Pandas 副本以进行切片
        df = _as_pandas_frame(self.overview_table).copy()
        
        # 特征筛选逻辑
        if features is not None:
            if isinstance(features, str):
                features = [features]
            df = df[df["feature"].isin(features)]

        return self._get_styler(
            df,
            title="Dataset Overview", 
            cmap="RdYlGn_r", 
            sort_by= ["dtype"] + (["missing_rate"] if sort_by is None else ([sort_by] if isinstance(sort_by, str) else sort_by)),
            sort_ascending=sort_ascending, 
            # 指定哪些列应用“红绿灯”配色 (高值=红)
            subset_cols=["missing_rate", "zeros_rate", "unique_rate", "top1_ratio"],
            fmt_as_pct=False # 概览表混合了多种类型，不强制全转百分比，由内部逻辑细分
        )

    def show_trend(self, 
                   metric: str, 
                   features: Optional[Union[str, List[str]]] = None, 
                   group_ascending: bool = True, 
                   sort_by: Union[List[str], str] = "total", 
                   sort_ascending: bool = False) -> "pd.io.formats.style.Styler":
        """
        展示指定指标的分组趋势。

        Parameters
        ----------
        metric : str
            指标名称，例如 ``"missing"``、``"mean"`` 或 ``"psi"``。
        features : str or List[str], optional
            需要展示的特征名称。若为 ``None``，展示全部特征。
        group_ascending : bool, default True
            分组列或时间切片列的横向排序方向。
        sort_by : str or List[str], default 'total'
            趋势表内部排序依据，可以是单列或多列列表。
        sort_ascending : bool, default False
            是否按 ``sort_by`` 升序排列。

        Returns
        -------
        pd.io.formats.style.Styler
            样式化趋势热力表。

        Raises
        ------
        ValueError
            当 ``metric`` 不在当前报告支持的指标范围内时抛出。
        """
        # 路由逻辑：查找指标属于哪个表
        source_type = self._metric_index.get(metric)
        if source_type is None:
            # 提供可用指标的快速提示，方便交互式探索。
            available = list(self._metric_index.keys())
            raise ValueError(f"Metric '{metric}' not found. Available metrics: {available[:10]}...")

        # 获取数据
        if source_type == "dq":
            df_raw = self.dq_tables[metric]
            # DQ 默认配置
            cmap = "RdYlGn_r"  # 红色代表高风险 (高缺失)
            fmt_pct = True     # DQ 指标通常是率 (Rate/Ratio)
            vmin, vmax = 0, 1  # 率通常在 0~1 之间
            
        else: # source_type == "stat"
            df_raw = self.stats_tables[metric]
            # Stats 默认配置
            cmap = "Blues"     # 蓝色代表数值高低 (中性)
            fmt_pct = False    # 统计值通常是绝对值
            vmin, vmax = None, None

        # 特殊指标微调 (Override)
        if metric == "psi":
            cmap = "RdYlGn_r" # PSI 高了是坏事
            fmt_pct = False   # PSI 是数值不是百分比
            vmin, vmax = 0.0, 0.5 # 锚定阈值
        
        df = _as_pandas_frame(df_raw).copy()

        # 特征筛选逻辑
        if features is not None:
            if isinstance(features, str):
                features = [features]
            df = df[df["feature"].isin(features)]

        # 排序
        df = df.sort_values(by=sort_by, ascending=sort_ascending)
        df = self._reorder_trend_cols(df, group_ascending=group_ascending)

        return self._get_styler(
            df,
            title=f"Trend Analysis: {metric}",
            cmap=cmap,
            fmt_as_pct=fmt_pct,
            vmin=vmin, 
            vmax=vmax,
            add_bars=True # 所有趋势表都允许显示 CV 条
        )

    def _reorder_trend_cols(self, df: pd.DataFrame, group_ascending: bool) -> pd.DataFrame:
        """重新排列趋势表的列顺序。"""
        # 定义元数据列和末尾统计列
        meta_cols = ["feature", "dtype", "distribution", "top1_value"]
        stat_cols = ["total", "group_mean", "group_var", "group_cv"]
        
        # 识别中间的分组列（如时间列）
        all_cols = df.columns.tolist()
        group_cols = [c for c in all_cols if c not in meta_cols + stat_cols]
        
        # 排序分组列 (受 group_ascending 控制)
        group_cols_sorted = sorted(group_cols, reverse=not group_ascending)
        
        # 组合最终顺序
        final_order = [c for c in meta_cols if c in all_cols] + \
                      group_cols_sorted + \
                      [c for c in stat_cols if c in all_cols]
        return df[final_order]

    def write_excel(self, 
                    path: str = "mars_report.xlsx", 
                    group_ascending: bool = True,
                    sort_by: Union[str, List[str]] = "total",
                    sort_ascending: bool = False) -> None:
        """
        导出画像 Excel 报告。

        Parameters
        ----------
        path : str, default "mars_report.xlsx"
            输出文件路径。
        group_ascending : bool, default True
            趋势页中分组列的横向排序方向。
        sort_by : str or list of str, default "total"
            趋势页内部的排序依据。
        sort_ascending : bool, default False
            趋势页内部是否按 ``sort_by`` 升序排列。

        Notes
        -----
        该方法依赖 ``xlsxwriter`` 导出带样式的多工作表 Excel 文件。
        """
        logger.info(f"Exporting report to: {path}...")
        
        # 1. 依赖检查
        try:
            import xlsxwriter
        except ImportError:
            logger.error("'xlsxwriter' is required for Excel export. Install it via: pip install xlsxwriter")
            return

        try:
            with pd.ExcelWriter(path, engine="xlsxwriter") as writer:
                #--------------------------------------------------------
                # 1. 导出概览页 (Overview)
                #--------------------------------------------------------
                overview_styler = self.show_overview()
                if overview_styler is not None:
                    overview_styler.to_excel(writer, sheet_name="Overview", index=False)
                
                #--------------------------------------------------------
                # 2. 统一导出所有趋势页 (Trend & DQ)
                #--------------------------------------------------------
                dq_keys = list(self.dq_tables.keys())
                stat_keys = list(self.stats_tables.keys())
                all_metrics = dq_keys + stat_keys
                
                for metric in all_metrics:
                    # 保证导出的表结构与 Notebook 展示完全一致
                    styler = self.show_trend(
                        metric, 
                        group_ascending=group_ascending, 
                        sort_by=sort_by, 
                        sort_ascending=sort_ascending
                    )
                    
                    if styler is not None:
                        prefix = "DQ" if metric in self.dq_tables else "Trend"
                        sheet_name = f"{prefix}_{metric.capitalize()}"[:31]
                        
                        styler.to_excel(writer, sheet_name=sheet_name, index=False)
                        
                        # 确保条件格式锚定的列与导出的表完全吻合
                        self._apply_excel_formatting(
                            writer, sheet_name, metric, 
                            group_ascending=group_ascending,
                            sort_by=sort_by,
                            sort_ascending=sort_ascending
                        )

                # 3. 自动列宽调整
                for sheet in writer.sheets.values():
                    sheet.autofit()
                    
            logger.info("Report exported successfully.")

        except Exception as e:
            logger.error(f"Failed to export Excel: {e}", exc_info=True)

    def _apply_excel_formatting(self, 
                                writer, 
                                sheet_name: str, 
                                metric: str, 
                                group_ascending: bool,
                                sort_by: Union[str, List[str]],
                                sort_ascending: bool):
        """
        为导出的趋势工作表应用条件格式。
        """
        if metric in self.dq_tables:
            raw_df = self.dq_tables[metric]
        else:
            raw_df = self.stats_tables[metric]
            
        # 必须和 show_trend 的内部重排逻辑一模一样，否则 Excel 样式会错位
        df_pd: pd.DataFrame = _as_pandas_frame(raw_df).copy()
        
        # 匹配行排序
        if sort_by in df_pd.columns or (isinstance(sort_by, list) and all(c in df_pd.columns for c in sort_by)):
            df_pd = df_pd.sort_values(by=sort_by, ascending=sort_ascending)
            
        # 匹配列排序
        df_pd = self._reorder_trend_cols(df_pd, group_ascending=group_ascending)
        
        # 动态识别趋势列范围，避免依赖固定列位置。
        meta_and_stat = set(["feature", "dtype", "distribution", "top1_value", "total", "group_mean", "group_var", "group_cv"])
        time_cols = [c for c in df_pd.columns if c not in meta_and_stat]
        
        if not time_cols:
            return

        worksheet = writer.sheets[sheet_name]
        
        # PSI 专用三色阶 (红绿灯)
        if metric == "psi":
            meta_cols = ["feature", "dtype", "distribution", "top1_value"]
            start_col = 0
            for i, col in enumerate(df_pd.columns):
                if col not in meta_cols:
                    start_col = i
                    break
            
            end_col = len(df_pd.columns) - 1
            
            worksheet.conditional_format(1, start_col, len(df_pd), end_col, {
                'type': '3_color_scale',
                'min_type': 'num', 'min_value': 0.05, 'min_color': '#63BE7B', # Green
                'mid_type': 'num', 'mid_value': 0.15, 'mid_color': '#FFEB84', # Yellow
                'max_type': 'num', 'max_value': 0.25, 'max_color': '#F8696B'  # Red
            })

        # 稳定性 Data Bars (针对 group_cv)
        if "group_cv" in df_pd.columns:
            col_idx = df_pd.columns.get_loc("group_cv")
            worksheet.conditional_format(1, col_idx, len(df_pd), col_idx, {
                'type': 'data_bar', 
                'bar_color': '#638EC6', 
                'bar_solid': True,
                'min_type': 'num', 'min_value': 0, 
                'max_type': 'num', 'max_value': 1
            })

    def _get_styler(
        self, 
        df_input: Any, 
        title: str, 
        cmap: str, 
        sort_by: Optional[List[str]] = None,
        sort_ascending: bool = False, # 统一内部 API 命名
        subset_cols: Optional[List[str]] = None, 
        add_bars: bool = False, 
        fmt_as_pct: bool = False,
        vmin: Optional[float] = None, 
        vmax: Optional[float] = None
    ) -> Optional["pd.io.formats.style.Styler"]:
        """
        生成统一的 Pandas Styler 样式对象。
        """
        if df_input is None:
            return None
        df: pd.DataFrame = _as_pandas_frame(df_input)
        if sort_by is not None:
            df = df.sort_values(by=sort_by, ascending=sort_ascending) # 使用统一参数进行底层排序
        if df.empty:
            return None

        # 元数据排除列表
        exclude_meta: List[str] = [
            "feature", "dtype", 
            "group_mean", "group_var", "group_cv",
            "distribution",
            "top1_value"
            ]
        
        # 确定色彩渐变范围
        if subset_cols:
            gradient_cols: List[str] = [c for c in subset_cols if c in df.columns]
        else:
            gradient_cols = [c for c in df.columns if c not in exclude_meta]

        styler = df.style.set_caption(f"<b>{title}</b>").hide(axis="index")
        
        # 应用热力图
        if gradient_cols:
            styler = styler.background_gradient(
                cmap=cmap, 
                subset=gradient_cols, 
                axis=None,
                vmin=vmin,
                vmax=vmax
            )
        
        # 应用数据条
        if add_bars and "group_cv" in df.columns:
            styler = styler.bar(subset=["group_cv"], color='#ff9999', vmin=0, vmax=1, width=90)
            styler = styler.format("{:.4f}", subset=["group_cv", "group_var"])

        # 数值格式化逻辑
        num_cols: pd.Index = df.select_dtypes(include=['number']).columns
        data_cols: List[str] = [c for c in num_cols if c not in ["group_var", "group_cv", "distribution"]]

        pct_format: str = "{:.2%}"  
        float_format: str = "{:.2f}"

        if fmt_as_pct:
            if data_cols:
                styler = styler.format(pct_format, subset=data_cols)
        else:
            pct_cols: List[str] = [
                c for c in df.columns 
                if ("rate" in c or "ratio" in c) and (c in num_cols)
            ]
            
            if pct_cols:
                styler = styler.format(pct_format, subset=pct_cols)
            
            float_cols: List[str] = [c for c in data_cols if c not in pct_cols]
            if float_cols:
                styler = styler.format(float_format, subset=float_cols)
        
        # 分布迷你图样式
        if "distribution" in df.columns:
            styler = styler.set_table_styles([
                {'selector': '.col_distribution', 'props': [
                    # 优先使用 Consolas (Win) 或 Menlo (Mac)，最后 fallback 到 monospace
                    ('font-family', '"Consolas", "Menlo", "Courier New", monospace'), 
                    ('color', '#1f77b4'),
                    ('white-space', 'pre'), # [关键] 防止 HTML 自动压缩连续空格
                    ('font-weight', 'bold'),
                    ('text-align', 'left')
                ]}
            ], overwrite=False)

        # 全局表格外观
        styler = styler.set_table_styles([
            {
                'selector': 'th', 
                'props': [('text-align', 'left'), ('background-color', '#f0f2f5'), ('color', '#333')]
            },
            {
                'selector': 'caption', 
                'props': [('font-size', '1.2em'), ('padding', '10px 0'), ('color', '#2c3e50')]
            }
        ], overwrite=False)

        return styler
    
class MarsEvaluationReport:
    """
    特征效能与稳定性评估报告容器。

    作为风控特征工程流水线的标准输出载体，该组件负责统一管理并呈现由 `MarsBinEvaluator` 
    产出的高维特征评估度量与多维趋势矩阵。系统封装了对底层评估数据帧的只读访问接口、
    面向交互式分析环境的富文本视图渲染，以及跨平台的高保真电子表格持久化导出能力，
    以支撑特征区分度审计与跨期分布漂移监控。

    Parameters
    ----------
    summary_table : Union[pl.DataFrame, pd.DataFrame]
        特征级汇总评估宽表。涵盖特征的全局预测力（如 IV, KS, AUC）与跨期稳定性边界
        （如最大 PSI, 最小风险逻辑一致性相关系数）的核心度量数据。

    trend_tables : Dict[str, Union[pl.DataFrame, pd.DataFrame]]
        核心评估指标的跨期趋势字典。键为具体风控指标名称（如 'psi', 'auc', 'iv', 'bad_rate', 
        'risk_corr'），值为对应特征在各时间切片或客群维度下的交叉透视矩阵。

    detail_table : Union[pl.DataFrame, pd.DataFrame]
        细粒度分箱明细表。包含特征在各时间切片下所有分箱区间的样本分布占比、坏账率、
        提升度（Lift）、证据权重（WOE）及累积风险推演指标。

    group_col : str, optional
        驱动趋势分析与分箱明细切片的分组维度标识。若当前处于单点截面评估模式，
        则该值为 None 或全局默认聚合标识。
    feature_data_source : dict of str to str, optional
        特征到数据源标签的映射。
    dt_col : str, optional
        原始日期列名。
    missing_by_day_table : Union[pl.DataFrame, pd.DataFrame], optional
        按日聚合的缺失明细表。
    report_meta : dict, optional
        报告元信息，例如目标列、绘图配置或上下文标签。

    Attributes
    ----------
    summary_table : DataFrame
        内部持有的特征级汇总评估宽表引用。
        
    trend_tables : dict of str to DataFrame
        内部持有的核心评估指标趋势字典引用。
        
    detail_table : DataFrame
        内部持有的细粒度分箱明细表引用。
        
    group_col : str
        内部挂载的分组维度标识。

    Notes
    -----
    该容器提供了针对特征评估诊断的统一交互层 API。其暴露的 `show_summary` 与 `show_trend` 
    方法通过动态构建样式渲染器（Pandas Styler），支持直接将风控指标梯度映射为预警色带
    与数据条，以加速区分度缺陷与单调性倒挂的物理识别。
    执行持久化导出时，底层引擎严格还原交互式视图中的条件格式与业务阈值规则，确保离线
    模型监控文档与线上审计视图的视觉连贯性。

    Examples
    --------
    >>> from mars.analysis import MarsBinEvaluator
    >>> evaluator = MarsBinEvaluator(target="is_bad")
    >>> report = evaluator.evaluate(df, profile_by="month")
    >>> 
    >>> # 触发交互式特征汇总审计视图
    >>> core_features = ["age", "debt_ratio", "revolving_util"]
    >>> report.show_summary(features=core_features)
    >>> 
    >>> # 追踪特定指标的时间序列漂移轨迹
    >>> report.show_trend("psi", sort_by="Total", sort_ascending=False, group_ascending=True)
    >>> 
    >>> # 执行包含全量分箱明细的监控报表持久化导出
    >>> report.write_excel("mars_feature_evaluation.xlsx")
    """

    def __init__(
        self, 
        summary_table: Union[pl.DataFrame, pd.DataFrame],
        trend_tables: Dict[str, Union[pl.DataFrame, pd.DataFrame]],
        detail_table: Union[pl.DataFrame, pd.DataFrame],
        group_col: Optional[str] = None,
        feature_data_source: Optional[Dict[str, str]] = None,
        dt_col: Optional[str] = None,
        missing_by_day_table: Optional[Union[pl.DataFrame, pd.DataFrame]] = None,
        report_meta: Optional[Dict[str, Any]] = None,
    ) -> None:
        """
        初始化报告容器。

        Parameters
        ----------
        summary_table : Union[pl.DataFrame, pd.DataFrame]
            特征级汇总表。
        trend_tables : Dict[str, Union[pl.DataFrame, pd.DataFrame]]
            指标趋势表字典。
        detail_table : Union[pl.DataFrame, pd.DataFrame]
            最细粒度的分箱明细表。
        group_col : str, optional
            分组列名（例如：'month' 或 'vintage'）。
        feature_data_source : dict of str to str, optional
            特征到数据源标签的映射。
        dt_col : str, optional
            原始日期列名。
        missing_by_day_table : Union[pl.DataFrame, pd.DataFrame], optional
            按日汇总的缺失率明细表。
        report_meta : dict, optional
            报告元信息，例如目标列、绘图配置或上下文标签。
        """
        # 直接存储原始数据，不再强制命名为 _pl，以支持多种类型
        self._summary = summary_table
        self._trend_dict = trend_tables
        self._detail = detail_table
        self.group_col = group_col 
        self.feature_data_source = feature_data_source or {}
        self.dt_col = dt_col
        self._missing_by_day = missing_by_day_table
        self._report_meta = report_meta or {}
        
    @property
    def summary_table(self) -> Union[pl.DataFrame, pd.DataFrame]:
        """
        返回特征汇总评估表。

        Returns
        -------
        pl.DataFrame or pd.DataFrame
            与构造时输入类型一致的汇总表。
        """
        return self._summary

    @property
    def trend_tables(self) -> Dict[str, Union[pl.DataFrame, pd.DataFrame]]:
        """
        返回指标趋势表字典。

        Returns
        -------
        dict of str to pl.DataFrame or pd.DataFrame
            键为指标名称，值为对应趋势宽表。
        """
        return self._trend_dict

    @property
    def detail_table(self) -> Union[pl.DataFrame, pd.DataFrame]:
        """
        返回分箱明细表。

        Returns
        -------
        pl.DataFrame or pd.DataFrame
            与构造时输入类型一致的分箱明细表。
        """
        return self._detail

    @property
    def missing_by_day_table(self) -> Optional[Union[pl.DataFrame, pd.DataFrame]]:
        """
        返回按日聚合的缺失明细表。

        Returns
        -------
        pl.DataFrame or pd.DataFrame or None
            若评估流程生成了按日缺失统计，则返回对应表；否则返回 ``None``。
        """
        return self._missing_by_day

    @property
    def report_meta(self) -> Dict[str, Any]:
        """
        返回报告元信息字典。

        Returns
        -------
        dict of str to Any
            生成报告时记录的辅助元数据。
        """
        return self._report_meta

    def get_evaluation_data(self) -> Tuple[
        Union[pl.DataFrame, pd.DataFrame], 
        Dict[str, Union[pl.DataFrame, pd.DataFrame]], 
        Union[pl.DataFrame, pd.DataFrame]
    ]:
        """
        获取评估报告的原始数据。

        Returns
        -------
        tuple
            依次返回 ``(summary_table, trend_tables, detail_table)``，
            且各对象类型与构造时输入保持一致。
        """
        return self.summary_table, self.trend_tables, self.detail_table

    def _repr_html_(self) -> str:
        """
        返回 Jupyter 环境下的评估摘要面板。
        """
        # 内部展示逻辑统一转为 Pandas 处理
        df_summary_pd = _as_pandas_frame(self.summary_table)
        n_feats = len(df_summary_pd)
        
        # 简单统计报警数 (修正为新的小写列名 psi_max)
        high_risk_psi = 0
        if "psi_max" in df_summary_pd.columns:
            high_risk_psi = sum(df_summary_pd["psi_max"] > 0.25)

        # 样式定义
        pill_style = (
            "background-color: #e8f4f8; color: #2980b9; border: 1px solid #bce0eb; "
            "padding: 2px 6px; border-radius: 4px; font-family: monospace; font-size: 0.9em; margin-right: 4px;"
        )
        
        # 动态生成 Trend 链接
        trend_keys = list(self.trend_tables.keys())
        trend_pills = "".join([f"<span style='{pill_style}'>'{k}'</span>" for k in trend_keys])

        lines = []
        # 查看类操作
        lines.append('👉 <code>.show_summary()</code> &nbsp;<span style="color:#7f8c8d">View Feature Ranking</span>')
        lines.append(f'👉 <code>.show_trend(metric)</code> <span style="color:#7f8c8d">metric: {trend_pills}</span>')
        
        # 数据访问与导出入口
        lines.append('<hr style="margin: 8px 0; border: 0; border-top: 1px dashed #ccc;">')
        lines.append('📥 <code>.get_evaluation_data()</code> &nbsp;<span style="color:#7f8c8d">Get Raw Data (summary, trends, detail)</span>')
        lines.append('💾 <code>.write_excel()</code> &nbsp;<span style="color:#7f8c8d">Export to Excel</span>')

        return f"""
        <div style="border-left: 5px solid #8e44ad; background-color: #f4f6f7; padding: 15px; border-radius: 0 5px 5px 0; font-family: 'Segoe UI', sans-serif;">
            <h3 style="margin:0 0 10px 0; color:#2c3e50;">📉 Mars Feature Evaluation</h3>
            
            <div style="display: flex; gap: 30px; margin-bottom: 12px; font-size: 0.95em;">
                <div><strong>🏷️ Features:</strong> {n_feats}</div>
                <div><strong>🚨 High PSI (>0.25):</strong> <span style="color: {'red' if high_risk_psi > 0 else 'green'}; font-weight:bold;">{high_risk_psi}</span></div>
                <div><strong>📅 Group By:</strong> {self.group_col if self.group_col else 'None (Total Only)'}</div>
            </div>
            
            <div style="font-size:0.9em; line-height:1.8; color:#2c3e50; background: white; padding: 10px; border: 1px solid #e0e0e0; border-radius: 4px;">
                { "<br>".join(lines) }
            </div>
        </div>
        """

    @staticmethod
    def _slugify(value: str) -> str:
        slug = "".join(ch.lower() if ch.isalnum() else "-" for ch in str(value))
        slug = "-".join(part for part in slug.split("-") if part)
        return slug or "section"

    @staticmethod
    def _wrap_html_section(title: str, body: str, section_id: str, subtitle: Optional[str] = None, open_by_default: bool = True) -> str:
        open_attr = " open" if open_by_default else ""
        subtitle_html = f'<div class="mars-section-subtitle">{html.escape(subtitle)}</div>' if subtitle else ""
        return f"""
        <details id="{section_id}" class="mars-section"{open_attr}>
            <summary>{html.escape(title)}</summary>
            {subtitle_html}
            <div class="mars-section-body">
                {body}
            </div>
        </details>
        """

    @staticmethod
    def _is_missing_html_value(value: Any) -> bool:
        if value is None:
            return True
        try:
            return bool(pd.isna(value))
        except TypeError:
            return False

    @classmethod
    def _format_html_value(
        cls,
        value: Any,
        *,
        as_percent: bool = False,
        precision: int = 2,
    ) -> str:
        if cls._is_missing_html_value(value):
            return ""

        if isinstance(value, (np.integer, int, np.floating, float)) and not isinstance(value, bool):
            num = float(value)
            if not np.isfinite(num):
                return ""
            if as_percent:
                return f"{num * 100:.{precision}f}%"
            return f"{num:.{precision}f}"

        return str(value)

    @staticmethod
    def _normalize_search_text(*parts: Any) -> str:
        joined = " ".join("" if part is None else str(part) for part in parts)
        return " ".join(joined.split()).strip().lower()

    @classmethod
    def _is_percent_column(
        cls,
        col_name: Any,
        *,
        metric_name: Optional[str] = None,
    ) -> bool:
        col_lower = str(col_name).strip().lower()
        metric_lower = str(metric_name or "").strip().lower()

        non_percent_labels = {"feature", "dtype", "data_source", "target", "y", "bin"}
        if col_lower in non_percent_labels:
            return False

        if metric_lower in {"missing", "bad_rate"} and col_lower not in non_percent_labels:
            return True

        if col_lower in {"pct", "risk", "missing", "missing_rate", "bad_rate", "cum_bad_rate"}:
            return True

        if col_lower.startswith("missing"):
            return True

        return (
            col_lower.endswith(("_rate", "_ratio", "_pct"))
            or col_lower.startswith(("pct_", "risk_"))
            or col_lower.endswith("%")
        )

    @staticmethod
    def _interpolate_rgb(start: Tuple[int, int, int], end: Tuple[int, int, int], ratio: float) -> Tuple[int, int, int]:
        return tuple(
            int(round(start[idx] + (end[idx] - start[idx]) * ratio))
            for idx in range(3)
        )

    @classmethod
    def _three_color_rgb(cls, ratio: float, *, reverse: bool = False) -> Tuple[int, int, int]:
        ratio = max(0.0, min(1.0, ratio))
        low = (248, 105, 107) if not reverse else (99, 190, 123)
        mid = (255, 235, 132)
        high = (99, 190, 123) if not reverse else (248, 105, 107)
        if ratio <= 0.5:
            return cls._interpolate_rgb(low, mid, ratio * 2.0)
        return cls._interpolate_rgb(mid, high, (ratio - 0.5) * 2.0)

    @classmethod
    def _column_colspan(cls, col_name: Any) -> int:
        return max(1, str(col_name).count("|") + 1)

    @classmethod
    def _format_sort_value(cls, value: Any, sort_type: str) -> str:
        if cls._is_missing_html_value(value):
            return ""

        if sort_type == "number":
            try:
                return f"{float(value):.12g}"
            except (TypeError, ValueError):
                return ""

        return str(value)

    @staticmethod
    def _reorder_group_columns(df: pd.DataFrame, leading_cols: List[str]) -> pd.DataFrame:
        if df.empty:
            return df

        head_cols = [c for c in leading_cols if c in df.columns]
        other_cols = [c for c in df.columns if c not in head_cols]
        non_total = sorted([c for c in other_cols if c != "Total"])
        tail_cols = ["Total"] if "Total" in other_cols else []
        return df[head_cols + non_total + tail_cols]
        """

        if bin_type_values:
            bin_options = ['<option value="__all__">全部 bin_type</option>'] + [
                f'<option value="{html.escape(bin_type)}">{html.escape(bin_type)}</option>'
                for bin_type in bin_type_values
            ]
            control_parts.append(
                f'<label class="mars-select-group">Bin Type'
                f'<select id="mars-pivot-bin-type" onchange="marsUpdatePivotViews()">{"".join(bin_options)}</select>'
                f'</label>'
            )

        view_blocks = []
        bin_type_scope = [None] + bin_type_values if bin_type_values else [None]
        for y_val in y_values:
            for bin_type in bin_type_scope:
                pivot_df = cls._build_pivot_frame(
                    detail_pd,
                    group_col=group_col,
                    y_value=y_val if "y" in detail_pd.columns else None,
                    bin_type_value=bin_type,
                )
                table_id = f"mars-pivot-{cls._slugify(y_val)}-{cls._slugify(bin_type or 'all')}"
                table_html = cls._build_sortable_table_html(
                    pivot_df,
                    table_id,
                    search_placeholder="Search grouped pivot...",
                    empty_text="No grouped pivot data for this selection.",
                )
                label_suffix = f" | {bin_type}" if bin_type else " | 全部 bin_type"
                view_blocks.append(
                    f'<div class="mars-pivot-view" data-y-value="{html.escape(y_val)}" '
                    f'data-bin-type-value="{html.escape(bin_type or "__all__")}">'
                    f'<div class="mars-view-label">{html.escape(y_val + label_suffix)}</div>'
                    f'{table_html}'
                    f'</div>'
                )

        controls_html = (
            f'<div class="mars-inline-controls">{"".join(control_parts)}</div>'
            if control_parts else ""
        )
        return controls_html + "".join(view_blocks)

        """

    @staticmethod
    def _resolve_chart_sort_column(summary_df: pd.DataFrame, requested: str) -> Optional[str]:
        if requested in summary_df.columns:
            return requested
        if "psi_max" in summary_df.columns:
            return "psi_max"

        numeric_cols = [
            c for c in summary_df.select_dtypes(include=[np.number]).columns
            if c not in {"bin_index"}
        ]
        return numeric_cols[0] if numeric_cols else None

    @staticmethod
    def _semantic_for_metric(metric: str) -> str:
        metric = str(metric).lower()
        if metric.startswith("missing") or metric in {"psi", "psi_max", "missing_rate"}:
            return "risk_high"
        if metric.startswith("lift") or metric in {"iv", "auc", "ks", "risk_corr", "rc_min"}:
            return "good_high"
        if metric == "mono":
            return "diverging"
        return "neutral"

    @staticmethod
    def _escape_attr(value: Any) -> str:
        return html.escape("" if value is None else str(value), quote=True)

    @staticmethod
    def _trend_style_rule(metric: Optional[str]) -> Optional[Dict[str, Any]]:
        metric_key = str(metric or "").lower()
        purple_rgb = (160, 98, 196)
        green = (99, 190, 123)
        yellow = (255, 235, 132)
        red = (248, 105, 107)

        rules: Dict[str, Dict[str, Any]] = {
            "missing": {"anchors": (0.0, 0.5, 1.0), "colors": (green, yellow, red)},
            "psi": {"anchors": (0.0, 0.1, 0.25), "colors": (green, yellow, red)},
            "iv": {"anchors": (0.01, 0.05, 0.1), "colors": (red, yellow, green), "purple_above": 0.2, "purple_rgb": purple_rgb},
            "ks": {"anchors": (4.0, 8.0, 12.0), "colors": (red, yellow, green), "purple_above": 16.0, "purple_rgb": purple_rgb},
            "auc": {"anchors": (0.525, 0.55, 0.575), "colors": (red, yellow, green), "purple_above": 0.625, "purple_rgb": purple_rgb},
            "lift": {"anchors": (1.2, 1.3, 1.4), "colors": (red, yellow, green), "purple_above": 1.5, "purple_rgb": purple_rgb},
            "risk_corr": {"anchors": (0.2, 0.5, 0.8), "colors": (red, yellow, green)},
        }
        return rules.get(metric_key)

    @classmethod
    def _summary_style_rule(cls, metric: Optional[str]) -> Optional[Dict[str, Any]]:
        metric_key = str(metric or "").lower()
        if metric_key in {"iv", "ks", "auc", "psi_max", "rc_min", "lift_max", "missing", "missing_min", "missing_max"}:
            mapped = {
                "psi_max": "psi",
                "rc_min": "risk_corr",
                "lift_max": "lift",
                "missing_min": "missing",
                "missing_max": "missing",
            }.get(metric_key, metric_key)
            return cls._trend_style_rule(mapped)
        if metric_key == "lift_min":
            return {
                "anchors": (0.5, 0.6, 0.7, 0.8),
                "colors": ((160, 98, 196), (99, 190, 123), (255, 235, 132), (248, 105, 107)),
            }
        return None

    @staticmethod
    def _sort_metric_display_df(df: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            return df
        if "Total" in df.columns:
            sort_vals = pd.to_numeric(df["Total"], errors="coerce")
            return (
                df.assign(__mars_total_sort=sort_vals)
                .sort_values(["__mars_total_sort", "feature"], ascending=[False, True], na_position="last")
                .drop(columns="__mars_total_sort")
            )
        if "feature" in df.columns:
            return df.sort_values("feature", ascending=True)
        return df

    @classmethod
    def _build_threshold_legend_html(
        cls,
        items: List[Tuple[str, str]],
        *,
        legend_id: str,
    ) -> str:
        if not items:
            return ""
        chips = "".join(
            f'<span class="mars-legend-chip"><strong>{html.escape(label)}</strong> {html.escape(desc)}</span>'
            for label, desc in items
        )
        return f'<div id="{legend_id}" class="mars-legend">{chips}</div>'

    @classmethod
    def _build_dataset_overview_html(cls, report_meta: Dict[str, Any]) -> str:
        if not report_meta:
            return ""

        def fmt_value(value: Any) -> str:
            if value is None or value == "":
                return "N/A"
            return html.escape(str(value))

        cards = [
            ("Rows", fmt_value(report_meta.get("row_count"))),
            ("Features", fmt_value(report_meta.get("feature_count"))),
            ("Profile By", fmt_value(report_meta.get("profile_by_input"))),
            ("Groups", fmt_value(report_meta.get("group_count"))),
        ]
        if report_meta.get("dt_col"):
            cards.append(
                (
                    "Time Range",
                    fmt_value(report_meta.get("start_dt")) + " ~ " + fmt_value(report_meta.get("end_dt")),
                )
            )
        targets = report_meta.get("targets") or []
        cards.append(("Targets", fmt_value(", ".join(str(v) for v in targets) if targets else None)))

        event_rate_map = report_meta.get("event_rate_by_target") or {}
        if event_rate_map:
            rate_text = " | ".join(
                f"{target}: {cls._format_html_value(value, as_percent=True)}" if value is not None else f"{target}: N/A"
                for target, value in event_rate_map.items()
            )
        else:
            rate_text = "N/A"
        cards.append(("Event Rate", html.escape(rate_text)))
        if report_meta.get("feature_start_aware_baseline"):
            active_features = report_meta.get("feature_start_baseline_features") or []
            cards.append(
                (
                    "Start-Aware Baseline",
                    fmt_value(f"Enabled ({len(active_features)} features)"),
                )
            )

        card_html = "".join(
            f'<div class="mars-kpi-card"><div class="mars-kpi-label">{label}</div><div class="mars-kpi-value">{value}</div></div>'
            for label, value in cards
        )
        return f'<section id="dataset-overview" class="mars-overview-grid">{card_html}</section>'

    @classmethod
    def _build_feature_jump_html(cls, features: List[str]) -> str:
        feature_values = sorted({str(feature) for feature in features if str(feature).strip()})
        if not feature_values:
            return ""
        option_html = "".join(
            f'<option value="{html.escape(feature)}"></option>'
            for feature in feature_values
        )
        return (
            '<div class="mars-feature-jump">'
            '<label class="mars-summary-filter-label" for="mars-feature-jump-input">Jump to Feature</label>'
            '<div class="mars-search-cluster">'
            '<input id="mars-feature-jump-input" class="mars-filter-input" type="search" '
            'list="mars-feature-jump-list" placeholder="Jump to Summary feature..." '
            'onkeydown="if(event.key===\'Enter\'){event.preventDefault();marsJumpToFeature();}" />'
            '<datalist id="mars-feature-jump-list">'
            f'{option_html}'
            '</datalist>'
            '<button type="button" class="mars-mini-button" onclick="marsJumpToFeature()">Go</button>'
            '</div>'
            '<div class="mars-footnote">Jumps to the matching row in Summary.</div>'
            '<div id="mars-feature-jump-error" class="mars-search-error"></div>'
            '</div>'
        )

    @staticmethod
    def _table_sticky_role(column_name: Any) -> Optional[str]:
        column_lower = str(column_name).strip().lower()
        if column_lower == "feature":
            return "feature"
        if column_lower == "dtype":
            return "secondary"
        return None

    @staticmethod
    def _sticky_class_for_role(role: Optional[str]) -> str:
        if not role:
            return ""
        return f" mars-sticky-col mars-{role}-col"

    @staticmethod
    def _sticky_inner_class_for_role(role: Optional[str]) -> str:
        if not role:
            return ""
        return " mars-sticky-cell-inner"

    @staticmethod
    def _build_scope_feedback_html(scope_id: str, *, empty_text: str) -> str:
        return (
            f'<div id="{scope_id}-status" class="mars-result-status" aria-live="polite"></div>'
            f'<div id="{scope_id}-empty" class="mars-empty mars-scope-empty" hidden>{html.escape(empty_text)}</div>'
        )

    @classmethod
    def _build_html_styles(cls) -> str:
        return """
                :root { --bg:#f5f7fb; --panel:#fff; --panel-soft:#f9fbfd; --ink:#203040; --muted:#607080; --line:#d9e3eb; --line-soft:#ebf1f6; --accent:#3b87ad; --danger:#c44f4f; --shadow:0 16px 36px rgba(51,82,108,.08); }
                body { margin:0; font-family:"Segoe UI","PingFang SC","Microsoft YaHei",sans-serif; background:radial-gradient(circle at top right,#edf6fb 0%,#f5f7fb 40%,#f8fbfd 100%); color:var(--ink); }
                .mars-page { max-width:1640px; margin:0 auto; padding:22px; }
                .mars-hero,.mars-section { background:var(--panel); border:1px solid var(--line); border-radius:18px; box-shadow:var(--shadow); }
                .mars-hero { padding:22px 24px; margin-bottom:16px; position:relative; overflow:hidden; }
                .mars-hero::after { content:""; position:absolute; inset:auto -80px -90px auto; width:240px; height:240px; background:radial-gradient(circle, rgba(59,135,173,.14) 0%, rgba(59,135,173,0) 72%); pointer-events:none; }
                .mars-hero h1 { margin:0 0 8px 0; font-size:30px; }
                #mars-page-top { position:relative; top:0; }
                .mars-hero p,.mars-footnote,.mars-section-subtitle,.mars-search-error,.mars-view-label,.mars-pivot-source-title,.mars-result-status,.mars-export-helper { color:var(--muted); position:relative; z-index:1; }
                .mars-meta,.mars-nav,.mars-inline-controls { display:flex; flex-wrap:wrap; gap:10px; }
                .mars-meta { margin-top:12px; position:relative; z-index:1; }
                .mars-pill,.mars-nav a { border:1px solid var(--line); background:#f7fbff; border-radius:999px; padding:6px 12px; font-size:13px; color:#36546d; text-decoration:none; }
                .mars-global-tools { margin-top:16px; display:grid; grid-template-columns:minmax(280px,420px) auto minmax(240px,340px) minmax(280px,1fr) minmax(180px,240px); gap:10px; align-items:start; position:relative; z-index:1; }
                .mars-filter-input,.mars-select-group select,.mars-clear-button,.mars-mini-button { border:1px solid var(--line); border-radius:12px; background:#fff; font-size:14px; }
                .mars-filter-input { padding:10px 12px; width:100%; box-sizing:border-box; }
                .mars-search-cluster { display:grid; grid-template-columns:minmax(0,1fr) auto; gap:8px; align-items:center; }
                .mars-select-group { display:inline-flex; gap:8px; align-items:center; font-size:13px; }
                .mars-select-group select { padding:8px 10px; }
                .mars-source-panel { border:1px solid var(--line); border-radius:14px; background:#fff; padding:10px 12px; min-width:280px; }
                .mars-source-header,.mars-source-options { display:flex; flex-wrap:wrap; gap:8px; }
                .mars-source-header { align-items:center; justify-content:space-between; margin-bottom:10px; }
                .mars-source-header strong { font-size:13px; color:#355b74; }
                .mars-source-link { border:0; background:transparent; color:var(--accent); cursor:pointer; font-size:12px; padding:0; }
                .mars-source-option { display:inline-flex; align-items:center; gap:6px; border:1px solid var(--line-soft); border-radius:999px; padding:5px 10px; background:#f9fbfe; font-size:13px; }
                .mars-clear-button,.mars-mini-button { padding:9px 12px; cursor:pointer; }
                .mars-toggle { display:inline-flex; align-items:center; gap:8px; font-size:13px; }
                .mars-export-block { display:grid; gap:6px; align-content:start; }
                .mars-export-helper { font-size:12px; line-height:1.35; }
                .mars-nav { margin:14px 0 18px 0; }
                .mars-overview-grid { display:grid; grid-template-columns:repeat(auto-fit, minmax(170px, 1fr)); gap:12px; }
                .mars-kpi-card { border:1px solid var(--line-soft); border-radius:14px; background:linear-gradient(180deg,#fbfdff 0%,#f7fbff 100%); padding:14px; }
                .mars-kpi-label { font-size:12px; color:var(--muted); margin-bottom:6px; text-transform:uppercase; letter-spacing:.04em; }
                .mars-kpi-value { font-size:16px; font-weight:700; color:#244258; line-height:1.35; word-break:break-word; }
                .mars-legend { display:flex; flex-wrap:wrap; gap:8px; margin-top:10px; }
                .mars-legend-chip { display:inline-flex; align-items:center; gap:6px; border:1px solid var(--line-soft); border-radius:999px; padding:6px 10px; background:#fff; font-size:12px; color:#436179; }
                .mars-section { margin-bottom:16px; overflow:hidden; }
                .mars-section>summary,.mars-metric-block>summary { cursor:pointer; list-style:none; font-weight:700; }
                .mars-section>summary { padding:16px 18px; background:#f7fbff; border-bottom:1px solid var(--line-soft); }
                .mars-section>summary::-webkit-details-marker,.mars-metric-block>summary::-webkit-details-marker { display:none; }
                .mars-section-body { padding:14px 18px 18px 18px; }
                .mars-section-subtitle { padding:12px 18px 0 18px; font-size:13px; }
                .mars-metric-block { border:1px solid var(--line-soft); border-radius:14px; background:var(--panel-soft); margin-bottom:12px; padding:12px; }
                .mars-metric-block>summary { margin-bottom:10px; color:#355b74; }
                .mars-table-wrap { min-width:0; }
                .mars-table-toolbar { display:grid; grid-template-columns:minmax(240px,360px); gap:6px; margin-bottom:10px; }
                .mars-chart-controls { display:grid; grid-template-columns:minmax(240px,360px) auto; gap:10px; align-items:start; }
                .mars-chart-search { min-width:240px; }
                .mars-summary-filter { border:1px solid var(--line-soft); border-radius:14px; background:#fbfdff; padding:12px; margin-bottom:10px; }
                .mars-summary-filter-label { display:block; margin-bottom:8px; font-size:13px; font-weight:600; color:#355b74; }
                .mars-result-status { min-height:16px; font-size:12px; margin:6px 0 10px 0; }
                .mars-table-scroll { position:relative; overflow:auto; border:1px solid var(--line-soft); border-radius:14px; background:#fff; }
                .mars-data-table { width:max-content; min-width:100%; border-collapse:separate; border-spacing:0; font-size:13px; }
                .mars-th,.mars-td { border-bottom:1px solid var(--line-soft); padding:8px 10px; white-space:nowrap; text-align:left; vertical-align:top; }
                .mars-th { position:sticky; top:0; background:#eef6fb; z-index:1; }
                .mars-td { position:relative; z-index:0; }
                .mars-sticky-col { position:sticky; background-clip:padding-box; overflow:hidden; }
                .mars-feature-col { min-width:var(--mars-feature-col-width, 220px); width:var(--mars-feature-col-width, 220px); max-width:var(--mars-feature-col-width, 220px); box-sizing:border-box; }
                .mars-secondary-col { min-width:var(--mars-secondary-col-width, 110px); width:var(--mars-secondary-col-width, 110px); max-width:var(--mars-secondary-col-width, 110px); box-sizing:border-box; }
                .mars-bin-col { min-width:var(--mars-bin-col-width, 140px); width:var(--mars-bin-col-width, 140px); max-width:var(--mars-bin-col-width, 140px); box-sizing:border-box; }
                .mars-data-table .mars-td.mars-feature-col,
                .mars-data-table .mars-td.mars-secondary-col,
                .mars-pivot-table .mars-td.mars-bin-col { background:#fff; }
                .mars-data-table .mars-th.mars-feature-col,
                .mars-data-table .mars-th.mars-secondary-col,
                .mars-pivot-table .mars-th.mars-bin-col { background:#eef6fb; }
                .mars-data-table .mars-th.mars-feature-col { left:0; z-index:6; box-shadow:2px 0 0 rgba(217,227,235,.85); }
                .mars-data-table .mars-td.mars-feature-col { left:0; z-index:4; box-shadow:2px 0 0 rgba(217,227,235,.85); }
                .mars-data-table .mars-th.mars-secondary-col { left:var(--mars-feature-col-width, 220px); z-index:5; box-shadow:2px 0 0 rgba(217,227,235,.72); }
                .mars-data-table .mars-td.mars-secondary-col { left:var(--mars-feature-col-width, 220px); z-index:3; box-shadow:2px 0 0 rgba(217,227,235,.72); }
                .mars-pivot-table .mars-th.mars-feature-col { left:0; z-index:7; box-shadow:2px 0 0 rgba(217,227,235,.85); }
                .mars-pivot-table .mars-td.mars-feature-col { left:0; z-index:5; box-shadow:2px 0 0 rgba(217,227,235,.85); }
                .mars-pivot-table .mars-th.mars-bin-col { left:var(--mars-feature-col-width, 220px); z-index:6; padding-right:18px; box-shadow:2px 0 0 rgba(217,227,235,.85); }
                .mars-pivot-table .mars-td.mars-bin-col { left:var(--mars-feature-col-width, 220px); z-index:4; box-shadow:2px 0 0 rgba(217,227,235,.85); }
                .mars-th.is-numeric,.mars-td.is-numeric { text-align:right; }
                .mars-sort-button { width:100%; min-width:0; overflow:hidden; border:0; background:transparent; padding:0; margin:0; color:inherit; font:inherit; display:inline-flex; align-items:center; justify-content:space-between; gap:8px; cursor:pointer; }
                .mars-sort-label { display:block; min-width:0; overflow:hidden; text-overflow:ellipsis; }
                .mars-cell-text { display:block; min-width:0; overflow:hidden; text-overflow:ellipsis; white-space:nowrap; }
                .mars-sticky-cell-inner { min-width:0; overflow:hidden; text-overflow:ellipsis; white-space:nowrap; }
                .mars-th.mars-feature-col { padding-right:18px; }
                .mars-resize-handle { position:absolute; top:0; right:0; width:10px; height:100%; cursor:col-resize; user-select:none; touch-action:none; }
                .mars-resize-handle::after { content:""; position:absolute; top:20%; bottom:20%; left:4px; width:2px; border-radius:2px; background:rgba(53,91,116,.22); }
                .mars-feature-jump { min-width:240px; }
                .mars-pivot-table .mars-th, .mars-pivot-table .mars-td { background-clip:padding-box; }
                .mars-jump-highlight { animation:mars-jump-pulse 1.2s ease-out 1; }
                .mars-jump-highlight-cell { animation:mars-jump-pulse 1.2s ease-out 1; }
                .mars-table-ownership-sentinel { height:0; margin:0; padding:0; pointer-events:none; }
                .mars-floating-header-host { position:fixed; top:0; left:0; width:0; display:none; border:0; border-radius:14px; background:#fff; box-shadow:0 14px 32px rgba(32,48,64,.16), inset 0 0 0 1px var(--line-soft); overflow:hidden; z-index:60; }
                .mars-floating-header-host.is-visible { display:block; }
                .mars-floating-header-scroll { overflow:hidden; background:#fff; }
                .mars-floating-header-table { width:max-content; min-width:100%; margin:0; table-layout:fixed; }
                .mars-floating-header-table tbody { display:none; }
                .mars-floating-header-table .mars-th { top:0; z-index:8; }
                .mars-floating-header-table .mars-th.mars-feature-col { z-index:10; }
                .mars-floating-header-table .mars-th.mars-secondary-col,
                .mars-floating-header-table .mars-th.mars-bin-col { z-index:9; }
                .mars-back-to-top { position:fixed; right:24px; bottom:24px; border:1px solid rgba(53,91,116,.18); border-radius:999px; background:rgba(255,255,255,.96); color:#355b74; box-shadow:0 14px 28px rgba(32,48,64,.14); padding:11px 16px; font-size:13px; font-weight:600; cursor:pointer; opacity:0; transform:translateY(12px); pointer-events:none; transition:opacity .18s ease, transform .18s ease, box-shadow .18s ease; z-index:70; }
                .mars-back-to-top.is-visible { opacity:1; transform:translateY(0); pointer-events:auto; }
                .mars-back-to-top:hover,.mars-back-to-top:focus-visible { box-shadow:0 18px 34px rgba(32,48,64,.2); outline:none; }
                .mars-data-table tbody tr.mars-jump-highlight > .mars-td { position:relative; z-index:2; filter:saturate(1.08) brightness(.98); box-shadow:inset 0 0 0 9999px rgba(255,237,177,.34), inset 0 2px 0 rgba(233,153,49,.86), inset 0 -2px 0 rgba(233,153,49,.86) !important; transition:box-shadow .32s ease, filter .32s ease, outline-color .32s ease; }
                .mars-data-table tbody tr.mars-jump-highlight > .mars-td:first-child { border-left:3px solid rgba(233,153,49,.86); }
                .mars-data-table tbody tr.mars-jump-highlight > .mars-td:last-child { border-right:3px solid rgba(233,153,49,.86); }
                .mars-data-table tbody tr.mars-jump-highlight > .mars-td .mars-cell-text { font-weight:600; }
                .mars-data-table tbody tr.mars-jump-highlight > .mars-td.mars-feature-col { color:#122636; }
                .mars-data-table tbody tr.mars-jump-highlight > .mars-td.mars-jump-highlight-cell { outline:2px solid rgba(245,158,11,.48); outline-offset:-2px; box-shadow:inset 0 0 0 9999px rgba(255,247,213,.82), inset 6px 0 0 #f59e0b, inset 0 2px 0 rgba(233,153,49,.9), inset 0 -2px 0 rgba(233,153,49,.9) !important; filter:saturate(1.12) brightness(1); }
                .mars-data-table tbody tr.mars-jump-highlight > .mars-td.mars-jump-highlight-cell .mars-cell-text { font-weight:700; color:#0f2131; text-shadow:0 1px 0 rgba(255,255,255,.55); }
                .mars-sort-indicator::before { content:"\\2195"; color:#8aa1b3; font-size:11px; }
                th[data-sort-dir="asc"] .mars-sort-indicator::before { content:"\\2191"; color:var(--accent); }
                th[data-sort-dir="desc"] .mars-sort-indicator::before { content:"\\2193"; color:var(--accent); }
                .mars-empty { border:1px dashed var(--line); border-radius:14px; padding:16px; background:#fbfdff; font-size:13px; }
                .mars-scope-empty { margin-top:10px; }
                .mars-scope-empty[hidden] { display:none !important; }
                .mars-chart-card { border:1px solid var(--line-soft); border-radius:14px; background:#fff; padding:12px; margin-bottom:12px; box-shadow:0 8px 20px rgba(51,82,108,.05); }
                .mars-pivot-source-title-cell { background:#edf6fb; color:#355b74; font-weight:700; letter-spacing:.02em; }
                .mars-pivot-feature { font-weight:600; color:#2f495e; }
                .mars-pivot-feature-blank .mars-cell-text { visibility:hidden; }
                .mars-pivot-spacer-row td { border-bottom:0; padding:5px 0; background:linear-gradient(180deg,transparent 0%,rgba(233,239,245,.65) 100%); }
                .mars-chart-card h4 { margin:0 0 10px 0; font-size:16px; }
                .mars-footnote { font-size:12px; margin-top:12px; }
        """.strip()

    @classmethod
    def _build_html_runtime_script(cls, *, summary_filter_columns: List[str]) -> str:
        template = """
                const marsSummaryFilterColumns = new Set(__SUMMARY_FILTER_COLUMNS__);
                const marsState = {
                    globalQuery:"",
                    regexMode:false,
                    localQueries:{},
                    selectedSources:[],
                    appliedSummaryExpression:"",
                    summaryAllowedFeatures:null,
                    refreshScheduled:false,
                    refreshFrameId:null,
                    postPaintFrameId:null,
                    refreshTimerId:null,
                    pendingRefreshTokens:[],
                    pendingLayoutToken:null,
                    layoutFrameId:null,
                    resizeState:null,
                    resizeFrameScheduled:false,
                    floatingHeaderTableId:"",
                    floatingHeaderScrollBox:null,
                    floatingHeaderFrameId:null,
                    jumpHighlightTimerId:null,
                    jumpHighlightArmTimerId:null,
                    jumpHighlightNode:null,
                    jumpHighlightCell:null
                };
                function marsBuildMatcher(query) { const q=(query||"").trim(); if(!q) return {ok:true,match:()=>true}; if(marsState.regexMode) { try { const regex=new RegExp(q,"i"); return {ok:true,match:(text)=>regex.test(text||"")}; } catch(err) { return {ok:false,error:err.message}; } } const terms=q.toLowerCase().split(/\\s+/).filter(Boolean); return {ok:true,match:(text)=>terms.every((term)=>(text||"").toLowerCase().includes(term))}; }
                function marsSetError(id, message) { const node=document.getElementById(id); if(node) node.textContent=message||""; }
                function marsNormalizeFeatureValue(value) { return (value||"").trim().toLowerCase(); }
                function marsResolveLocalScope(scopeId) { return scopeId==="mars-chart-cards" ? "charts" : `table:${scopeId}`; }
                function marsMergeRefreshToken(scopeToken) {
                    const token=(scopeToken||"all").trim() || "all";
                    if(token==="all") { marsState.pendingRefreshTokens=["all"]; return; }
                    if(marsState.pendingRefreshTokens.includes("all")) return;
                    if(!marsState.pendingRefreshTokens.includes(token)) marsState.pendingRefreshTokens.push(token);
                }
                function marsMergeLayoutToken(scopeToken) {
                    const token=(scopeToken||"all").trim() || "all";
                    if(token==="all" || marsState.pendingLayoutToken==="all" || !marsState.pendingLayoutToken) {
                        marsState.pendingLayoutToken = token==="all" ? "all" : marsState.pendingLayoutToken || token;
                        return;
                    }
                    if(marsState.pendingLayoutToken!==token) marsState.pendingLayoutToken="all";
                }
                function marsQueueRefresh(scopeToken="all", delayMs=0) {
                    marsMergeRefreshToken(scopeToken);
                    if(marsState.refreshFrameId) window.cancelAnimationFrame(marsState.refreshFrameId);
                    if(marsState.postPaintFrameId) window.cancelAnimationFrame(marsState.postPaintFrameId);
                    if(marsState.refreshTimerId) window.clearTimeout(marsState.refreshTimerId);
                    marsState.refreshScheduled = true;
                    marsState.refreshFrameId = window.requestAnimationFrame(() => {
                        marsState.refreshFrameId = null;
                        marsState.postPaintFrameId = window.requestAnimationFrame(() => {
                            marsState.postPaintFrameId = null;
                            marsState.refreshTimerId = window.setTimeout(() => {
                                marsState.refreshTimerId = null;
                                marsState.refreshScheduled = false;
                                marsFlushRefreshQueue();
                            }, Math.max(0, Number(delayMs) || 0));
                        });
                    });
                }
                function marsQueueTextRefresh(scopeToken="all") { marsQueueRefresh(scopeToken, 80); }
                function marsQueueLayoutSync(scopeToken="all") {
                    marsMergeLayoutToken(scopeToken);
                    if(marsState.layoutFrameId) return;
                    marsState.layoutFrameId = window.requestAnimationFrame(() => {
                        marsState.layoutFrameId = null;
                        const token = marsState.pendingLayoutToken || "all";
                        marsState.pendingLayoutToken = null;
                        marsSyncScopeLayouts(token);
                    });
                }
                function marsSetGlobalQuery(value) { marsState.globalQuery=value||""; marsQueueTextRefresh("all"); }
                function marsSetLocalQuery(scopeId, value) { marsState.localQueries[scopeId]=value||""; marsQueueTextRefresh(marsResolveLocalScope(scopeId)); }
                function marsSetRegexMode(enabled) { marsState.regexMode=!!enabled; marsQueueRefresh("all"); }
                function marsSetDataSources() { const boxes=Array.from(document.querySelectorAll(".mars-source-checkbox")); marsState.selectedSources=boxes.filter((box)=>box.checked).map((box)=>box.value); marsQueueRefresh("all"); }
                function marsHandleDataSourceToggle() { marsSetDataSources(); }
                function marsHandlePivotTargetChange() { marsQueueRefresh("pivot"); marsQueueLayoutSync("pivot"); }
                function marsHandleChartTargetChange() { marsQueueRefresh("charts"); }
                function marsSelectAllSources() { document.querySelectorAll(".mars-source-checkbox").forEach((box)=>{ box.checked=true; }); marsSetDataSources(); }
                function marsClearSources() { document.querySelectorAll(".mars-source-checkbox").forEach((box)=>{ box.checked=false; }); marsSetDataSources(); }
                function marsClearGlobalSearch() { const input=document.getElementById("mars-global-search"); if(input) input.value=""; marsState.globalQuery=""; marsQueueTextRefresh("all"); }
                function marsTokenizeSummaryExpression(expr) {
                    const text=(expr||"").trim();
                    if(!text) return {ok:true,tokens:[]};
                    const tokenPattern=/\\s*(>=|<=|==|!=|>|<|\\&|\\||\\(|\\)|-?(?:\\d+\\.\\d*|\\d*\\.\\d+|\\d+)(?:[eE][+-]?\\d+)?|[A-Za-z_][A-Za-z0-9_]*)\\s*/gy;
                    const tokens=[];
                    let cursor=0;
                    while(cursor < text.length) {
                        tokenPattern.lastIndex = cursor;
                        const match=tokenPattern.exec(text);
                        if(!match) return {ok:false,error:"Invalid expression syntax."};
                        tokens.push(match[1]);
                        if(tokenPattern.lastIndex<=cursor) return {ok:false,error:"Invalid expression syntax."};
                        cursor=tokenPattern.lastIndex;
                    }
                    return {ok:true,tokens};
                }
                function marsParseSummaryExpression(expr) {
                    const tokenResult=marsTokenizeSummaryExpression(expr);
                    if(!tokenResult.ok) return tokenResult;
                    const tokens=tokenResult.tokens;
                    if(!tokens.length) return {ok:true,ast:null};
                    let idx=0;
                    function peek() { return tokens[idx]; }
                    function consume(expected) {
                        const token=tokens[idx];
                        if(expected && token!==expected) throw new Error(`Expected '${expected}'`);
                        idx+=1;
                        return token;
                    }
                    function parsePrimary() {
                        const token=peek();
                        if(token===undefined) throw new Error("Unexpected end of expression.");
                        if(token==="(") { consume("("); const node=parseOr(); if(peek()!==")") throw new Error("Missing closing parenthesis."); consume(")"); return node; }
                        if(/^-?(?:\\d+\\.\\d*|\\d*\\.\\d+|\\d+)(?:[eE][+-]?\\d+)?$/.test(token)) { consume(); return {type:"number", value:Number(token)}; }
                        if(/^[A-Za-z_][A-Za-z0-9_]*$/.test(token)) {
                            if(!marsSummaryFilterColumns.has(token)) throw new Error(`Unknown metric: ${token}`);
                            consume();
                            return {type:"identifier", value:token};
                        }
                        throw new Error(`Unexpected token: ${token}`);
                    }
                    function parseComparison() {
                        const left=parsePrimary();
                        const token=peek();
                        if(["<", "<=", ">", ">=", "==", "!="].includes(token)) {
                            consume();
                            const right=parsePrimary();
                            return {type:"compare", op:token, left, right};
                        }
                        if(!["identifier", "compare", "and", "or"].includes(left.type)) throw new Error("Standalone values must be metric names.");
                        return left;
                    }
                    function parseAnd() {
                        let node=parseComparison();
                        while(peek()==="&") { consume("&"); node={type:"and", left:node, right:parseComparison()}; }
                        return node;
                    }
                    function parseOr() {
                        let node=parseAnd();
                        while(peek()==="|") { consume("|"); node={type:"or", left:node, right:parseAnd()}; }
                        return node;
                    }
                    try {
                        const ast=parseOr();
                        if(idx!==tokens.length) throw new Error(`Unexpected token: ${peek()}`);
                        return {ok:true,ast};
                    } catch(err) {
                        return {ok:false,error:err.message};
                    }
                }
                function marsEvaluateSummaryNode(node, metrics) {
                    if(!node) return true;
                    if(node.type==="number") return node.value;
                    if(node.type==="identifier") return Number(metrics?.[node.value]);
                    if(node.type==="compare") {
                        const left=Number(marsEvaluateSummaryNode(node.left, metrics));
                        const right=Number(marsEvaluateSummaryNode(node.right, metrics));
                        if(!Number.isFinite(left) || !Number.isFinite(right)) return false;
                        return node.op===">" ? left>right : node.op===">=" ? left>=right : node.op==="<" ? left<right : node.op==="<=" ? left<=right : node.op==="==" ? left===right : left!==right;
                    }
                    if(node.type==="and") return Boolean(marsEvaluateSummaryNode(node.left, metrics)) && Boolean(marsEvaluateSummaryNode(node.right, metrics));
                    if(node.type==="or") return Boolean(marsEvaluateSummaryNode(node.left, metrics)) || Boolean(marsEvaluateSummaryNode(node.right, metrics));
                    return false;
                }
                function marsSetSummaryExpression(value) {
                    const expr=(value||"").trim();
                    if(!expr) {
                        marsState.appliedSummaryExpression="";
                        marsSetError("mars-summary-expression-error", "");
                        marsQueueTextRefresh("all");
                        return;
                    }
                    const parsed=marsParseSummaryExpression(expr);
                    if(!parsed.ok) {
                        marsSetError("mars-summary-expression-error", parsed.error || "Invalid expression.");
                        marsQueueTextRefresh("all");
                        return;
                    }
                    marsState.appliedSummaryExpression=expr;
                    marsSetError("mars-summary-expression-error", "");
                    marsQueueTextRefresh("all");
                }
                function marsUpdateTableSpecialRows(table) { const rows=Array.from(table.querySelectorAll("tbody tr")); const visibleBySource=new Set(); const visibleByFeatureSource=new Set(); rows.forEach((row)=>{ const role=row.dataset.role||"data"; if(role==="data"&&row.style.display!=="none") { const source=row.dataset.dataSource||""; const feature=row.dataset.feature||""; visibleBySource.add(source); visibleByFeatureSource.add(`${source}||${feature}`); } }); rows.forEach((row)=>{ const role=row.dataset.role||"data"; if(role==="source") { row.style.display=visibleBySource.has(row.dataset.dataSource||"")?"":"none"; } else if(role==="spacer") { const key=`${row.dataset.dataSource||""}||${row.dataset.feature||""}`; row.style.display=visibleByFeatureSource.has(key)?"":"none"; } }); }
                function marsSourceSelected(source) { if(source==="__aggregate__") return true; const hasBoxes=document.querySelectorAll(".mars-source-checkbox").length>0; if(!hasBoxes) return true; return marsState.selectedSources.includes(source||"UNMAPPED"); }
                function marsReadRowMetrics(row) { let metrics={}; try { metrics=JSON.parse(row.dataset.metrics||"{}"); } catch(err) { metrics={}; } return metrics; }
                function marsSummaryRowAllowedWithoutLocal(row, globalMatcher=null, summaryParsed=null) {
                    if(!row) return false;
                    const matcher=globalMatcher||marsBuildMatcher(marsState.globalQuery);
                    if(!matcher.ok) return false;
                    const parsed=summaryParsed||marsParseSummaryExpression(marsState.appliedSummaryExpression);
                    if(!parsed.ok) return false;
                    const source=row.dataset.dataSource||"UNMAPPED";
                    const text=row.dataset.searchText||row.textContent||"";
                    return marsSourceSelected(source) && matcher.match(text) && marsEvaluateSummaryNode(parsed.ast, marsReadRowMetrics(row));
                }
                function marsGetSummaryFeatureAllowSet() {
                    const table=document.getElementById("mars-summary-table");
                    if(!table) return null;
                    const globalMatcher=marsBuildMatcher(marsState.globalQuery);
                    if(!globalMatcher.ok) return null;
                    const parsed=marsParseSummaryExpression(marsState.appliedSummaryExpression);
                    if(!parsed.ok) return marsState.summaryAllowedFeatures;
                    const features=new Set();
                    table.querySelectorAll("tbody tr[data-feature]").forEach((row)=>{
                        const feature=row.dataset.feature||"";
                        if(feature && marsSummaryRowAllowedWithoutLocal(row, globalMatcher, parsed)) features.add(feature);
                    });
                    return features;
                }
                function marsFeatureAllowed(feature) { if(!(marsState.summaryAllowedFeatures instanceof Set)) return true; return marsState.summaryAllowedFeatures.has(feature||""); }
                function marsSetScopeStatus(scopeId, visibleCount, totalCount, noun) {
                    const node=document.getElementById(`${scopeId}-status`);
                    if(!node) return;
                    const visible=Math.max(0, Number(visibleCount) || 0);
                    const total=Math.max(0, Number(totalCount) || 0);
                    if(total===0 || visible===0) { node.textContent=`0 ${noun} matched current filters.`; return; }
                    if(visible===total) { node.textContent=`${visible} ${noun} shown.`; return; }
                    node.textContent=`${visible} of ${total} ${noun} shown.`;
                }
                function marsToggleScopeEmpty(scopeId, visible) {
                    const node=document.getElementById(`${scopeId}-empty`);
                    if(node) node.hidden=!visible;
                }
                function marsUpdateTableFeedback(tableId, totalCount, visibleCount) {
                    marsSetScopeStatus(tableId, visibleCount, totalCount, "rows");
                    marsToggleScopeEmpty(tableId, visibleCount===0);
                }
                function marsApplyTableFilter(tableId) {
                    const table=document.getElementById(tableId);
                    if(!table) return;
                    const globalMatcher=marsBuildMatcher(marsState.globalQuery);
                    if(!globalMatcher.ok) { marsSetError("mars-global-error", `Invalid regex: ${globalMatcher.error}`); return; }
                    marsSetError("mars-global-error", "");
                    const localMatcher=marsBuildMatcher(marsState.localQueries[tableId]||"");
                    if(!localMatcher.ok) { marsSetError(`${tableId}-error`, `Invalid regex: ${localMatcher.error}`); return; }
                    marsSetError(`${tableId}-error`, "");
                    const summaryParsed=marsParseSummaryExpression(marsState.appliedSummaryExpression);
                    const isSummary=table.dataset.tableKind==="summary";
                    const dataRows = Array.from(table.querySelectorAll("tbody tr")).filter((row)=>(row.dataset.role||"data")==="data");
                    dataRows.forEach((row)=>{
                        const source=row.dataset.dataSource||"UNMAPPED";
                        const feature=row.dataset.feature||"";
                        const text=row.dataset.searchText||row.textContent||"";
                        const globalVisible=marsSourceSelected(source)&&globalMatcher.match(text);
                        if(!globalVisible) { row.style.display="none"; return; }
                        const summaryVisible=isSummary
                            ? (summaryParsed.ok ? marsSummaryRowAllowedWithoutLocal(row, globalMatcher, summaryParsed) : true)
                            : marsFeatureAllowed(feature);
                        const visible=summaryVisible&&localMatcher.match(text);
                        row.style.display=visible?"":"none";
                    });
                    marsUpdateTableSpecialRows(table);
                    const visibleCount = dataRows.filter((row)=>row.style.display!=="none").length;
                    marsUpdateTableFeedback(tableId, dataRows.length, visibleCount);
                }
                function marsSortTable(tableId, trigger) { const table=document.getElementById(tableId); if(!table) return; const th=trigger.closest("th"); const colIndex=Number(th.dataset.colIndex||Array.from(th.parentNode.children).indexOf(th)); if(colIndex<0) return; const sourceHeader=table.querySelector(`thead th[data-col-index="${colIndex}"]`) || th; const tbody=table.querySelector("tbody"); const rows=Array.from(tbody.querySelectorAll("tr")).filter((row)=>(row.dataset.role||"data")==="data"); let nextDir="asc"; if(table.dataset.sortCol===String(colIndex)) nextDir=table.dataset.sortDir==="asc"?"desc":"asc"; const sortType=sourceHeader.dataset.sortType||th.dataset.sortType||"text"; rows.sort((a,b)=>{ const va=a.children[colIndex]?.dataset.sortValue||""; const vb=b.children[colIndex]?.dataset.sortValue||""; if(sortType==="number") { const na=Number(va), nb=Number(vb); const sa=Number.isFinite(na)?na:(nextDir==="asc"?Infinity:-Infinity); const sb=Number.isFinite(nb)?nb:(nextDir==="asc"?Infinity:-Infinity); return nextDir==="asc"?sa-sb:sb-sa; } return nextDir==="asc"?va.localeCompare(vb,undefined,{numeric:true,sensitivity:"base"}):vb.localeCompare(va,undefined,{numeric:true,sensitivity:"base"}); }); rows.forEach((row)=>tbody.appendChild(row)); table.dataset.sortCol=String(colIndex); table.dataset.sortDir=nextDir; table.querySelectorAll("thead th[data-sort-dir]").forEach((cell)=>cell.removeAttribute("data-sort-dir")); sourceHeader.dataset.sortDir=nextDir; marsApplyTableFilter(tableId); marsQueueLayoutSync(`table:${tableId}`); marsScheduleViewportRefresh(); }
                function marsUpdatePivotViews() {
                    const targetValue=document.getElementById("mars-pivot-target")?.value||null;
                    document.querySelectorAll(".mars-pivot-view").forEach((view)=>{
                        const sameTarget=!targetValue||view.dataset.yValue===targetValue;
                        view.style.display=sameTarget?"":"none";
                    });
                }
                function marsUpdateChartViews() {
                    const targetValue=document.getElementById("mars-chart-target")?.value||null;
                    const globalMatcher=marsBuildMatcher(marsState.globalQuery);
                    const localMatcher=marsBuildMatcher(marsState.localQueries["mars-chart-cards"]||"");
                    if(!globalMatcher.ok) { marsSetError("mars-global-error", `Invalid regex: ${globalMatcher.error}`); return; }
                    marsSetError("mars-global-error", "");
                    if(!localMatcher.ok) { marsSetError("mars-chart-cards-error", `Invalid regex: ${localMatcher.error}`); return; }
                    marsSetError("mars-chart-cards-error", "");
                    let totalCards=0;
                    let visibleCards=0;
                    document.querySelectorAll(".mars-chart-view").forEach((view)=>{
                        const visibleTarget=!targetValue||view.dataset.yValue===targetValue;
                        view.style.display=visibleTarget?"":"none";
                        if(!visibleTarget) return;
                        view.querySelectorAll(".mars-chart-card").forEach((card)=>{
                            totalCards += 1;
                            const source=card.dataset.dataSource||"UNMAPPED";
                            const feature=card.dataset.feature||"";
                            const text=card.dataset.searchText||card.textContent||"";
                            const globalVisible=marsSourceSelected(source)&&globalMatcher.match(text)&&marsFeatureAllowed(feature);
                            const visible=globalVisible&&localMatcher.match(text);
                            card.style.display=visible?"":"none";
                            if(visible) visibleCards += 1;
                        });
                    });
                    marsSetScopeStatus("mars-chart-cards", visibleCards, totalCards, "charts");
                    marsToggleScopeEmpty("mars-chart-cards", visibleCards===0);
                }
                function marsBuildExportFeatureMap() {
                    const table=document.getElementById("mars-summary-table");
                    if(!table) return {};
                    const summaryParsed=marsParseSummaryExpression(marsState.appliedSummaryExpression);
                    const featureMap=new Map();
                    const sourceOrder=Array.from(document.querySelectorAll(".mars-source-checkbox")).map((box)=>box.value);
                    table.querySelectorAll("tbody tr[data-feature]").forEach((row)=>{
                        const source=row.dataset.dataSource||"UNMAPPED";
                        const feature=row.dataset.feature||"";
                        if(!feature || !marsSourceSelected(source)) return;
                        if(summaryParsed.ok && !marsEvaluateSummaryNode(summaryParsed.ast, marsReadRowMetrics(row))) return;
                        if(!featureMap.has(source)) featureMap.set(source, new Set());
                        featureMap.get(source).add(feature);
                    });
                    const payload={};
                    const assignedSources=new Set();
                    sourceOrder.forEach((source)=>{
                        const values=featureMap.has(source) ? Array.from(featureMap.get(source)).sort((a,b)=>a.localeCompare(b, undefined, {numeric:true, sensitivity:"base"})) : [];
                        if(values.length) {
                            payload[source]=values;
                            assignedSources.add(source);
                        }
                    });
                    featureMap.forEach((features, source)=>{
                        if(assignedSources.has(source)) return;
                        const values=Array.from(features).sort((a,b)=>a.localeCompare(b, undefined, {numeric:true, sensitivity:"base"}));
                        if(values.length) payload[source]=values;
                    });
                    return payload;
                }
                function marsDownloadTextFile(text, fileName) { const blob=new Blob([text], {type:"text/plain;charset=utf-8"}); const link=document.createElement("a"); link.href=URL.createObjectURL(blob); link.download=fileName; link.click(); URL.revokeObjectURL(link.href); }
                function marsExportFeatures() { const featureMap=marsBuildExportFeatureMap(); marsDownloadTextFile(JSON.stringify(featureMap, null, 2), "mars_features.txt"); }
                function marsGetFloatingHeaderHost() { return document.getElementById("mars-floating-header-host"); }
                function marsGetFloatingHeaderScroll() { return document.getElementById("mars-floating-header-scroll"); }
                function marsGetTableScrollBox(table) { return table?.closest(".mars-table-scroll") || null; }
                function marsAncestorsDetailsOpen(node) {
                    let parent=node?.closest("details");
                    while(parent) {
                        if(!parent.open) return false;
                        parent=parent.parentElement?.closest("details");
                    }
                    return true;
                }
                function marsHasClientRects(node) {
                    return Boolean(node?.getClientRects && node.getClientRects().length);
                }
                function marsIntersectsViewport(rect) {
                    return rect.width > 0 && rect.height > 0 && rect.bottom > 0 && rect.top < window.innerHeight;
                }
                function marsTableIsActuallyVisible(table, scrollBox, thead) {
                    if(!table || !scrollBox || !thead) return false;
                    if(!marsAncestorsDetailsOpen(scrollBox)) return false;
                    if(!marsHasClientRects(scrollBox) || !marsHasClientRects(table) || !marsHasClientRects(thead)) return false;
                    const scrollRect=scrollBox.getBoundingClientRect();
                    const tableRect=table.getBoundingClientRect();
                    const theadRect=thead.getBoundingClientRect();
                    if(scrollRect.width <= 0 || scrollRect.height <= 0 || tableRect.width <= 0 || tableRect.height <= 0 || theadRect.width <= 0 || theadRect.height <= 0) return false;
                    return marsIntersectsViewport(scrollRect) && marsIntersectsViewport(tableRect);
                }
                function marsHideFloatingHeader() {
                    const host=marsGetFloatingHeaderHost();
                    const scrollHost=marsGetFloatingHeaderScroll();
                    if(scrollHost) scrollHost.innerHTML="";
                    if(host) {
                        host.hidden=true;
                        host.classList.remove("is-visible");
                        host.style.left="0px";
                        host.style.width="0px";
                        host.removeAttribute("data-table-id");
                    }
                    marsState.floatingHeaderTableId="";
                    marsState.floatingHeaderScrollBox=null;
                }
                function marsGetFirstVisibleDataRowTop(table) {
                    const rows=Array.from(table?.querySelectorAll("tbody tr") || []).filter((row) => {
                        if(row.offsetParent===null || row.style.display==="none") return false;
                        return (row.dataset.role || "data")==="data";
                    });
                    for(const row of rows) {
                        const rect=row.getBoundingClientRect();
                        if(rect.bottom > 0) return rect.top;
                    }
                    if(rows.length) return rows[0].getBoundingClientRect().top;
                    const tbody=table?.querySelector("tbody");
                    if(tbody) return tbody.getBoundingClientRect().top;
                    return table?.getBoundingClientRect().top ?? Number.POSITIVE_INFINITY;
                }
                function marsCollectLeafColumnWidths(table) {
                    const thead=table?.querySelector("thead");
                    const rows=Array.from(thead?.rows || []);
                    if(!rows.length) return [];
                    const occupancy=[];
                    const leafColumns=[];
                    const totalRows=rows.length;
                    rows.forEach((row, rowIndex) => {
                        occupancy[rowIndex] = occupancy[rowIndex] || [];
                        let colIndex=0;
                        Array.from(row.cells).forEach((cell) => {
                            while(occupancy[rowIndex][colIndex]) colIndex += 1;
                            const colSpan=Math.max(1, Number(cell.colSpan) || 1);
                            const rowSpan=Math.max(1, Number(cell.rowSpan) || 1);
                            for(let r=rowIndex; r<Math.min(totalRows, rowIndex + rowSpan); r += 1) {
                                occupancy[r] = occupancy[r] || [];
                                for(let c=colIndex; c<colIndex + colSpan; c += 1) occupancy[r][c]=true;
                            }
                            if(rowIndex + rowSpan >= totalRows) {
                                const baseWidth=Math.max(1, Number(cell.getBoundingClientRect().width || cell.offsetWidth || 0));
                                const sharedWidth=baseWidth / colSpan;
                                for(let c=0; c<colSpan; c += 1) {
                                    leafColumns[colIndex + c] = Math.max(1, Math.ceil(sharedWidth));
                                }
                            }
                            colIndex += colSpan;
                        });
                    });
                    return leafColumns.filter((width)=>Number.isFinite(width) && width > 0);
                }
                function marsBuildFloatingHeaderColGroup(table) {
                    const widths=marsCollectLeafColumnWidths(table);
                    if(!widths.length) return null;
                    const colgroup=document.createElement("colgroup");
                    widths.forEach((width) => {
                        const col=document.createElement("col");
                        col.style.width=`${width}px`;
                        col.style.minWidth=`${width}px`;
                        col.style.maxWidth=`${width}px`;
                        colgroup.appendChild(col);
                    });
                    return colgroup;
                }
                function marsCloneFloatingHeader(table) {
                    const host=marsGetFloatingHeaderHost();
                    const scrollHost=marsGetFloatingHeaderScroll();
                    const sourceScrollBox=marsGetTableScrollBox(table);
                    const thead=table?.querySelector("thead");
                    const colgroup=marsBuildFloatingHeaderColGroup(table);
                    if(!host || !scrollHost || !sourceScrollBox || !thead || !colgroup) {
                        marsHideFloatingHeader();
                        return;
                    }
                    const cloneTable=document.createElement("table");
                    cloneTable.className=`${table.className} mars-floating-header-table`;
                    cloneTable.setAttribute("aria-hidden", "true");
                    const inlineStyle=table.getAttribute("style");
                    if(inlineStyle) cloneTable.setAttribute("style", inlineStyle);
                    cloneTable.appendChild(colgroup);
                    cloneTable.appendChild(thead.cloneNode(true));
                    scrollHost.innerHTML="";
                    scrollHost.appendChild(cloneTable);
                    host.hidden=false;
                    host.classList.add("is-visible");
                    host.dataset.tableId=table.id;
                    marsState.floatingHeaderTableId=table.id;
                    marsState.floatingHeaderScrollBox=sourceScrollBox;
                    marsSyncFloatingHeaderMetrics(table);
                }
                function marsSyncFloatingHeaderMetrics(table) {
                    const host=marsGetFloatingHeaderHost();
                    const scrollHost=marsGetFloatingHeaderScroll();
                    const sourceScrollBox=marsGetTableScrollBox(table);
                    const cloneTable=scrollHost?.querySelector("table");
                    const thead=table?.querySelector("thead");
                    if(!host || !scrollHost || !sourceScrollBox || !cloneTable || !thead) {
                        marsHideFloatingHeader();
                        return;
                    }
                    const scrollRect=sourceScrollBox.getBoundingClientRect();
                    const headerRect=thead.getBoundingClientRect();
                    if(scrollRect.width <= 0 || headerRect.height <= 0) {
                        marsHideFloatingHeader();
                        return;
                    }
                    const colgroup=marsBuildFloatingHeaderColGroup(table);
                    if(!colgroup) {
                        marsHideFloatingHeader();
                        return;
                    }
                    const existingColgroup=cloneTable.querySelector("colgroup");
                    if(existingColgroup) cloneTable.replaceChild(colgroup, existingColgroup);
                    else cloneTable.insertBefore(colgroup, cloneTable.firstChild);
                    const contentLeft=scrollRect.left + (sourceScrollBox.clientLeft || 0);
                    const visibleWidth=Math.max(0, Math.ceil(sourceScrollBox.clientWidth || scrollRect.width || 0));
                    host.style.left=`${Math.max(0, contentLeft)}px`;
                    host.style.width=`${visibleWidth}px`;
                    host.style.top="0px";
                    scrollHost.style.height=`${Math.ceil(headerRect.height)}px`;
                    const inlineStyle=table.getAttribute("style");
                    if(inlineStyle) cloneTable.setAttribute("style", inlineStyle);
                    const tableWidth=Math.max(
                        colgroup.childElementCount
                            ? Array.from(colgroup.children).reduce((sum, col) => sum + (parseFloat(col.style.width) || 0), 0)
                            : 0,
                        Math.ceil(table.scrollWidth || 0),
                        Math.ceil(table.getBoundingClientRect().width || 0),
                    );
                    cloneTable.style.width=`${tableWidth}px`;
                    cloneTable.style.minWidth=`${tableWidth}px`;
                    cloneTable.style.maxWidth=`${tableWidth}px`;
                    scrollHost.scrollLeft=sourceScrollBox.scrollLeft;
                }
                function marsResolveFloatingHeaderOwner() {
                    const visibleTables=[];
                    document.querySelectorAll(".mars-table-scroll[data-table-id]").forEach((scrollBox) => {
                        const table=scrollBox.querySelector("table.mars-data-table[id]");
                        const thead=table?.querySelector("thead");
                        if(!marsTableIsActuallyVisible(table, scrollBox, thead)) return;
                        const theadRect=thead.getBoundingClientRect();
                        const tableRect=table.getBoundingClientRect();
                        visibleTables.push({
                            table,
                            scrollBox,
                            theadTop:theadRect.top,
                            headerHeight:Math.max(1, Math.ceil(theadRect.height || 0)),
                            tableBottom:tableRect.bottom,
                            firstDataRowTop:marsGetFirstVisibleDataRowTop(table),
                        });
                    });
                    if(!visibleTables.length) return null;
                    const hostHeight=Math.ceil(marsGetFloatingHeaderHost()?.getBoundingClientRect().height || 0);
                    const hasVisibleReadingTable=visibleTables.some(({ theadTop, firstDataRowTop, tableBottom, headerHeight }) => {
                        const readingBandBottom=Math.max(1, hostHeight || headerHeight);
                        return tableBottom > 0 && (theadTop <= readingBandBottom || firstDataRowTop <= readingBandBottom);
                    });
                    if(!hasVisibleReadingTable) return null;
                    const ownerCandidates=visibleTables.filter(({ theadTop, tableBottom, headerHeight }) => {
                        const floatingHeaderHeight=Math.max(1, hostHeight || headerHeight);
                        return theadTop <= 0 && tableBottom > floatingHeaderHeight;
                    });
                    if(!ownerCandidates.length) return null;
                    ownerCandidates.sort((a,b)=>b.theadTop-a.theadTop);
                    const owner=ownerCandidates[0];
                    const readingLine=Math.max(1, hostHeight || owner.headerHeight) + 1;
                    const shouldReleaseOwner=visibleTables.some((item) => {
                        if(item.table.id===owner.table.id) return false;
                        return item.theadTop > 0 && item.firstDataRowTop <= readingLine;
                    });
                    if(shouldReleaseOwner) return null;
                    return owner;
                }
                function marsRefreshFloatingHeader() {
                    const candidate=marsResolveFloatingHeaderOwner();
                    if(!candidate) {
                        marsHideFloatingHeader();
                        return;
                    }
                    if(marsState.floatingHeaderTableId!==candidate.table.id) {
                        marsCloneFloatingHeader(candidate.table);
                        return;
                    }
                    marsState.floatingHeaderScrollBox=candidate.scrollBox;
                    marsSyncFloatingHeaderMetrics(candidate.table);
                }
                function marsScheduleViewportRefresh() {
                    if(marsState.floatingHeaderFrameId) return;
                    marsState.floatingHeaderFrameId=window.requestAnimationFrame(() => {
                        marsState.floatingHeaderFrameId=null;
                        marsRefreshFloatingHeader();
                        marsUpdateBackToTopVisibility();
                    });
                }
                function marsHandleTableHorizontalScroll(event) {
                    const scrollBox=event.currentTarget;
                    if(scrollBox!==marsState.floatingHeaderScrollBox) return;
                    const scrollHost=marsGetFloatingHeaderScroll();
                    if(scrollHost) scrollHost.scrollLeft=scrollBox.scrollLeft;
                }
                function marsRegisterTableScrollListeners() {
                    document.querySelectorAll(".mars-table-scroll[data-table-id]").forEach((scrollBox) => {
                        if(scrollBox.dataset.headerScrollBound==="1") return;
                        scrollBox.dataset.headerScrollBound="1";
                        scrollBox.addEventListener("scroll", marsHandleTableHorizontalScroll, {passive:true});
                    });
                }
                function marsBackToTop() {
                    const anchor=document.getElementById("mars-page-top");
                    if(anchor) {
                        anchor.scrollIntoView({behavior:"smooth", block:"start"});
                        return;
                    }
                    window.scrollTo({top:0, behavior:"smooth"});
                }
                function marsUpdateBackToTopVisibility() {
                    const button=document.getElementById("mars-back-to-top");
                    if(!button) return;
                    button.classList.toggle("is-visible", window.scrollY > 600);
                }
                function marsColumnWidthProperty(columnKey) { return columnKey==="feature" ? "--mars-feature-col-width" : columnKey==="secondary" ? "--mars-secondary-col-width" : "--mars-bin-col-width"; }
                function marsColumnDefaultWidth(columnKey) { return columnKey==="feature" ? 220 : columnKey==="secondary" ? 110 : 140; }
                function marsColumnMinWidth(columnKey) { return columnKey==="feature" ? 140 : 90; }
                function marsApplyColumnWidth(table, columnKey, width) {
                    if(!table) return;
                    const safeWidth=Math.max(marsColumnMinWidth(columnKey), Number(width)||marsColumnDefaultWidth(columnKey));
                    table.style.setProperty(marsColumnWidthProperty(columnKey), `${safeWidth}px`);
                }
                function marsSyncStickyLayout(table) {
                    if(!table) return;
                    const featureHeader=table.querySelector("thead .mars-feature-col");
                    if(featureHeader) {
                        const featureWidth=Math.max(140, Math.ceil(featureHeader.getBoundingClientRect().width || marsColumnDefaultWidth("feature")));
                        marsApplyColumnWidth(table, "feature", featureWidth);
                    }
                    const secondaryHeader=table.querySelector("thead .mars-secondary-col");
                    if(secondaryHeader) {
                        const secondaryWidth=Math.max(90, Math.ceil(secondaryHeader.getBoundingClientRect().width || marsColumnDefaultWidth("secondary")));
                        marsApplyColumnWidth(table, "secondary", secondaryWidth);
                    }
                    const binHeader=table.querySelector("thead .mars-bin-col");
                    if(binHeader) {
                        const binWidth=Math.max(90, Math.ceil(binHeader.getBoundingClientRect().width || marsColumnDefaultWidth("bin")));
                        marsApplyColumnWidth(table, "bin", binWidth);
                    }
                }
                function marsTablesForScope(scopeToken) {
                    if(scopeToken==="all") return Array.from(document.querySelectorAll("table.mars-data-table[id]"));
                    if(scopeToken==="pivot") return Array.from(document.querySelectorAll("table.mars-pivot-table[id]"));
                    if(scopeToken.startsWith("table:")) {
                        const table=document.getElementById(scopeToken.slice(6));
                        return table ? [table] : [];
                    }
                    return [];
                }
                function marsSyncScopeLayouts(scopeToken="all") {
                    marsTablesForScope(scopeToken).forEach((table)=>marsSyncStickyLayout(table));
                    marsRegisterTableScrollListeners();
                    marsRefreshFloatingHeader();
                }
                function marsOpenAncestorSections(node) {
                    let parent=node?.closest("details");
                    while(parent) {
                        parent.open=true;
                        parent=parent.parentElement?.closest("details");
                    }
                }
                function marsFindSummaryFeatureNode(feature, visibleOnly=false) {
                    const target=marsNormalizeFeatureValue(feature);
                    const nodes=Array.from(document.querySelectorAll("#mars-summary-table tbody tr[data-feature]"));
                    const candidateNodes=visibleOnly ? nodes.filter((node)=>node.style.display!=="none" && node.offsetParent!==null) : nodes;
                    for(const node of candidateNodes) {
                        if(marsNormalizeFeatureValue(node.dataset.feature)===target) return node;
                    }
                    for(const node of candidateNodes) {
                        if(marsNormalizeFeatureValue(node.dataset.feature).includes(target)) return node;
                    }
                    return null;
                }
                function marsClearSummaryLocalQuery() {
                    marsState.localQueries["mars-summary-table"]="";
                    const input=document.getElementById("mars-summary-table-query");
                    if(input) input.value="";
                }
                function marsClearJumpHighlight() {
                    if(marsState.jumpHighlightArmTimerId) {
                        window.clearTimeout(marsState.jumpHighlightArmTimerId);
                        marsState.jumpHighlightArmTimerId=null;
                    }
                    if(marsState.jumpHighlightTimerId) {
                        window.clearTimeout(marsState.jumpHighlightTimerId);
                        marsState.jumpHighlightTimerId=null;
                    }
                    if(marsState.jumpHighlightNode) marsState.jumpHighlightNode.classList.remove("mars-jump-highlight");
                    if(marsState.jumpHighlightCell) marsState.jumpHighlightCell.classList.remove("mars-jump-highlight-cell");
                    marsState.jumpHighlightNode=null;
                    marsState.jumpHighlightCell=null;
                }
                function marsActivateJumpHighlight(node, featureCell) {
                    marsClearJumpHighlight();
                    if(!node) return;
                    marsState.jumpHighlightNode=node;
                    marsState.jumpHighlightCell=featureCell||null;
                    node.classList.add("mars-jump-highlight");
                    if(featureCell) featureCell.classList.add("mars-jump-highlight-cell");
                    marsState.jumpHighlightTimerId=window.setTimeout(() => {
                        if(marsState.jumpHighlightNode===node) {
                            node.classList.remove("mars-jump-highlight");
                            if(featureCell) featureCell.classList.remove("mars-jump-highlight-cell");
                            marsState.jumpHighlightTimerId=null;
                            marsState.jumpHighlightNode=null;
                            marsState.jumpHighlightCell=null;
                        }
                    }, 3000);
                }
                function marsFocusSummaryFeature(node) {
                    if(!node) return;
                    const featureCell=node.querySelector(".mars-feature-col");
                    const scrollBox=node.closest(".mars-table-scroll");
                    marsOpenAncestorSections(node);
                    marsClearJumpHighlight();
                    window.requestAnimationFrame(() => {
                        node.scrollIntoView({behavior:"smooth", block:"center", inline:"nearest"});
                        if(featureCell) featureCell.scrollIntoView({behavior:"smooth", block:"nearest", inline:"start"});
                        if(scrollBox) scrollBox.scrollTo({left:0, behavior:"smooth"});
                        marsState.jumpHighlightArmTimerId=window.setTimeout(() => {
                            marsState.jumpHighlightArmTimerId=null;
                            marsActivateJumpHighlight(node, featureCell);
                        }, 140);
                    });
                }
                function marsJumpToFeature() {
                    const input=document.getElementById("mars-feature-jump-input");
                    const value=(input?.value||"").trim();
                    if(!value) {
                        marsSetError("mars-feature-jump-error", "Enter a feature name to jump.");
                        return;
                    }
                    let node=marsFindSummaryFeatureNode(value, true);
                    if(node) {
                        marsSetError("mars-feature-jump-error", "");
                        marsFocusSummaryFeature(node);
                        return;
                    }
                    node=marsFindSummaryFeatureNode(value, false);
                    if(!node) {
                        marsSetError("mars-feature-jump-error", `Feature "${value}" does not exist in Summary.`);
                        return;
                    }
                    const globalMatcher=marsBuildMatcher(marsState.globalQuery);
                    const summaryParsed=marsParseSummaryExpression(marsState.appliedSummaryExpression);
                    if(marsSummaryRowAllowedWithoutLocal(node, globalMatcher, summaryParsed)) {
                        marsClearSummaryLocalQuery();
                        marsQueueRefresh("table:mars-summary-table");
                        window.requestAnimationFrame(() => {
                            window.requestAnimationFrame(() => {
                                const refreshedNode=marsFindSummaryFeatureNode(value, true) || marsFindSummaryFeatureNode(value, false);
                                marsSetError("mars-feature-jump-error", "");
                                marsFocusSummaryFeature(refreshedNode);
                            });
                        });
                        return;
                    }
                    marsSetError("mars-feature-jump-error", `Feature "${value}" is hidden by data source, global search, or summary filter.`);
                }
                function marsStartColumnResize(event, tableId, columnKey) {
                    event.preventDefault();
                    event.stopPropagation();
                    const table=document.getElementById(tableId);
                    if(!table) return;
                    const property=marsColumnWidthProperty(columnKey);
                    const computed=getComputedStyle(table);
                    const startWidth=parseFloat(computed.getPropertyValue(property)) || marsColumnDefaultWidth(columnKey);
                    marsState.resizeState={ tableId, columnKey, property, startX:event.clientX, startWidth, pendingWidth:startWidth };
                    document.body.style.cursor="col-resize";
                    document.body.style.userSelect="none";
                }
                function marsHandleColumnResize(event) {
                    if(!marsState.resizeState) return;
                    const { startX, startWidth, columnKey } = marsState.resizeState;
                    const table=document.getElementById(marsState.resizeState.tableId);
                    if(!table) return;
                    const minWidth=marsColumnMinWidth(columnKey);
                    const nextWidth=Math.max(minWidth, startWidth + (event.clientX - startX));
                    marsState.resizeState.pendingWidth=nextWidth;
                    if(marsState.resizeFrameScheduled) return;
                    marsState.resizeFrameScheduled=true;
                    window.requestAnimationFrame(() => {
                        marsState.resizeFrameScheduled=false;
                        if(!marsState.resizeState) return;
                        const activeTable=document.getElementById(marsState.resizeState.tableId);
                        marsApplyColumnWidth(activeTable, marsState.resizeState.columnKey, marsState.resizeState.pendingWidth);
                        marsRefreshFloatingHeader();
                    });
                }
                function marsStopColumnResize() {
                    if(!marsState.resizeState) return;
                    const table=document.getElementById(marsState.resizeState.tableId);
                    if(table) marsSyncStickyLayout(table);
                    marsState.resizeState=null;
                    document.body.style.cursor="";
                    document.body.style.userSelect="";
                    marsScheduleViewportRefresh();
                }
                function marsRefreshSummaryContext() {
                    const summaryFeatures=marsGetSummaryFeatureAllowSet();
                    marsState.summaryAllowedFeatures=summaryFeatures instanceof Set ? summaryFeatures : null;
                }
                function marsRefreshSummaryTable() { marsApplyTableFilter("mars-summary-table"); }
                function marsRefreshGenericTables() {
                    document.querySelectorAll("table.mars-data-table[id]").forEach((table)=>{
                        if(table.id==="mars-summary-table" || table.classList.contains("mars-pivot-table")) return;
                        marsApplyTableFilter(table.id);
                    });
                }
                function marsRefreshPivotScope() {
                    marsUpdatePivotViews();
                    document.querySelectorAll("table.mars-pivot-table[id]").forEach((table)=>marsApplyTableFilter(table.id));
                    marsQueueLayoutSync("pivot");
                }
                function marsRefreshScopeToken(scopeToken) {
                    if(scopeToken==="all") {
                        marsRefreshSummaryContext();
                        marsRefreshSummaryTable();
                        marsRefreshGenericTables();
                        marsRefreshPivotScope();
                        marsUpdateChartViews();
                        return;
                    }
                    if(scopeToken==="pivot") { marsRefreshPivotScope(); return; }
                    if(scopeToken==="charts") { marsUpdateChartViews(); return; }
                    if(scopeToken==="summary") { marsRefreshSummaryTable(); return; }
                    if(scopeToken.startsWith("table:")) { marsApplyTableFilter(scopeToken.slice(6)); }
                }
                function marsFlushRefreshQueue() {
                    const tokens = marsState.pendingRefreshTokens.length ? marsState.pendingRefreshTokens.slice() : ["all"];
                    marsState.pendingRefreshTokens = [];
                    if(tokens.includes("all")) {
                        marsRefreshScopeToken("all");
                        marsScheduleViewportRefresh();
                        return;
                    }
                    tokens.forEach((token)=>marsRefreshScopeToken(token));
                    marsScheduleViewportRefresh();
                }
                window.addEventListener("mousemove", marsHandleColumnResize);
                window.addEventListener("mouseup", marsStopColumnResize);
                window.addEventListener("resize", () => { marsQueueLayoutSync("all"); marsScheduleViewportRefresh(); });
                window.addEventListener("scroll", marsScheduleViewportRefresh, {passive:true});
                document.addEventListener("toggle", () => { marsHideFloatingHeader(); marsQueueLayoutSync("all"); marsScheduleViewportRefresh(); }, true);
                window.addEventListener("DOMContentLoaded", () => {
                    marsRegisterTableScrollListeners();
                    marsSetDataSources();
                    marsQueueLayoutSync("all");
                    marsQueueRefresh("all");
                    marsUpdateBackToTopVisibility();
                    marsRefreshFloatingHeader();
                });
        """
        return template.replace("__SUMMARY_FILTER_COLUMNS__", json.dumps(summary_filter_columns, ensure_ascii=False))

    @classmethod
    def _build_html_document(
        cls,
        *,
        report_name: str,
        styles: str,
        body_html: str,
        runtime_script: str,
    ) -> str:
        template = """
        <!DOCTYPE html>
        <html lang="zh">
        <head>
            <meta charset="utf-8" />
            <meta name="viewport" content="width=device-width, initial-scale=1" />
            <title>__TITLE__</title>
            <style>
__STYLES__
            </style>
        </head>
        <body>
__BODY_HTML__
            <script>
__RUNTIME_SCRIPT__
            </script>
        </body>
        </html>
        """
        return (
            template
            .replace("__TITLE__", html.escape(report_name))
            .replace("__STYLES__", styles)
            .replace("__BODY_HTML__", body_html)
            .replace("__RUNTIME_SCRIPT__", runtime_script)
        )

    @classmethod
    def _build_global_tools_html(
        cls,
        *,
        feature_jump_html: str,
        source_options: str,
    ) -> str:
        export_block_html = (
            '<div class="mars-export-block">'
            '<button type="button" class="mars-clear-button" onclick="marsExportFeatures()">Export Feature List</button>'
            '<div class="mars-export-helper">Export uses Summary expression + Data Source only. Global and local searches only affect display.</div>'
            '</div>'
        )
        return (
            '<div class="mars-global-tools">'
            '<div class="mars-search-cluster">'
            '<input id="mars-global-search" class="mars-filter-input" type="search" placeholder="Global search across tables and charts..." oninput="marsSetGlobalQuery(this.value)" />'
            '<button type="button" class="mars-clear-button" onclick="marsClearGlobalSearch()">Clear Search</button>'
            '</div>'
            '<label class="mars-toggle"><input id="mars-regex-mode" type="checkbox" onchange="marsSetRegexMode(this.checked)" /> Regex Mode</label>'
            f'{feature_jump_html}'
            '<div class="mars-source-panel">'
            '<div class="mars-source-header">'
            '<strong>Data Source</strong>'
            '<div>'
            '<button type="button" class="mars-source-link" onclick="marsSelectAllSources()">All</button>'
            '<button type="button" class="mars-source-link" onclick="marsClearSources()">Clear</button>'
            '</div>'
            '</div>'
            f'<div id="mars-data-source-group" class="mars-source-options">{source_options}</div>'
            '</div>'
            f'{export_block_html}'
            '</div>'
        )

    def _build_summary_section_html(
        self,
        *,
        summary_pd: pd.DataFrame,
        feature_sources: Dict[str, str],
        sort_by: str,
        ascending: bool,
    ) -> Optional[str]:
        if summary_pd.empty:
            return None

        summary_df = summary_pd.copy()
        if sort_by in summary_df.columns:
            summary_df = summary_df.sort_values(sort_by, ascending=ascending)
        if "mono" in summary_df.columns:
            summary_df = summary_df.drop(columns=["mono"])

        summary_semantics = {col: self._semantic_for_metric(col) for col in summary_df.columns}
        summary_percent_cols = [col for col in summary_df.columns if self._is_percent_column(col)]
        summary_style_rules = {
            col: self._summary_style_rule(col)
            for col in summary_df.columns
            if self._summary_style_rule(col) is not None
        }
        summary_legend_html = self._build_threshold_legend_html(
            [
                ("IV", "0.01 / 0.05 / 0.10, purple above 0.20"),
                ("KS", "4 / 8 / 12, purple above 16"),
                ("AUC", "0.525 / 0.55 / 0.575, purple above 0.625"),
                ("PSI", "0 / 0.1 / 0.25"),
                ("Missing", "0 / 0.5 / 1"),
                ("Lift Min", "0.5 / 0.6 / 0.7 / 0.8, purple at or below 0.5"),
                ("Lift Max", "1.2 / 1.3 / 1.4, purple above 1.5"),
            ],
            legend_id="mars-summary-legend",
        )
        summary_filter_html = (
            '<div class="mars-summary-filter">'
            '<label class="mars-summary-filter-label" for="mars-summary-expression">Feature Filter Expression</label>'
            '<input id="mars-summary-expression" class="mars-filter-input" type="search" '
            'placeholder="e.g. iv > 0.05 & (psi_max < 0.1 | missing < 0.2)" '
            'oninput="marsSetSummaryExpression(this.value)" />'
            '<div id="mars-summary-expression-error" class="mars-search-error"></div>'
            '<div class="mars-footnote">Available metrics: iv, ks, auc, psi_max, rc_min, lift_min, lift_max, missing, missing_min, missing_max.</div>'
            '<div class="mars-footnote">Supported operators: &gt;, &gt;=, &lt;, &lt;=, ==, !=, &amp;, |, ( ). Use &amp; for AND and | for OR.</div>'
            f'{summary_legend_html}'
            '</div>'
        )
        summary_table_html = self._build_enhanced_table_html(
            summary_df,
            "mars-summary-table",
            search_placeholder="Search summary table...",
            feature_sources=feature_sources,
            semantic_map=summary_semantics,
            percent_cols=summary_percent_cols,
            style_rule_map=summary_style_rules,
            extra_toolbar_html=summary_filter_html,
            table_kind="summary",
        )
        return self._wrap_html_section(
            "Summary",
            summary_table_html,
            "summary-section",
            subtitle="Feature-level ranking and monitoring summary.",
        )

    def _build_trend_sections_html(
        self,
        *,
        trend_pd_map: Dict[str, pd.DataFrame],
        missing_by_day_pd: Optional[pd.DataFrame],
        feature_sources: Dict[str, str],
    ) -> List[Tuple[str, str, str]]:
        sections: List[Tuple[str, str, str]] = []

        if missing_by_day_pd is not None and not missing_by_day_pd.empty:
            missing_day_df = self._reorder_group_columns(missing_by_day_pd.copy(), ["feature", "dtype"])
            missing_day_semantics = {col: "risk_high" for col in missing_day_df.columns if col not in {"feature", "dtype"}}
            missing_day_percent_cols = [
                col for col in missing_day_df.columns
                if self._is_percent_column(col, metric_name="missing")
            ]
            missing_day_style_rules = {
                col: self._trend_style_rule("missing")
                for col in missing_day_df.columns
                if col not in {"feature", "dtype"}
            }
            missing_day_html = self._build_enhanced_table_html(
                missing_day_df,
                "mars-missing-by-day",
                search_placeholder="Search daily missing trend...",
                feature_sources=feature_sources,
                semantic_map=missing_day_semantics,
                percent_cols=missing_day_percent_cols,
                style_rule_map=missing_day_style_rules,
            )
            sections.append((
                "missing-day-section",
                "Missing By Day",
                self._wrap_html_section(
                    "Missing Trend By Day",
                    missing_day_html,
                    "missing-day-section",
                    subtitle=f"Daily missing-rate trend derived from dt_col={self.dt_col}.",
                ),
            ))

        trend_blocks: List[str] = []
        trend_legend_html = self._build_threshold_legend_html(
            [
                ("Missing", "0 / 0.5 / 1"),
                ("PSI", "0 / 0.1 / 0.25"),
                ("IV", "0.01 / 0.05 / 0.10, purple above 0.20"),
                ("KS", "4 / 8 / 12, purple above 16"),
                ("Lift", "1.2 / 1.3 / 1.4, purple above 1.5"),
                ("AUC", "0.525 / 0.55 / 0.575, purple above 0.625"),
                ("Risk Corr", "0.2 / 0.5 / 0.8"),
            ],
            legend_id="mars-trend-legend",
        )
        for metric in ["missing", "psi", "iv", "ks", "lift", "auc", "risk_corr"]:
            if metric not in trend_pd_map:
                continue
            trend_df = self._sort_metric_display_df(
                self._reorder_group_columns(trend_pd_map[metric].copy(), ["feature", "dtype"])
            )
            trend_semantics = {col: self._semantic_for_metric(metric) for col in trend_df.columns if col not in {"feature", "dtype"}}
            trend_percent_cols = [
                col for col in trend_df.columns
                if self._is_percent_column(col, metric_name=metric)
            ]
            trend_style_rules = {
                col: self._trend_style_rule(metric)
                for col in trend_df.columns
                if col not in {"feature", "dtype"}
            }
            table_html = self._build_enhanced_table_html(
                trend_df,
                f"mars-trend-{self._slugify(metric)}",
                search_placeholder=f"Search {metric} trend...",
                feature_sources=feature_sources,
                semantic_map=trend_semantics,
                percent_cols=trend_percent_cols,
                style_rule_map=trend_style_rules,
            )
            trend_blocks.append(f'<details class="mars-metric-block" open><summary>{html.escape(metric.upper())}</summary>{table_html}</details>')
        if trend_blocks:
            sections.append((
                "trend-section",
                "Trend Tables",
                self._wrap_html_section(
                    "Trend Tables",
                    trend_legend_html + "".join(trend_blocks),
                    "trend-section",
                    subtitle="Trend tables with local search, default Total ranking, and Excel-like color scales.",
                ),
            ))

        return sections

    def _build_chart_section_html(
        self,
        *,
        detail_pd: pd.DataFrame,
        summary_pd: pd.DataFrame,
        feature_sources: Dict[str, str],
        max_plots: int,
        sort_by: str,
        ascending: bool,
    ) -> Optional[str]:
        if detail_pd.empty:
            return None

        chart_y_values = [str(v) for v in detail_pd["y"].dropna().astype(str).drop_duplicates().tolist()] if "y" in detail_pd.columns else ["Target"]
        chart_controls = (
            '<div class="mars-inline-controls mars-chart-controls">'
            '<input class="mars-filter-input mars-chart-search" type="search" '
            'placeholder="Search chart features..." '
            'oninput="marsSetLocalQuery(\'mars-chart-cards\', this.value)" />'
            '<div id="mars-chart-cards-error" class="mars-search-error"></div>'
        )
        if len(chart_y_values) > 1:
            chart_options = "".join(f'<option value="{html.escape(y_val)}">{html.escape(y_val)}</option>' for y_val in chart_y_values)
            chart_controls += f'<label class="mars-select-group">Chart Target<select id="mars-chart-target" onchange="marsHandleChartTargetChange()">{chart_options}</select></label>'
        chart_controls += "</div>"

        chart_views: List[str] = []
        try:
            from mars.utils.plotter import MarsPlotter

            for y_val in chart_y_values:
                chart_detail_pd = detail_pd[detail_pd["y"].astype(str) == y_val].copy() if "y" in detail_pd.columns else detail_pd.copy()
                chart_summary_pd = summary_pd[summary_pd["target"].astype(str) == y_val].copy() if "target" in summary_pd.columns else summary_pd.copy()
                chart_sort_col = self._resolve_chart_sort_column(chart_summary_pd, sort_by)
                if not chart_summary_pd.empty and chart_sort_col:
                    chart_summary_pd = chart_summary_pd.sort_values(chart_sort_col, ascending=ascending)
                if not chart_summary_pd.empty and "feature" in chart_summary_pd.columns:
                    chart_features = chart_summary_pd["feature"].drop_duplicates().tolist()[:max_plots]
                else:
                    chart_features = chart_detail_pd["feature"].drop_duplicates().tolist()[:max_plots]

                chart_cards: List[str] = []
                for feature in chart_features:
                    block_html = MarsPlotter.render_feature_binning_risk_trend_html(
                        df_detail=chart_detail_pd,
                        feature=feature,
                        group_col=self.group_col or "mars_group",
                        target_name=y_val,
                        dpi=150,
                    )
                    if not block_html:
                        continue
                    data_source = feature_sources.get(str(feature), "UNMAPPED")
                    chart_cards.append(
                        f'<article class="mars-chart-card" data-feature="{self._escape_attr(feature)}" data-data-source="{self._escape_attr(data_source)}" '
                        f'data-search-text="{self._escape_attr(self._normalize_search_text(feature, y_val, data_source))}"><h4>{html.escape(str(feature))}</h4>{block_html}</article>'
                    )
                if not chart_cards:
                    chart_cards.append('<div class="mars-empty">No chart data available for this target.</div>')
                chart_views.append(f'<div class="mars-chart-view" data-y-value="{self._escape_attr(y_val)}">{"".join(chart_cards)}</div>')
        except Exception as exc:
            logger.warning("HTML chart rendering skipped due to error: %s", exc)

        if not chart_views:
            return None

        chart_feedback_html = self._build_scope_feedback_html("mars-chart-cards", empty_text="No charts match current filters.")
        return self._wrap_html_section(
            "Charts",
            chart_controls + chart_feedback_html + f'<div id="mars-chart-cards">{"".join(chart_views)}</div>',
            "chart-section",
            subtitle="Risk trend charts rendered from the shared plotting path.",
            open_by_default=False,
        )

    @classmethod
    def _build_threshold_style(cls, value: float, rule: Dict[str, Any]) -> str:
        anchors = tuple(float(v) for v in rule["anchors"])
        colors = tuple(rule["colors"])
        purple_above = rule.get("purple_above")
        purple_rgb = tuple(rule.get("purple_rgb", (160, 98, 196)))

        if len(anchors) != len(colors):
            raise ValueError("Threshold style rules require the same number of anchors and colors.")

        red, green, blue = colors[-1]
        if value <= anchors[0]:
            red, green, blue = colors[0]
        else:
            segment_found = False
            for idx in range(len(anchors) - 1):
                start_anchor = anchors[idx]
                end_anchor = anchors[idx + 1]
                if value <= end_anchor:
                    ratio = 0.5 if abs(end_anchor - start_anchor) < 1e-12 else (value - start_anchor) / (end_anchor - start_anchor)
                    red, green, blue = cls._interpolate_rgb(colors[idx], colors[idx + 1], ratio)
                    segment_found = True
                    break
            if not segment_found and purple_above is not None:
                high = anchors[-1]
                upper = float(purple_above)
                ratio = 1.0 if abs(upper - high) < 1e-12 else min(max((value - high) / (upper - high), 0.0), 1.0)
                red, green, blue = cls._interpolate_rgb(colors[-1], purple_rgb, ratio)

        alpha = 0.84 if purple_above is not None and value >= float(purple_above) else 0.72
        font_style = " color: #fff; font-weight: 600;" if purple_above is not None and value >= float(purple_above) else ""
        return f"background-color: rgba({red}, {green}, {blue}, {alpha});{font_style}"

    @classmethod
    def _cell_style(
        cls,
        value: Any,
        *,
        semantic: str,
        vmin: Optional[float],
        vmax: Optional[float],
        style_rule: Optional[Dict[str, Any]] = None,
        data_bar: bool = False,
    ) -> str:
        if cls._is_missing_html_value(value):
            return ""

        try:
            num = float(value)
        except (TypeError, ValueError):
            return ""

        if not np.isfinite(num):
            return ""

        styles: List[str] = []
        if style_rule is not None:
            styles.append(cls._build_threshold_style(num, style_rule))

        if vmin is not None and vmax is not None and np.isfinite(vmin) and np.isfinite(vmax):
            span = vmax - vmin
            ratio = 0.5 if abs(span) < 1e-12 else (num - vmin) / span
            ratio = max(0.0, min(1.0, ratio))
            if style_rule is None:
                if semantic == "risk_high":
                    red, green, blue = cls._three_color_rgb(ratio, reverse=True)
                    styles.append(f"background-color: rgba({red}, {green}, {blue}, 0.72);")
                elif semantic == "good_high":
                    red, green, blue = cls._three_color_rgb(ratio, reverse=False)
                    styles.append(f"background-color: rgba({red}, {green}, {blue}, 0.72);")
                elif semantic == "diverging":
                    max_abs = max(abs(vmin), abs(vmax), 1e-12)
                    diverging_ratio = min(abs(num) / max_abs, 1.0)
                    if num >= 0:
                        red, green, blue = cls._interpolate_rgb((255, 235, 132), (99, 190, 123), diverging_ratio)
                    else:
                        red, green, blue = cls._interpolate_rgb((255, 235, 132), (248, 105, 107), diverging_ratio)
                    styles.append(f"background-color: rgba({red}, {green}, {blue}, 0.72);")

            if data_bar and style_rule is None:
                bar_ratio = ratio if semantic != "diverging" else min(abs(num) / max(abs(vmin), abs(vmax), 1e-12), 1.0)
                bar_color = "#8bbf9d" if semantic not in {"risk_high", "diverging"} else "#ea8f8f"
                if semantic == "good_high":
                    bar_color = "#7fc68d"
                if semantic == "diverging" and num >= 0:
                    bar_color = "#7fc68d"
                styles.append(
                    f"background-image: linear-gradient(90deg, {bar_color} 0%, {bar_color} {bar_ratio * 100:.2f}%, transparent {bar_ratio * 100:.2f}%);"
                    "background-repeat: no-repeat;"
                )
        return "".join(styles)

    @classmethod
    def _build_enhanced_table_html(
        cls,
        df: pd.DataFrame,
        table_id: str,
        *,
        search_placeholder: str,
        feature_sources: Optional[Dict[str, str]] = None,
        semantic_map: Optional[Dict[str, str]] = None,
        data_bar_cols: Optional[List[str]] = None,
        percent_cols: Optional[List[str]] = None,
        style_rule_map: Optional[Dict[str, Dict[str, Any]]] = None,
        extra_toolbar_html: str = "",
        table_kind: str = "generic",
        empty_text: str = "No data available.",
    ) -> str:
        if df.empty:
            return f'<div class="mars-empty">{html.escape(empty_text)}</div>'

        semantic_map = semantic_map or {}
        data_bar_cols = set(data_bar_cols or [])
        percent_cols = set(percent_cols or [])
        style_rule_map = style_rule_map or {}
        feature_sources = feature_sources or {}

        sort_types = {
            col: "number" if pd.api.types.is_numeric_dtype(df[col]) else "text"
            for col in df.columns
        }
        numeric_bounds = {}
        for col in df.columns:
            if sort_types[col] == "number":
                numeric_series = pd.to_numeric(df[col], errors="coerce")
                if numeric_series.notna().any():
                    numeric_bounds[col] = (float(numeric_series.min()), float(numeric_series.max()))
                else:
                    numeric_bounds[col] = (None, None)

        header_cells = []
        for col in df.columns:
            sort_type = sort_types[col]
            numeric_class = " is-numeric" if sort_type == "number" else ""
            sticky_role = cls._table_sticky_role(col)
            sticky_class = cls._sticky_class_for_role(sticky_role)
            sticky_inner_class = cls._sticky_inner_class_for_role(sticky_role)
            resize_handle = (
                f'<span class="mars-resize-handle" onmousedown="marsStartColumnResize(event, \'{table_id}\', \'feature\')"></span>'
                if sticky_role == "feature" else ""
            )
            header_cells.append(
                f'<th class="mars-th{numeric_class}{sticky_class}" data-sort-type="{sort_type}" data-col-index="{len(header_cells)}">'
                f'<button type="button" class="mars-sort-button{sticky_inner_class}" onclick="marsSortTable(\'{table_id}\', this)">'
                f'<span class="mars-sort-label">{html.escape(str(col))}</span><span class="mars-sort-indicator"></span>'
                f'</button>{resize_handle}</th>'
            )

        body_rows = []
        for _, row in df.iterrows():
            feature = str(row["feature"]) if "feature" in df.columns else ""
            data_source = str(row.get("data_source", feature_sources.get(feature, "UNMAPPED")))
            metric_payload = {
                str(col): float(row[col])
                for col in df.columns
                if sort_types.get(col) == "number" and not cls._is_missing_html_value(row[col])
            }
            search_text = cls._normalize_search_text(
                feature,
                data_source,
                *[
                    cls._format_html_value(row[col], as_percent=(col in percent_cols))
                    for col in df.columns
                ],
            )

            row_cells = []
            for col in df.columns:
                sort_type = sort_types[col]
                display_val = cls._format_html_value(row[col], as_percent=(col in percent_cols))
                sort_val = cls._format_sort_value(row[col], sort_type)
                numeric_class = " is-numeric" if sort_type == "number" else ""
                semantic = semantic_map.get(col, "neutral")
                style_rule = style_rule_map.get(col)
                vmin, vmax = numeric_bounds.get(col, (None, None))
                style = cls._cell_style(
                    row[col],
                    semantic=semantic,
                    vmin=vmin,
                    vmax=vmax,
                    style_rule=style_rule,
                    data_bar=col in data_bar_cols,
                )
                sticky_role = cls._table_sticky_role(col)
                sticky_class = cls._sticky_class_for_role(sticky_role)
                sticky_inner_class = cls._sticky_inner_class_for_role(sticky_role)
                row_cells.append(
                    f'<td class="mars-td{numeric_class}{sticky_class}" data-col="{cls._escape_attr(col)}" '
                    f'data-sort-value="{cls._escape_attr(sort_val)}" style="{cls._escape_attr(style)}">'
                    f'<span class="mars-cell-text{sticky_inner_class}">{html.escape(display_val)}</span></td>'
                )

            body_rows.append(
                f'<tr data-feature="{cls._escape_attr(feature)}" data-data-source="{cls._escape_attr(data_source)}" '
                f'data-search-text="{cls._escape_attr(search_text)}" '
                f'data-metrics="{cls._escape_attr(json.dumps(metric_payload, ensure_ascii=False, separators=(",", ":")))}">'
                f'{"".join(row_cells)}</tr>'
            )

        return f"""
        <div id="{table_id}-wrap" class="mars-table-wrap">
            <div class="mars-table-toolbar">
                <input
                    id="{table_id}-query"
                    class="mars-filter-input"
                    type="search"
                    placeholder="{html.escape(search_placeholder)}"
                    oninput="marsSetLocalQuery('{table_id}', this.value)"
                />
                <div id="{table_id}-error" class="mars-search-error"></div>
            </div>
            {extra_toolbar_html}
            {cls._build_scope_feedback_html(table_id, empty_text="No rows match current filters.")}
            <div class="mars-table-ownership-sentinel" data-table-id="{table_id}" data-sentinel-role="start" aria-hidden="true"></div>
            <div class="mars-table-scroll" data-table-id="{table_id}">
                <table id="{table_id}" class="mars-data-table" data-sort-col="" data-sort-dir="" data-table-kind="{cls._escape_attr(table_kind)}" style="--mars-feature-col-width: 220px; --mars-secondary-col-width: 110px;">
                    <thead><tr>{''.join(header_cells)}</tr></thead>
                    <tbody>{''.join(body_rows)}</tbody>
                </table>
            </div>
            <div class="mars-table-ownership-sentinel" data-table-id="{table_id}" data-sentinel-role="end" aria-hidden="true"></div>
        </div>
        """

    @classmethod
    def _build_grouped_pivot_section_html(
        cls,
        detail_pd: pd.DataFrame,
        *,
        group_col: str,
        feature_sources: Dict[str, str],
    ) -> str:
        if detail_pd.empty or group_col not in detail_pd.columns:
            return '<div class="mars-empty">No grouped pivot data available.</div>'

        work_df = detail_pd.copy()
        if "bin_label" in work_df.columns:
            work_df = work_df[work_df["bin_label"].astype(str) != "Total"]
        work_df = work_df[work_df[group_col].astype(str) != "Total"]
        if "bin_type" in work_df.columns:
            work_df = work_df[work_df["bin_type"].astype(str).isin(["首尾组", "正常组", "空值组"])]
        if work_df.empty:
            return '<div class="mars-empty">No grouped pivot data available.</div>'

        missing_idx = -1
        try:
            from mars.feature.binner import MarsBinnerBase

            missing_idx = int(MarsBinnerBase.IDX_MISSING)
        except Exception:
            missing_idx = -1

        work_df["bin_index_num"] = pd.to_numeric(work_df.get("bin_index"), errors="coerce")
        missing_mask = work_df["bin_index_num"].eq(missing_idx)
        if "bin_label" in work_df.columns:
            missing_mask = missing_mask | work_df["bin_label"].astype(str).str.lower().eq("missing")

        max_bin_index = (
            work_df.loc[~missing_mask, ["feature", group_col, "bin_index_num"]]
            .groupby(["feature", group_col], dropna=False)["bin_index_num"]
            .transform("max")
        )
        work_df["pivot_bin_type"] = np.where(
            missing_mask,
            "空值组",
            np.where(
                work_df["bin_index_num"].eq(0) | work_df["bin_index_num"].eq(max_bin_index),
                "首尾组",
                "正常组",
            ),
        )

        if "data_source" not in work_df.columns:
            work_df["data_source"] = work_df["feature"].map(feature_sources).fillna("UNMAPPED")
        else:
            work_df["data_source"] = work_df["data_source"].fillna("UNMAPPED")

        y_values = (
            [str(v) for v in work_df["y"].dropna().astype(str).drop_duplicates().tolist()]
            if "y" in work_df.columns
            else ["Target"]
        )
        control_parts = []
        if len(y_values) > 1:
            options = "".join(
                f'<option value="{html.escape(y_value)}">{html.escape(y_value)}</option>'
                for y_value in y_values
            )
            control_parts.append(
                f'<label class="mars-select-group">Pivot Target'
                f'<select id="mars-pivot-target" onchange="marsHandlePivotTargetChange()">{options}</select>'
                f'</label>'
            )

        metrics = [
            ("pct", "pct", "good_high", True, True),
            ("risk", "risk", "risk_high", True, True),
            ("lift", "lift", "good_high", False, True),
            ("bad_count", "bad count", "risk_high", False, False),
            ("total_count", "total count", "good_high", False, False),
            ("iv", "iv", "good_high", False, False),
        ]
        view_blocks: List[str] = []

        for y_val in y_values:
            view_df = work_df.copy()
            if "y" in view_df.columns:
                view_df = view_df[view_df["y"].astype(str) == y_val]
            if view_df.empty:
                continue

            for col in ["pct", "bad", "count", "lift", "iv_bin"]:
                if col not in view_df.columns:
                    view_df[col] = np.nan
            view_df["bad"] = pd.to_numeric(view_df["bad"], errors="coerce").fillna(0.0)
            view_df["count"] = pd.to_numeric(view_df["count"], errors="coerce").fillna(0.0)
            view_df["pct"] = pd.to_numeric(view_df["pct"], errors="coerce").fillna(0.0)
            view_df["lift"] = pd.to_numeric(view_df["lift"], errors="coerce")
            view_df["iv_bin"] = pd.to_numeric(view_df["iv_bin"], errors="coerce").fillna(0.0)

            grouped = (
                view_df.groupby(
                    ["feature", "bin_label", "bin_index", "pivot_bin_type", group_col],
                    dropna=False,
                )
                .agg(
                    bad_count=("bad", "sum"),
                    total_count=("count", "sum"),
                    lift=("lift", "max"),
                    iv=("iv_bin", "sum"),
                )
                .reset_index()
            )
            group_denominator = grouped.groupby(
                ["feature", group_col],
                dropna=False,
            )["total_count"].transform("sum")
            grouped["pct"] = grouped["total_count"] / group_denominator.replace(0, np.nan)
            grouped["pct"] = grouped["pct"].fillna(0.0)
            grouped["risk"] = grouped["bad_count"] / (grouped["total_count"] + 1e-9)

            totals = (
                grouped.groupby(["feature", "bin_label", "bin_index", "pivot_bin_type"], dropna=False)
                .agg(
                    bad_count=("bad_count", "sum"),
                    total_count=("total_count", "sum"),
                    lift=("lift", "max"),
                    iv=("iv", "sum"),
                )
                .reset_index()
            )
            feature_denominator = totals.groupby(["feature"], dropna=False)["total_count"].transform("sum")
            totals["pct"] = totals["total_count"] / feature_denominator.replace(0, np.nan)
            totals["pct"] = totals["pct"].fillna(0.0)
            totals["risk"] = totals["bad_count"] / (totals["total_count"] + 1e-9)
            totals[group_col] = "Total"
            feature_rank = (
                totals.groupby(["feature"], dropna=False)["iv"]
                .sum()
                .reset_index()
                .sort_values(["iv", "feature"], ascending=[False, True])
            )
            grouped = pd.concat([grouped, totals], ignore_index=True, sort=False)
            ordered_groups = ["Total"] + sorted(
                [g for g in grouped[group_col].astype(str).unique().tolist() if g != "Total"]
            )

            metric_bounds = {}
            for metric_key, _, _, _, _ in metrics:
                vals = pd.to_numeric(grouped[metric_key], errors="coerce")
                metric_bounds[metric_key] = (
                    float(vals.min()) if vals.notna().any() else None,
                    float(vals.max()) if vals.notna().any() else None,
                )

            header_top = [
                '<th rowspan="2" class="mars-th mars-sticky-col mars-feature-col">Feature'
                f'<span class="mars-resize-handle" onmousedown="marsStartColumnResize(event, \'mars-pivot-{cls._slugify(y_val)}\', \'feature\')"></span>'
                '</th>',
                '<th rowspan="2" class="mars-th mars-sticky-col mars-bin-col">Bin'
                f'<span class="mars-resize-handle" onmousedown="marsStartColumnResize(event, \'mars-pivot-{cls._slugify(y_val)}\', \'bin\')"></span>'
                '</th>',
            ]
            header_bottom: List[str] = []
            for _, metric_label, _, _, _ in metrics:
                header_top.append(
                    f'<th class="mars-th is-numeric" colspan="{len(ordered_groups)}">{html.escape(metric_label)}</th>'
                )
                for group_value in ordered_groups:
                    header_bottom.append(f'<th class="mars-th is-numeric">{html.escape(str(group_value))}</th>')

            row_chunks: List[str] = []
            total_columns = 2 + len(metrics) * len(ordered_groups)
            grouped["bin_index"] = pd.to_numeric(grouped["bin_index"], errors="coerce")
            grouped = grouped.sort_values(["feature", "bin_index", "bin_label", group_col], na_position="last")
            feature_order = feature_rank["feature"].astype(str).tolist()
            for feature in feature_order:
                feature_df = grouped[grouped["feature"].astype(str) == feature].copy()
                if feature_df.empty:
                    continue
                bin_frames = []
                for (bin_label, bin_index, pivot_bin_type), bin_df in feature_df.groupby(
                    ["bin_label", "bin_index", "pivot_bin_type"],
                    dropna=False,
                ):
                    bin_frames.append((bin_index, str(bin_label), str(pivot_bin_type), bin_df.set_index(group_col)))

                bin_frames.sort(key=lambda item: (float(item[0]) if pd.notna(item[0]) else np.inf, item[1]))
                first_row = True
                for _, bin_label, pivot_bin_type, row_df in bin_frames:
                    row_cells = [
                        f'<td class="mars-td mars-pivot-feature mars-sticky-col mars-feature-col{" mars-pivot-feature-blank" if not first_row else ""}"><span class="mars-cell-text mars-sticky-cell-inner">{html.escape(feature if first_row else "")}</span></td>',
                        f'<td class="mars-td mars-pivot-bin mars-sticky-col mars-bin-col"><span class="mars-cell-text mars-sticky-cell-inner">{html.escape(bin_label)}</span></td>',
                    ]
                    for metric_key, _, semantic, is_percent, show_bar in metrics:
                        vmin, vmax = metric_bounds[metric_key]
                        for group_value in ordered_groups:
                            value = row_df[metric_key].get(group_value, np.nan)
                            display_val = cls._format_html_value(value, as_percent=is_percent)
                            sort_val = cls._format_sort_value(value, "number")
                            style = cls._cell_style(value, semantic=semantic, vmin=vmin, vmax=vmax, data_bar=show_bar)
                            row_cells.append(
                                f'<td class="mars-td is-numeric" data-col="{cls._escape_attr(metric_key)}|{cls._escape_attr(group_value)}" '
                                f'data-sort-value="{cls._escape_attr(sort_val)}" style="{cls._escape_attr(style)}">{html.escape(display_val)}</td>'
                            )

                    search_text = cls._normalize_search_text(feature, bin_label, pivot_bin_type)
                    row_chunks.append(
                        f'<tr data-feature="{cls._escape_attr(feature)}" data-data-source="__aggregate__" '
                        f'data-search-text="{cls._escape_attr(search_text)}">{"".join(row_cells)}</tr>'
                    )
                    first_row = False

                row_chunks.append(
                    f'<tr class="mars-pivot-spacer-row" data-role="spacer" data-feature="{cls._escape_attr(feature)}" '
                    f'data-data-source="__aggregate__"><td colspan="{total_columns}"></td></tr>'
                )

            if row_chunks:
                table_id = f"mars-pivot-{cls._slugify(y_val)}"
                table_html = (
                    f'<div class="mars-table-toolbar">'
                    f'<input id="{table_id}-query" class="mars-filter-input" type="search" placeholder="Search grouped pivot..." '
                    f'oninput="marsSetLocalQuery(\'{table_id}\', this.value)" />'
                    f'<div id="{table_id}-error" class="mars-search-error"></div>'
                    f'</div>'
                    f'{cls._build_scope_feedback_html(table_id, empty_text="No grouped pivot rows match current filters.")}'
                    f'<div class="mars-table-ownership-sentinel" data-table-id="{cls._escape_attr(table_id)}" data-sentinel-role="start" aria-hidden="true"></div>'
                    f'<div class="mars-table-scroll" data-table-id="{cls._escape_attr(table_id)}">'
                    f'<table id="{table_id}" class="mars-data-table mars-pivot-table" data-sort-col="" data-sort-dir="" data-table-kind="pivot" style="--mars-feature-col-width: 220px; --mars-bin-col-width: 140px;">'
                    f'<thead><tr>{"".join(header_top)}</tr><tr>{"".join(header_bottom)}</tr></thead>'
                    f'<tbody>{"".join(row_chunks)}</tbody></table></div>'
                    f'<div class="mars-table-ownership-sentinel" data-table-id="{cls._escape_attr(table_id)}" data-sentinel-role="end" aria-hidden="true"></div>'
                )
                view_blocks.append(
                    f'<div class="mars-pivot-view" data-y-value="{cls._escape_attr(y_val)}">'
                    f'<div class="mars-view-label">{html.escape(y_val)}</div>{table_html}</div>'
                )

        if not view_blocks:
            return '<div class="mars-empty">No grouped pivot data available.</div>'
        controls_html = f'<div class="mars-inline-controls">{"".join(control_parts)}</div>' if control_parts else ""
        return controls_html + "".join(view_blocks)

    def write_html(
        self,
        path: str = "mars_bin_report.html",
        *,
        report_name: str = "MARS Evaluation Report",
        max_plots: int = 20,
        sort_by: str = "iv",
        ascending: bool = False,
        include_summary: bool = True,
        include_trends: bool = True,
        include_detail: bool = True,
        include_charts: bool = True,
    ) -> None:
        """
        导出自包含的交互式 HTML 报告。

        Parameters
        ----------
        path : str, default "mars_bin_report.html"
            输出文件路径。
        report_name : str, default "MARS Evaluation Report"
            HTML 页面标题与报告名称。
        max_plots : int, default 20
            图表区域最多展示的特征数量。
        sort_by : str, default "iv"
            图表和汇总视图默认使用的排序指标。
        ascending : bool, default False
            是否按 ``sort_by`` 升序排列。
        include_summary : bool, default True
            是否包含汇总表区域。
        include_trends : bool, default True
            是否包含趋势分析区域。
        include_detail : bool, default True
            是否包含分箱明细区域。
        include_charts : bool, default True
            是否包含图表区域。

        Notes
        -----
        导出的 HTML 为单文件报告，适合脱离 Notebook 独立分享或归档。
        """
        return self._write_html_v2(
            path=path,
            report_name=report_name,
            max_plots=max_plots,
            sort_by=sort_by,
            ascending=ascending,
            include_summary=include_summary,
            include_trends=include_trends,
            include_detail=include_detail,
            include_charts=include_charts,
        )
        '''
        summary_pd = _as_pandas_frame(self.summary_table).copy()
        detail_pd = _as_pandas_frame(self.detail_table).copy()
        trend_pd_map = {metric: _as_pandas_frame(df).copy() for metric, df in self.trend_tables.items()}

        n_features = len(summary_pd) if not summary_pd.empty else detail_pd["feature"].nunique() if "feature" in detail_pd.columns else 0
        group_label = self.group_col if self.group_col else "None (Total Only)"
        _ = include_detail  # Kept for backward compatibility; HTML export no longer renders detail table.

        html_parts: List[str] = []
        nav_items: List[Tuple[str, str]] = []

        if include_summary and not summary_pd.empty:
            summary_df = summary_pd.copy()
            if sort_by in summary_df.columns:
                summary_df = summary_df.sort_values(sort_by, ascending=ascending)
            summary_table_html = self._build_sortable_table_html(
                summary_df,
                "mars-summary-table",
                search_placeholder="Search summary table...",
            )
            html_parts.append(
                self._wrap_html_section(
                    "Summary",
                    summary_table_html,
                    "summary-section",
                    subtitle="Feature-level ranking and monitoring summary.",
                )
            )
            nav_items.append(("summary-section", "Summary"))

        general_trend_metrics = [
            metric for metric in ["psi", "auc", "ks", "iv", "risk_corr"]
            if metric in trend_pd_map
        ]
        if include_trends and general_trend_metrics:
            trend_blocks: List[str] = []
            for metric in general_trend_metrics:
                trend_df = self._reorder_group_columns(
                    trend_pd_map[metric].copy(),
                    ["feature", "dtype"],
                )
                trend_blocks.append(
                    self._build_metric_table_block(
                        metric.upper(),
                        trend_df,
                        f"mars-trend-{self._slugify(metric)}",
                        search_placeholder=f"Search {metric} trend...",
                    )
                )
            html_parts.append(
                self._wrap_html_section(
                    "Trend Tables",
                    "".join(trend_blocks),
                    "trend-section",
                    subtitle="Core cross-period monitoring metrics, excluding bad_rate.",
                )
            )
            nav_items.append(("trend-section", "Trends"))

        if include_trends and "missing" in trend_pd_map:
            missing_df = self._reorder_group_columns(
                trend_pd_map["missing"].copy(),
                ["feature", "dtype"],
            )
            html_parts.append(
                self._wrap_html_section(
                    "Missing Trend",
                    self._build_sortable_table_html(
                        missing_df,
                        "mars-missing-trend",
                        search_placeholder="Search missing trend...",
                    ),
                    "missing-section",
                    subtitle="Feature-level missing-rate trend across groups.",
                )
            )
            nav_items.append(("missing-section", "Missing"))

        if include_trends and "lift" in trend_pd_map:
            lift_df = self._reorder_group_columns(
                trend_pd_map["lift"].copy(),
                ["feature", "dtype"],
            )
            html_parts.append(
                self._wrap_html_section(
                    "Lift Trend",
                    self._build_sortable_table_html(
                        lift_df,
                        "mars-lift-trend",
                        search_placeholder="Search lift trend...",
                    ),
                    "lift-section",
                    subtitle="Feature-level max-lift trend across groups.",
                )
            )
            nav_items.append(("lift-section", "Lift"))

        if not detail_pd.empty:
            pivot_group_col = self.group_col or "mars_group"
            pivot_body = self._build_pivot_section_html(detail_pd, group_col=pivot_group_col)
            html_parts.append(
                self._wrap_html_section(
                    "Grouped Pivot",
                    pivot_body,
                    "pivot-section",
                    subtitle="Pivoted monitoring view aligned with the Excel report structure.",
                    open_by_default=False,
                )
            )
            nav_items.append(("pivot-section", "Pivot"))

        if include_charts and not detail_pd.empty:
            chart_y_values = (
                [str(v) for v in detail_pd["y"].dropna().astype(str).drop_duplicates().tolist()]
                if "y" in detail_pd.columns and detail_pd["y"].notna().any()
                else ["Target"]
            )
            chart_controls = ""
            if len(chart_y_values) > 1:
                chart_options = "".join(
                    f'<option value="{html.escape(y_val)}">{html.escape(y_val)}</option>'
                    for y_val in chart_y_values
                )
                chart_controls = (
                    f'<div class="mars-inline-controls">'
                    f'<label class="mars-select-group">Chart Target'
                    f'<select id="mars-chart-target" onchange="marsUpdateChartViews()">{chart_options}</select>'
                    f'</label></div>'
                )

            chart_views: List[str] = []
            try:
                from mars.utils.plotter import MarsPlotter

                for y_val in chart_y_values:
                    if "y" in detail_pd.columns:
                        chart_detail_pd = detail_pd[detail_pd["y"].astype(str) == y_val].copy()
                    else:
                        chart_detail_pd = detail_pd.copy()

                    if "target" in summary_pd.columns:
                        chart_summary_pd = summary_pd[summary_pd["target"].astype(str) == y_val].copy()
                    else:
                        chart_summary_pd = summary_pd.copy()

                    chart_sort_col = self._resolve_chart_sort_column(chart_summary_pd, sort_by)
                    if not chart_summary_pd.empty and chart_sort_col:
                        chart_summary_pd = chart_summary_pd.sort_values(chart_sort_col, ascending=ascending)

                    if not chart_summary_pd.empty and "feature" in chart_summary_pd.columns:
                        chart_features = chart_summary_pd["feature"].drop_duplicates().tolist()[:max_plots]
                    else:
                        chart_features = chart_detail_pd["feature"].drop_duplicates().tolist()[:max_plots]

                    chart_cards: List[str] = []
                    for feature in chart_features:
                        block_html = MarsPlotter.render_feature_binning_risk_trend_html(
                            df_detail=chart_detail_pd,
                            feature=feature,
                            group_col=self.group_col or "mars_group",
                            target_name=y_val,
                            dpi=150,
                        )
                        if not block_html:
                            continue
                        chart_cards.append(
                            f'<article class="mars-chart-card" data-search-text="{html.escape((feature + " " + y_val).lower())}">'
                            f'<h4>{html.escape(feature)}</h4>{block_html}</article>'
                        )

                    if not chart_cards:
                        chart_cards.append('<div class="mars-empty">No chart data available for this target.</div>')

                    chart_views.append(
                        f'<div class="mars-chart-view" data-y-value="{html.escape(y_val)}">'
                        f'{"".join(chart_cards)}</div>'
                    )
            except Exception as e:
                logger.warning("HTML chart rendering skipped due to error: %s", e)

            if chart_views:
                html_parts.append(
                    self._wrap_html_section(
                        "Charts",
                        chart_controls + "".join(chart_views),
                        "chart-section",
                        subtitle="Risk trend charts rendered from the shared plotting path.",
                        open_by_default=False,
                    )
                )
                nav_items.append(("chart-section", "Charts"))

        nav_html = "".join(
            f'<a href="#{html.escape(section_id)}">{html.escape(label)}</a>'
            for section_id, label in nav_items
        )

        page_html = """
        <!DOCTYPE html>
        <html lang="zh">
        <head>
            <meta charset="utf-8" />
            <meta name="viewport" content="width=device-width, initial-scale=1" />
            <title>{html.escape(safe_report_name)}</title>
            <style>
                :root {
                    --bg: #f4f1ea;
                    --panel: #fffdf8;
                    --panel-alt: #f9f5ee;
                    --ink: #1f2933;
                    --muted: #677483;
                    --line: #d9d1c3;
                    --line-soft: #ebe4d8;
                    --accent: #9b6b3d;
                    --accent-soft: #efe2d0;
                    --danger: #b63a3a;
                }
                body {
                    margin: 0;
                    font-family: "Segoe UI", "PingFang SC", "Microsoft YaHei", sans-serif;
                    background: radial-gradient(circle at top right, #efe8dd 0%, var(--bg) 34%, #f6f3ee 100%);
                    color: var(--ink);
                }
                .mars-page {
                    max-width: 1480px;
                    margin: 0 auto;
                    padding: 24px 24px 40px 24px;
                }
                .mars-hero {
                    background: linear-gradient(135deg, rgba(255, 253, 248, 0.96), rgba(249, 245, 238, 0.98));
                    border: 1px solid var(--line);
                    border-radius: 22px;
                    padding: 24px 26px;
                    box-shadow: 0 18px 50px rgba(76, 56, 33, 0.08);
                    margin-bottom: 18px;
                }
                .mars-hero h1 {
                    margin: 0 0 8px 0;
                    font-size: 30px;
                    letter-spacing: 0.01em;
                }
                .mars-hero p {
                    margin: 0;
                    color: var(--muted);
                    font-size: 14px;
                }
                .mars-meta {
                    display: flex;
                    flex-wrap: wrap;
                    gap: 10px;
                    margin-top: 14px;
                }
                .mars-pill {
                    background: rgba(255, 255, 255, 0.92);
                    border: 1px solid var(--line);
                    border-radius: 999px;
                    padding: 7px 12px;
                    font-size: 13px;
                }
                .mars-global-tools {
                    margin-top: 16px;
                    display: grid;
                    grid-template-columns: minmax(280px, 460px) auto auto;
                    gap: 10px;
                    align-items: center;
                }
                .mars-filter-input,
                .mars-select-group select,
                .mars-clear-button {
                    border: 1px solid var(--line);
                    border-radius: 12px;
                    background: rgba(255, 255, 255, 0.9);
                    font-size: 14px;
                }
                .mars-filter-input {
                    padding: 10px 12px;
                    width: 100%;
                    box-sizing: border-box;
                }
                .mars-toggle {
                    display: inline-flex;
                    align-items: center;
                    gap: 8px;
                    font-size: 13px;
                    color: var(--muted);
                }
                .mars-clear-button {
                    padding: 10px 14px;
                    cursor: pointer;
                }
                .mars-search-error {
                    color: var(--danger);
                    font-size: 12px;
                    min-height: 16px;
                }
                .mars-nav {
                    display: flex;
                    flex-wrap: wrap;
                    gap: 10px;
                    margin: 16px 0 20px 0;
                }
                .mars-nav a {
                    text-decoration: none;
                    color: #6d4b2f;
                    background: rgba(255, 253, 248, 0.82);
                    border: 1px solid var(--line);
                    padding: 8px 12px;
                    border-radius: 999px;
                    font-size: 13px;
                }
                .mars-section {
                    background: rgba(255, 253, 248, 0.97);
                    border: 1px solid var(--line);
                    border-radius: 18px;
                    margin-bottom: 18px;
                    overflow: hidden;
                }
                .mars-section > summary {
                    cursor: pointer;
                    list-style: none;
                    font-weight: 700;
                    padding: 16px 18px;
                    background: rgba(249, 245, 238, 0.92);
                    border-bottom: 1px solid var(--line-soft);
                }
                .mars-section > summary::-webkit-details-marker {
                    display: none;
                }
                .mars-section-subtitle {
                    padding: 12px 18px 0 18px;
                    font-size: 13px;
                    color: var(--muted);
                }
                .mars-section-body {
                    padding: 14px 18px 18px 18px;
                }
                .mars-metric-block {
                    background: rgba(255, 255, 255, 0.7);
                    border: 1px solid var(--line-soft);
                    border-radius: 14px;
                    padding: 14px;
                    margin-bottom: 14px;
                }
                .mars-metric-block h3,
                .mars-view-label {
                    margin: 0 0 10px 0;
                    font-size: 15px;
                    color: #7a5635;
                }
                .mars-inline-controls {
                    display: flex;
                    flex-wrap: wrap;
                    gap: 12px;
                    margin-bottom: 12px;
                }
                .mars-select-group {
                    display: inline-flex;
                    align-items: center;
                    gap: 8px;
                    font-size: 13px;
                    color: var(--muted);
                }
                .mars-select-group select {
                    padding: 8px 10px;
                }
                .mars-table-wrap {
                    width: 100%;
                }
                .mars-table-toolbar {
                    display: grid;
                    grid-template-columns: minmax(220px, 420px);
                    gap: 6px;
                    margin-bottom: 10px;
                }
                .mars-table-scroll {
                    overflow-x: auto;
                    border: 1px solid var(--line-soft);
                    border-radius: 14px;
                    background: rgba(255, 255, 255, 0.76);
                }
                .mars-data-table {
                    width: 100%;
                    border-collapse: separate;
                    border-spacing: 0;
                    font-size: 13px;
                }
                .mars-th,
                .mars-td {
                    border-bottom: 1px solid var(--line-soft);
                    padding: 8px 10px;
                    vertical-align: top;
                    text-align: left;
                    white-space: nowrap;
                }
                .mars-th {
                    position: sticky;
                    top: 0;
                    background: rgba(249, 245, 238, 0.98);
                    z-index: 1;
                }
                .mars-td.is-numeric,
                .mars-th.is-numeric {
                    text-align: right;
                }
                .mars-sort-button {
                    width: 100%;
                    border: 0;
                    background: transparent;
                    padding: 0;
                    margin: 0;
                    color: inherit;
                    font: inherit;
                    display: inline-flex;
                    align-items: center;
                    justify-content: space-between;
                    gap: 8px;
                    cursor: pointer;
                }
                .mars-sort-indicator::before {
                    content: "↕";
                    color: #b29a7f;
                    font-size: 11px;
                }
                th[data-sort-dir="asc"] .mars-sort-indicator::before {
                    content: "↑";
                    color: var(--accent);
                }
                th[data-sort-dir="desc"] .mars-sort-indicator::before {
                    content: "↓";
                    color: var(--accent);
                }
                .mars-pivot-view,
                .mars-chart-view {
                    margin-bottom: 14px;
                }
                .mars-chart-card {
                    border: 1px solid var(--line-soft);
                    border-radius: 18px;
                    padding: 14px;
                    background: rgba(255, 255, 255, 0.76);
                    margin-bottom: 14px;
                }
                .mars-chart-card h4 {
                    margin: 0 0 10px 0;
                    font-size: 16px;
                }
                .mars-empty {
                    border: 1px dashed var(--line);
                    border-radius: 14px;
                    background: rgba(255, 255, 255, 0.7);
                    color: var(--muted);
                    padding: 16px;
                    font-size: 13px;
                }
                .mars-footnote {
                    font-size: 12px;
                    color: var(--muted);
                    margin-top: 12px;
                }
            </style>
        </head>
        <body>
            <div class="mars-page">
                <div class="mars-hero">
                    <h1>MARS Evaluation Report</h1>
                    <p>Lightweight monitoring HTML with shared charts, sortable tables, grouped pivot views, and dual-mode search.</p>
                    <div class="mars-meta">
                        __META_PILLS__
                    </div>
                    <div class="mars-global-tools">
                        <input id="mars-global-search" class="mars-filter-input" type="search" placeholder="Global search across tables and charts..." oninput="marsSetGlobalQuery(this.value)" />
                        <label class="mars-toggle"><input id="mars-regex-mode" type="checkbox" onchange="marsSetRegexMode(this.checked)" /> Regex Mode</label>
                        <button type="button" class="mars-clear-button" onclick="marsClearGlobalSearch()">Clear Search</button>
                    </div>
                    <div id="mars-global-error" class="mars-search-error"></div>
                </div>

                <div class="mars-nav">__NAV_HTML__</div>

                __BODY_HTML__

                <div class="mars-footnote">
                    HTML export is self-contained. `detail_table` is preserved in Python/Excel workflows but intentionally omitted from the HTML page.
                </div>
            </div>

            <script>
                const marsState = {
                    globalQuery: "",
                    regexMode: false,
                    localQueries: {}
                };

                function marsBuildMatcher(query) {
                    const q = (query || "").trim();
                    if (!q) {
                        return { ok: true, match: () => true };
                    }

                    if (marsState.regexMode) {
                        try {
                            const regex = new RegExp(q, "i");
                            return { ok: true, match: (text) => regex.test(text || "") };
                        } catch (err) {
                            return { ok: false, error: err.message };
                        }
                    }

                    const terms = q.toLowerCase().split(/\\s+/).filter(Boolean);
                    return {
                        ok: true,
                        match: (text) => {
                            const haystack = (text || "").toLowerCase();
                            return terms.every((term) => haystack.includes(term));
                        }
                    };
                }

                function marsSetError(id, message) {
                    const node = document.getElementById(id);
                    if (node) {
                        node.textContent = message || "";
                    }
                }

                function marsSetGlobalQuery(value) {
                    marsState.globalQuery = value || "";
                    marsRefreshFilters();
                }

                function marsSetLocalQuery(tableId, value) {
                    marsState.localQueries[tableId] = value || "";
                    marsApplyTableFilter(tableId);
                }

                function marsSetRegexMode(enabled) {
                    marsState.regexMode = !!enabled;
                    marsRefreshFilters();
                }

                function marsClearGlobalSearch() {
                    const input = document.getElementById("mars-global-search");
                    if (input) {
                        input.value = "";
                    }
                    marsState.globalQuery = "";
                    marsRefreshFilters();
                }

                function marsApplyTableFilter(tableId) {
                    const table = document.getElementById(tableId);
                    if (!table) return;

                    const globalMatcher = marsBuildMatcher(marsState.globalQuery);
                    if (!globalMatcher.ok) {
                        marsSetError("mars-global-error", `Invalid regex: ${globalMatcher.error}`);
                        return;
                    }
                    marsSetError("mars-global-error", "");

                    const localQuery = marsState.localQueries[tableId] || "";
                    const localMatcher = marsBuildMatcher(localQuery);
                    if (!localMatcher.ok) {
                        marsSetError(`${tableId}-error`, `Invalid regex: ${localMatcher.error}`);
                        return;
                    }
                    marsSetError(`${tableId}-error`, "");

                    const rows = table.querySelectorAll("tbody tr");
                    rows.forEach((row) => {
                        const text = row.dataset.searchText || row.textContent || "";
                        const keep = globalMatcher.match(text) && localMatcher.match(text);
                        row.style.display = keep ? "" : "none";
                    });
                }

                function marsSortTable(tableId, trigger) {
                    const table = document.getElementById(tableId);
                    if (!table) return;

                    const th = trigger.closest("th");
                    const headers = Array.from(table.querySelectorAll("thead th"));
                    const colIndex = headers.indexOf(th);
                    if (colIndex < 0) return;

                    const currentCol = table.dataset.sortCol;
                    let nextDir = "asc";
                    if (currentCol === String(colIndex)) {
                        nextDir = table.dataset.sortDir === "asc" ? "desc" : "asc";
                    }

                    const tbody = table.querySelector("tbody");
                    const rows = Array.from(tbody.querySelectorAll("tr"));
                    const sortType = th.dataset.sortType || "text";

                    rows.sort((rowA, rowB) => {
                        const cellA = rowA.children[colIndex];
                        const cellB = rowB.children[colIndex];
                        const valA = cellA ? (cellA.dataset.sortValue || "") : "";
                        const valB = cellB ? (cellB.dataset.sortValue || "") : "";

                        if (sortType === "number") {
                            const numA = Number(valA);
                            const numB = Number(valB);
                            const safeA = Number.isFinite(numA) ? numA : (nextDir === "asc" ? Infinity : -Infinity);
                            const safeB = Number.isFinite(numB) ? numB : (nextDir === "asc" ? Infinity : -Infinity);
                            return nextDir === "asc" ? safeA - safeB : safeB - safeA;
                        }

                        return nextDir === "asc"
                            ? valA.localeCompare(valB, undefined, { numeric: true, sensitivity: "base" })
                            : valB.localeCompare(valA, undefined, { numeric: true, sensitivity: "base" });
                    });

                    rows.forEach((row) => tbody.appendChild(row));
                    table.dataset.sortCol = String(colIndex);
                    table.dataset.sortDir = nextDir;
                    headers.forEach((header, index) => {
                        header.dataset.sortDir = index === colIndex ? nextDir : "";
                    });

                    marsApplyTableFilter(tableId);
                }

                function marsUpdatePivotViews() {
                    const targetSelect = document.getElementById("mars-pivot-target");
                    const binTypeSelect = document.getElementById("mars-pivot-bin-type");
                    const targetValue = targetSelect ? targetSelect.value : null;
                    const binTypeValue = binTypeSelect ? binTypeSelect.value : "__all__";

                    document.querySelectorAll(".mars-pivot-view").forEach((view) => {
                        const sameTarget = !targetValue || view.dataset.yValue === targetValue;
                        const sameBinType = !binTypeSelect || binTypeValue === "__all__" || view.dataset.binTypeValue === binTypeValue;
                        view.style.display = sameTarget && sameBinType ? "" : "none";
                    });
                }

                function marsUpdateChartViews() {
                    const targetSelect = document.getElementById("mars-chart-target");
                    const targetValue = targetSelect ? targetSelect.value : null;
                    const globalMatcher = marsBuildMatcher(marsState.globalQuery);
                    if (!globalMatcher.ok) {
                        marsSetError("mars-global-error", `Invalid regex: ${globalMatcher.error}`);
                        return;
                    }

                    document.querySelectorAll(".mars-chart-view").forEach((view) => {
                        const visibleTarget = !targetValue || view.dataset.yValue === targetValue;
                        view.style.display = visibleTarget ? "" : "none";
                        if (!visibleTarget) return;

                        view.querySelectorAll(".mars-chart-card").forEach((card) => {
                            const text = card.dataset.searchText || card.textContent || "";
                            card.style.display = globalMatcher.match(text) ? "" : "none";
                        });
                    });
                }

                function marsRefreshFilters() {
                    marsUpdatePivotViews();
                    marsUpdateChartViews();
                    document.querySelectorAll("table.mars-data-table").forEach((table) => {
                        marsApplyTableFilter(table.id);
                    });
                }

                window.addEventListener("DOMContentLoaded", marsRefreshFilters);
            </script>
        </body>
        </html>
        """

        meta_pills = "".join([
            f'<div class="mars-pill">Features: {n_features}</div>',
            f'<div class="mars-pill">Trend Metrics: {len(trend_pd_map)}</div>',
            f'<div class="mars-pill">Group By: {html.escape(str(group_label))}</div>',
        ])
        page_html = page_html.replace("__META_PILLS__", meta_pills)
        page_html = page_html.replace("__NAV_HTML__", nav_html)
        page_html = page_html.replace("__BODY_HTML__", "".join(html_parts))

        with open(path, "w", encoding="utf-8") as f:
            f.write(page_html)

        logger.info("Exported evaluation report to HTML: %s", path)

        '''

    def _write_html_v2(
        self,
        *,
        path: str,
        report_name: str,
        max_plots: int,
        sort_by: str,
        ascending: bool,
        include_summary: bool,
        include_trends: bool,
        include_detail: bool,
        include_charts: bool,
    ) -> None:
        summary_pd = _as_pandas_frame(self.summary_table).copy()
        detail_pd = _as_pandas_frame(self.detail_table).copy()
        trend_pd_map = {metric: _as_pandas_frame(df).copy() for metric, df in self.trend_tables.items()}
        missing_by_day_pd = _as_pandas_frame(self.missing_by_day_table).copy() if self.missing_by_day_table is not None else None

        _ = include_detail
        feature_sources = dict(self.feature_data_source or {})
        if not feature_sources and not summary_pd.empty and {"feature", "data_source"}.issubset(summary_pd.columns):
            feature_sources = dict(zip(summary_pd["feature"].astype(str), summary_pd["data_source"].astype(str)))
        if not feature_sources and not detail_pd.empty and {"feature", "data_source"}.issubset(detail_pd.columns):
            source_df = detail_pd[["feature", "data_source"]].dropna().drop_duplicates()
            feature_sources = dict(zip(source_df["feature"].astype(str), source_df["data_source"].astype(str)))

        n_features = len(summary_pd) if not summary_pd.empty else detail_pd["feature"].nunique() if "feature" in detail_pd.columns else 0
        group_label = self.group_col if self.group_col else "None (Total Only)"
        all_sources = sorted(
            set(feature_sources.values())
            | set(summary_pd["data_source"].astype(str).tolist() if "data_source" in summary_pd.columns else [])
            | set(detail_pd["data_source"].astype(str).tolist() if "data_source" in detail_pd.columns else [])
        )
        safe_report_name = report_name or "MARS Evaluation Report"
        summary_filter_columns = [
            "iv", "ks", "auc", "psi_max", "rc_min",
            "lift_min", "lift_max",
            "missing", "missing_min", "missing_max",
        ]
        feature_jump_html = self._build_feature_jump_html(
            summary_pd["feature"].astype(str).tolist() if "feature" in summary_pd.columns else detail_pd["feature"].astype(str).tolist() if "feature" in detail_pd.columns else []
        )

        html_parts: List[str] = []
        nav_items: List[Tuple[str, str]] = []

        overview_html = self._build_dataset_overview_html(self.report_meta)
        if overview_html:
            html_parts.append(self._wrap_html_section("Dataset Overview", overview_html, "overview-section", subtitle="Dataset context, grouping setup, and target-level baseline stats."))
            nav_items.append(("overview-section", "Overview"))

        if include_summary:
            summary_html = self._build_summary_section_html(
                summary_pd=summary_pd,
                feature_sources=feature_sources,
                sort_by=sort_by,
                ascending=ascending,
            )
            if summary_html:
                html_parts.append(summary_html)
                nav_items.append(("summary-section", "Summary"))

        if include_trends:
            for section_id, label, section_html in self._build_trend_sections_html(
                trend_pd_map=trend_pd_map,
                missing_by_day_pd=missing_by_day_pd,
                feature_sources=feature_sources,
            ):
                html_parts.append(section_html)
                nav_items.append((section_id, label))

        if not detail_pd.empty:
            pivot_body = self._build_grouped_pivot_section_html(detail_pd, group_col=self.group_col or "mars_group", feature_sources=feature_sources)
            html_parts.append(self._wrap_html_section("Grouped Pivot", pivot_body, "pivot-section", subtitle="Binned distribution and risk comparison across groups.", open_by_default=False))
            nav_items.append(("pivot-section", "Grouped Pivot"))

        if include_charts:
            chart_html = self._build_chart_section_html(
                detail_pd=detail_pd,
                summary_pd=summary_pd,
                feature_sources=feature_sources,
                max_plots=max_plots,
                sort_by=sort_by,
                ascending=ascending,
            )
            if chart_html:
                html_parts.append(chart_html)
                nav_items.append(("chart-section", "Charts"))

        nav_html = "".join(f'<a href="#{html.escape(section_id)}">{html.escape(label)}</a>' for section_id, label in nav_items)
        source_options = "".join(
            f'<label class="mars-source-option"><input type="checkbox" class="mars-source-checkbox" '
            f'value="{html.escape(source)}" checked onchange="marsHandleDataSourceToggle()" />'
            f'<span>{html.escape(source)}</span></label>'
            for source in all_sources
        )

        """Legacy inline template retained temporarily during report.py cleanup.
        <!DOCTYPE html>
        <html lang="zh">
        <head>
            <meta charset="utf-8" />
            <meta name="viewport" content="width=device-width, initial-scale=1" />
            <title>{html.escape(safe_report_name)}</title>
            <style>
                :root {{ --bg:#f5f7fb; --panel:#fff; --panel-soft:#f9fbfd; --ink:#203040; --muted:#607080; --line:#d9e3eb; --line-soft:#ebf1f6; --accent:#3b87ad; --danger:#c44f4f; --shadow:0 16px 36px rgba(51,82,108,.08); }}
                body {{ margin:0; font-family:"Segoe UI","PingFang SC","Microsoft YaHei",sans-serif; background:radial-gradient(circle at top right,#edf6fb 0%,#f5f7fb 40%,#f8fbfd 100%); color:var(--ink); }}
                .mars-page {{ max-width:1640px; margin:0 auto; padding:22px; }}
                .mars-hero,.mars-section {{ background:var(--panel); border:1px solid var(--line); border-radius:18px; box-shadow:var(--shadow); }}
                .mars-hero {{ padding:22px 24px; margin-bottom:16px; position:relative; overflow:hidden; }}
                .mars-hero::after {{ content:""; position:absolute; inset:auto -80px -90px auto; width:240px; height:240px; background:radial-gradient(circle, rgba(59,135,173,.14) 0%, rgba(59,135,173,0) 72%); pointer-events:none; }}
                .mars-hero h1 {{ margin:0 0 8px 0; font-size:30px; }}
                .mars-hero p,.mars-footnote,.mars-section-subtitle,.mars-search-error,.mars-view-label,.mars-pivot-source-title {{ color:var(--muted); position:relative; z-index:1; }}
                .mars-meta,.mars-nav,.mars-inline-controls {{ display:flex; flex-wrap:wrap; gap:10px; }}
                .mars-meta {{ margin-top:12px; position:relative; z-index:1; }}
                .mars-pill,.mars-nav a {{ border:1px solid var(--line); background:#f7fbff; border-radius:999px; padding:6px 12px; font-size:13px; color:#36546d; text-decoration:none; }}
                .mars-global-tools {{ margin-top:16px; display:grid; grid-template-columns:minmax(280px,420px) auto minmax(240px,340px) minmax(280px,1fr) auto; gap:10px; align-items:start; position:relative; z-index:1; }}
                .mars-filter-input,.mars-select-group select,.mars-clear-button,.mars-mini-button {{ border:1px solid var(--line); border-radius:12px; background:#fff; font-size:14px; }}
                .mars-filter-input {{ padding:10px 12px; width:100%; box-sizing:border-box; }}
                .mars-search-cluster {{ display:grid; grid-template-columns:minmax(0,1fr) auto; gap:8px; align-items:center; }}
                .mars-select-group {{ display:inline-flex; gap:8px; align-items:center; font-size:13px; }}
                .mars-select-group select {{ padding:8px 10px; }}
                .mars-source-panel {{ border:1px solid var(--line); border-radius:14px; background:#fff; padding:10px 12px; min-width:280px; }}
                .mars-source-header,.mars-source-options {{ display:flex; flex-wrap:wrap; gap:8px; }}
                .mars-source-header {{ align-items:center; justify-content:space-between; margin-bottom:10px; }}
                .mars-source-header strong {{ font-size:13px; color:#355b74; }}
                .mars-source-link {{ border:0; background:transparent; color:var(--accent); cursor:pointer; font-size:12px; padding:0; }}
                .mars-source-option {{ display:inline-flex; align-items:center; gap:6px; border:1px solid var(--line-soft); border-radius:999px; padding:5px 10px; background:#f9fbfe; font-size:13px; }}
                .mars-clear-button,.mars-mini-button {{ padding:9px 12px; cursor:pointer; }}
                .mars-toggle {{ display:inline-flex; align-items:center; gap:8px; font-size:13px; }}
                .mars-nav {{ margin:14px 0 18px 0; }}
                .mars-overview-grid {{ display:grid; grid-template-columns:repeat(auto-fit, minmax(170px, 1fr)); gap:12px; }}
                .mars-kpi-card {{ border:1px solid var(--line-soft); border-radius:14px; background:linear-gradient(180deg,#fbfdff 0%,#f7fbff 100%); padding:14px; }}
                .mars-kpi-label {{ font-size:12px; color:var(--muted); margin-bottom:6px; text-transform:uppercase; letter-spacing:.04em; }}
                .mars-kpi-value {{ font-size:16px; font-weight:700; color:#244258; line-height:1.35; word-break:break-word; }}
                .mars-legend {{ display:flex; flex-wrap:wrap; gap:8px; margin-top:10px; }}
                .mars-legend-chip {{ display:inline-flex; align-items:center; gap:6px; border:1px solid var(--line-soft); border-radius:999px; padding:6px 10px; background:#fff; font-size:12px; color:#436179; }}
                .mars-section {{ margin-bottom:16px; overflow:hidden; }}
                .mars-section>summary,.mars-metric-block>summary {{ cursor:pointer; list-style:none; font-weight:700; }}
                .mars-section>summary {{ padding:16px 18px; background:#f7fbff; border-bottom:1px solid var(--line-soft); }}
                .mars-section>summary::-webkit-details-marker,.mars-metric-block>summary::-webkit-details-marker {{ display:none; }}
                .mars-section-body {{ padding:14px 18px 18px 18px; }}
                .mars-section-subtitle {{ padding:12px 18px 0 18px; font-size:13px; }}
                .mars-metric-block {{ border:1px solid var(--line-soft); border-radius:14px; background:var(--panel-soft); margin-bottom:12px; padding:12px; }}
                .mars-metric-block>summary {{ margin-bottom:10px; color:#355b74; }}
                .mars-table-toolbar {{ display:grid; grid-template-columns:minmax(240px,360px); gap:6px; margin-bottom:10px; }}
                .mars-chart-controls {{ display:grid; grid-template-columns:minmax(240px,360px) auto; gap:10px; align-items:start; }}
                .mars-chart-search {{ min-width:240px; }}
                .mars-summary-filter {{ border:1px solid var(--line-soft); border-radius:14px; background:#fbfdff; padding:12px; margin-bottom:10px; }}
                .mars-summary-filter-label {{ display:block; margin-bottom:8px; font-size:13px; font-weight:600; color:#355b74; }}
                .mars-table-scroll {{ overflow:auto; border:1px solid var(--line-soft); border-radius:14px; background:#fff; }}
                .mars-data-table {{ width:max-content; min-width:100%; border-collapse:separate; border-spacing:0; font-size:13px; }}
                .mars-th,.mars-td {{ border-bottom:1px solid var(--line-soft); padding:8px 10px; white-space:nowrap; text-align:left; vertical-align:top; }}
                .mars-th {{ position:sticky; top:0; background:#eef6fb; z-index:1; }}
                .mars-feature-col {{ min-width:var(--mars-feature-col-width, 220px); width:var(--mars-feature-col-width, 220px); max-width:var(--mars-feature-col-width, 220px); box-sizing:border-box; overflow:hidden; text-overflow:ellipsis; background-clip:padding-box; }}
                .mars-secondary-col {{ min-width:var(--mars-secondary-col-width, 110px); width:var(--mars-secondary-col-width, 110px); max-width:var(--mars-secondary-col-width, 110px); box-sizing:border-box; overflow:hidden; text-overflow:ellipsis; background-clip:padding-box; }}
                .mars-bin-col {{ min-width:var(--mars-bin-col-width, 140px); width:var(--mars-bin-col-width, 140px); max-width:var(--mars-bin-col-width, 140px); box-sizing:border-box; overflow:hidden; text-overflow:ellipsis; background-clip:padding-box; }}
                .mars-data-table .mars-td.mars-feature-col, .mars-data-table .mars-td.mars-secondary-col, .mars-pivot-table .mars-td.mars-bin-col {{ background:#fff; }}
                .mars-data-table .mars-th.mars-feature-col, .mars-data-table .mars-th.mars-secondary-col, .mars-pivot-table .mars-th.mars-bin-col {{ background:#eef6fb; }}
                .mars-data-table .mars-th.mars-feature-col, .mars-data-table .mars-td.mars-feature-col {{ position:sticky; left:0; z-index:3; box-shadow:2px 0 0 rgba(217,227,235,.85); }}
                .mars-data-table .mars-th.mars-secondary-col, .mars-data-table .mars-td.mars-secondary-col {{ position:sticky; left:var(--mars-feature-col-width, 220px); z-index:2; box-shadow:2px 0 0 rgba(217,227,235,.72); }}
                .mars-pivot-table .mars-th-sticky-left-2, .mars-pivot-table .mars-bin-col {{ position:sticky; left:var(--mars-feature-col-width, 220px); box-shadow:2px 0 0 rgba(217,227,235,.85); background-clip:padding-box; }}
                .mars-th.is-numeric,.mars-td.is-numeric {{ text-align:right; }}
                .mars-sort-button {{ width:100%; min-width:0; overflow:hidden; border:0; background:transparent; padding:0; margin:0; color:inherit; font:inherit; display:inline-flex; align-items:center; justify-content:space-between; gap:8px; cursor:pointer; }}
                .mars-sort-label {{ display:block; min-width:0; overflow:hidden; text-overflow:ellipsis; }}
                .mars-cell-text {{ display:block; overflow:hidden; text-overflow:ellipsis; }}
                .mars-th.mars-feature-col, .mars-th.mars-bin-col {{ position:sticky; }}
                .mars-th.mars-feature-col {{ padding-right:18px; }}
                .mars-resize-handle {{ position:absolute; top:0; right:0; width:10px; height:100%; cursor:col-resize; user-select:none; touch-action:none; }}
                .mars-resize-handle::after {{ content:""; position:absolute; top:20%; bottom:20%; left:4px; width:2px; border-radius:2px; background:rgba(53,91,116,.22); }}
                .mars-feature-jump {{ min-width:240px; }}
                .mars-pivot-table .mars-th, .mars-pivot-table .mars-td {{ position:relative; background-clip:padding-box; }}
                .mars-pivot-table .mars-th.mars-feature-col {{ z-index:5; }}
                .mars-pivot-table .mars-td.mars-feature-col {{ z-index:4; }}
                .mars-pivot-table .mars-th.mars-bin-col {{ z-index:4; padding-right:18px; }}
                .mars-pivot-table .mars-td.mars-bin-col {{ z-index:3; }}
                .mars-pivot-table .mars-td.mars-feature-col, .mars-pivot-table .mars-td.mars-bin-col {{ background:#fff; }}
                .mars-jump-highlight {{ animation:mars-jump-pulse 1.2s ease-out 1; }}
                .mars-jump-highlight-cell {{ animation:mars-jump-pulse 1.2s ease-out 1; }}
                @keyframes mars-jump-pulse {{ 0% {{ box-shadow:0 0 0 0 rgba(59,135,173,.35); background-color:rgba(208,234,246,.68); }} 100% {{ box-shadow:0 0 0 14px rgba(59,135,173,0); background-color:inherit; }} }}
                .mars-sort-indicator::before {{ content:"↕"; color:#8aa1b3; font-size:11px; }}
                th[data-sort-dir="asc"] .mars-sort-indicator::before {{ content:"↑"; color:var(--accent); }}
                th[data-sort-dir="desc"] .mars-sort-indicator::before {{ content:"↓"; color:var(--accent); }}
                .mars-empty {{ border:1px dashed var(--line); border-radius:14px; padding:16px; background:#fbfdff; font-size:13px; }}
                .mars-sort-indicator::before {{ content:"\\2195"; color:#8aa1b3; font-size:11px; }}
                th[data-sort-dir="asc"] .mars-sort-indicator::before {{ content:"\\2191"; color:var(--accent); }}
                th[data-sort-dir="desc"] .mars-sort-indicator::before {{ content:"\\2193"; color:var(--accent); }}
                .mars-chart-card {{ border:1px solid var(--line-soft); border-radius:14px; background:#fff; padding:12px; margin-bottom:12px; box-shadow:0 8px 20px rgba(51,82,108,.05); }}
                .mars-pivot-source-title-cell {{ background:#edf6fb; color:#355b74; font-weight:700; letter-spacing:.02em; }}
                .mars-pivot-feature {{ font-weight:600; color:#2f495e; }}
                .mars-pivot-feature-blank .mars-cell-text {{ visibility:hidden; }}
                .mars-pivot-spacer-row td {{ border-bottom:0; padding:5px 0; background:linear-gradient(180deg,transparent 0%,rgba(233,239,245,.65) 100%); }}
                .mars-chart-card h4 {{ margin:0 0 10px 0; font-size:16px; }}
                .mars-footnote {{ font-size:12px; margin-top:12px; }}
            </style>
        </head>
        <body>
            <div class="mars-page">
                <div class="mars-hero">
                    <h1>{html.escape(safe_report_name)}</h1>
                    <p>Interactive monitoring report with source-aware tables, Excel-like color scales, grouped pivot views, and shared trend charts.</p>
                <div class="mars-meta">
                    <div class="mars-pill">Features: {n_features}</div>
                    <div class="mars-pill">Trend Metrics: {len(trend_pd_map)}</div>
                    <div class="mars-pill">Group By: {html.escape(str(group_label))}</div>
                </div>
                <div class="mars-global-tools">
                    <div class="mars-search-cluster">
                        <input id="mars-global-search" class="mars-filter-input" type="search" placeholder="Global search across tables and charts..." oninput="marsSetGlobalQuery(this.value)" />
                        <button type="button" class="mars-clear-button" onclick="marsClearGlobalSearch()">Clear Search</button>
                    </div>
                    <label class="mars-toggle"><input id="mars-regex-mode" type="checkbox" onchange="marsSetRegexMode(this.checked)" /> Regex Mode</label>
                    {feature_jump_html}
                    <div class="mars-source-panel">
                        <div class="mars-source-header">
                            <strong>Data Source</strong>
                            <div>
                                <button type="button" class="mars-source-link" onclick="marsSelectAllSources()">All</button>
                                    <button type="button" class="mars-source-link" onclick="marsClearSources()">Clear</button>
                                </div>
                        </div>
                        <div id="mars-data-source-group" class="mars-source-options">{source_options}</div>
                    </div>
                    <button type="button" class="mars-clear-button" onclick="marsExportFeatures()">Export Feature List</button>
                </div>
                    <div id="mars-global-error" class="mars-search-error"></div>
                </div>
                <div class="mars-nav">{nav_html}</div>
                {"".join(html_parts)}
                <div class="mars-footnote">HTML export is self-contained. detail_table remains available in Python and Excel workflows.</div>
            </div>
            <script>
                const marsSummaryFilterColumns = new Set({json.dumps(summary_filter_columns, ensure_ascii=False)});
                const marsState = {{
                    globalQuery:"",
                    regexMode:false,
                    localQueries:{{}},
                    activeTableId:null,
                    selectedSources:[],
                    appliedSummaryExpression:"",
                    summaryAllowedFeatures:null,
                    refreshScheduled:false,
                    refreshFrameId:null,
                    postPaintFrameId:null,
                    refreshTimerId:null,
                    resizeState:null,
                    resizeFrameScheduled:false
                }};
                function marsBuildMatcher(query) {{ const q=(query||"").trim(); if(!q) return {{ok:true,match:()=>true}}; if(marsState.regexMode) {{ try {{ const regex=new RegExp(q,"i"); return {{ok:true,match:(text)=>regex.test(text||"")}}; }} catch(err) {{ return {{ok:false,error:err.message}}; }} }} const terms=q.toLowerCase().split(/\\s+/).filter(Boolean); return {{ok:true,match:(text)=>terms.every((term)=>(text||"").toLowerCase().includes(term))}}; }}
                function marsSetError(id, message) {{ const node=document.getElementById(id); if(node) node.textContent=message||""; }}
                function marsNormalizeFeatureValue(value) {{ return (value||"").trim().toLowerCase(); }}
                function marsSetActiveScope(scopeId) {{ marsState.activeTableId=scopeId||null; }}
                function marsSetActiveTable(tableId) {{ marsSetActiveScope(tableId); }}
                function marsQueueRefresh(delayMs=0) {{
                    if(marsState.refreshFrameId) window.cancelAnimationFrame(marsState.refreshFrameId);
                    if(marsState.postPaintFrameId) window.cancelAnimationFrame(marsState.postPaintFrameId);
                    if(marsState.refreshTimerId) window.clearTimeout(marsState.refreshTimerId);
                    marsState.refreshScheduled = true;
                    marsState.refreshFrameId = window.requestAnimationFrame(() => {{
                        marsState.refreshFrameId = null;
                        marsState.postPaintFrameId = window.requestAnimationFrame(() => {{
                            marsState.postPaintFrameId = null;
                            marsState.refreshTimerId = window.setTimeout(() => {{
                                marsState.refreshTimerId = null;
                                marsState.refreshScheduled = false;
                                marsRefreshFilters();
                            }}, Math.max(0, Number(delayMs) || 0));
                        }});
                    }});
                }}
                function marsQueueTextRefresh() {{ marsQueueRefresh(80); }}
                function marsSetGlobalQuery(value) {{ marsState.globalQuery=value||""; marsQueueTextRefresh(); }}
                function marsSetLocalQuery(scopeId, value) {{ marsState.localQueries[scopeId]=value||""; marsState.activeTableId=scopeId; marsQueueTextRefresh(); }}
                function marsSetRegexMode(enabled) {{ marsState.regexMode=!!enabled; marsQueueRefresh(); }}
                function marsSetDataSources() {{ const boxes=Array.from(document.querySelectorAll(".mars-source-checkbox")); marsState.selectedSources=boxes.filter((box)=>box.checked).map((box)=>box.value); marsQueueRefresh(); }}
                function marsHandleDataSourceToggle() {{ marsSetDataSources(); }}
                function marsHandlePivotTargetChange() {{ marsQueueRefresh(); }}
                function marsHandleChartTargetChange() {{ marsQueueRefresh(); }}
                function marsSelectAllSources() {{ document.querySelectorAll(".mars-source-checkbox").forEach((box)=>{{ box.checked=true; }}); marsSetDataSources(); }}
                function marsClearSources() {{ document.querySelectorAll(".mars-source-checkbox").forEach((box)=>{{ box.checked=false; }}); marsSetDataSources(); }}
                function marsClearGlobalSearch() {{ const input=document.getElementById("mars-global-search"); if(input) input.value=""; marsState.globalQuery=""; marsQueueTextRefresh(); }}
                function marsTokenizeSummaryExpression(expr) {{
                    const text=(expr||"").trim();
                    if(!text) return {{ok:true,tokens:[]}};
                    const tokenPattern=/\\s*(>=|<=|==|!=|>|<|\\&|\\||\\(|\\)|-?(?:\\d+\\.\\d*|\\d*\\.\\d+|\\d+)(?:[eE][+-]?\\d+)?|[A-Za-z_][A-Za-z0-9_]*)\\s*/gy;
                    const tokens=[];
                    let match;
                    while((match=tokenPattern.exec(text))!==null) {{
                        tokens.push(match[1]);
                    }}
                    if(tokenPattern.lastIndex!==text.length) {{
                        return {{ok:false,error:"Invalid expression syntax."}};
                    }}
                    return {{ok:true,tokens}};
                }}
                function marsParseSummaryExpression(expr) {{
                    const tokenResult=marsTokenizeSummaryExpression(expr);
                    if(!tokenResult.ok) return tokenResult;
                    const tokens=tokenResult.tokens;
                    if(!tokens.length) return {{ok:true,ast:null}};
                    let idx=0;
                    function peek() {{ return tokens[idx]; }}
                    function consume(expected) {{
                        const token=tokens[idx];
                        if(expected && token!==expected) throw new Error(`Expected '${{expected}}'`);
                        idx+=1;
                        return token;
                    }}
                    function parsePrimary() {{
                        const token=peek();
                        if(token===undefined) throw new Error("Unexpected end of expression.");
                        if(token==="(") {{ consume("("); const node=parseOr(); if(peek()!==")") throw new Error("Missing closing parenthesis."); consume(")"); return node; }}
                        if(/^-?(?:\\d+\\.\\d*|\\d*\\.\\d+|\\d+)(?:[eE][+-]?\\d+)?$/.test(token)) {{ consume(); return {{type:"number", value:Number(token)}}; }}
                        if(/^[A-Za-z_][A-Za-z0-9_]*$/.test(token)) {{
                            if(!marsSummaryFilterColumns.has(token)) throw new Error(`Unknown metric: ${{token}}`);
                            consume();
                            return {{type:"identifier", value:token}};
                        }}
                        throw new Error(`Unexpected token: ${{token}}`);
                    }}
                    function parseComparison() {{
                        const left=parsePrimary();
                        const token=peek();
                        if(["<", "<=", ">", ">=", "==", "!="].includes(token)) {{
                            consume();
                            const right=parsePrimary();
                            return {{type:"compare", op:token, left, right}};
                        }}
                        if(left.type!=="identifier") throw new Error("Standalone values must be metric names.");
                        return left;
                    }}
                    function parseAnd() {{
                        let node=parseComparison();
                        while(peek()==="&") {{ consume("&"); node={{type:"and", left:node, right:parseComparison()}}; }}
                        return node;
                    }}
                    function parseOr() {{
                        let node=parseAnd();
                        while(peek()==="|") {{ consume("|"); node={{type:"or", left:node, right:parseAnd()}}; }}
                        return node;
                    }}
                    try {{
                        const ast=parseOr();
                        if(idx!==tokens.length) throw new Error(`Unexpected token: ${{peek()}}`);
                        return {{ok:true,ast}};
                    }} catch(err) {{
                        return {{ok:false,error:err.message}};
                    }}
                }}
                function marsEvaluateSummaryNode(node, metrics) {{
                    if(!node) return true;
                    if(node.type==="number") return node.value;
                    if(node.type==="identifier") return Number(metrics?.[node.value]);
                    if(node.type==="compare") {{
                        const left=Number(marsEvaluateSummaryNode(node.left, metrics));
                        const right=Number(marsEvaluateSummaryNode(node.right, metrics));
                        if(!Number.isFinite(left) || !Number.isFinite(right)) return false;
                        return node.op===">" ? left>right : node.op===">=" ? left>=right : node.op==="<" ? left<right : node.op==="<=" ? left<=right : node.op==="==" ? left===right : left!==right;
                    }}
                    if(node.type==="and") return Boolean(marsEvaluateSummaryNode(node.left, metrics)) && Boolean(marsEvaluateSummaryNode(node.right, metrics));
                    if(node.type==="or") return Boolean(marsEvaluateSummaryNode(node.left, metrics)) || Boolean(marsEvaluateSummaryNode(node.right, metrics));
                    return false;
                }}
                function marsSetSummaryExpression(value) {{
                    const expr=(value||"").trim();
                    if(!expr) {{
                        marsState.appliedSummaryExpression="";
                        marsSetError("mars-summary-expression-error", "");
                        marsQueueTextRefresh();
                        return;
                    }}
                    const parsed=marsParseSummaryExpression(expr);
                    if(!parsed.ok) {{
                        marsSetError("mars-summary-expression-error", parsed.error || "Invalid expression.");
                        marsQueueTextRefresh();
                        return;
                    }}
                    marsState.appliedSummaryExpression=expr;
                    marsSetError("mars-summary-expression-error", "");
                    marsQueueTextRefresh();
                }}
                function marsUpdateTableSpecialRows(table) {{ const rows=Array.from(table.querySelectorAll("tbody tr")); const visibleBySource=new Set(); const visibleByFeatureSource=new Set(); rows.forEach((row)=>{{ const role=row.dataset.role||"data"; if(role==="data"&&row.style.display!=="none") {{ const source=row.dataset.dataSource||""; const feature=row.dataset.feature||""; visibleBySource.add(source); visibleByFeatureSource.add(`${{source}}||${{feature}}`); }} }}); rows.forEach((row)=>{{ const role=row.dataset.role||"data"; if(role==="source") {{ row.style.display=visibleBySource.has(row.dataset.dataSource||"")?"":"none"; }} else if(role==="spacer") {{ const key=`${{row.dataset.dataSource||""}}||${{row.dataset.feature||""}}`; row.style.display=visibleByFeatureSource.has(key)?"":"none"; }} }}); }}
                function marsSourceSelected(source) {{ if(source==="__aggregate__") return true; const hasBoxes=document.querySelectorAll(".mars-source-checkbox").length>0; if(!hasBoxes) return true; return marsState.selectedSources.includes(source||"UNMAPPED"); }}
                function marsReadRowMetrics(row) {{ let metrics={{}}; try {{ metrics=JSON.parse(row.dataset.metrics||"{{}}"); }} catch(err) {{ metrics={{}}; }} return metrics; }}
                function marsSummaryRowAllowedWithoutLocal(row, globalMatcher=null, summaryParsed=null) {{
                    if(!row) return false;
                    const matcher=globalMatcher||marsBuildMatcher(marsState.globalQuery);
                    if(!matcher.ok) return false;
                    const parsed=summaryParsed||marsParseSummaryExpression(marsState.appliedSummaryExpression);
                    if(!parsed.ok) return false;
                    const source=row.dataset.dataSource||"UNMAPPED";
                    const text=row.dataset.searchText||row.textContent||"";
                    return marsSourceSelected(source) && matcher.match(text) && marsEvaluateSummaryNode(parsed.ast, marsReadRowMetrics(row));
                }}
                function marsGetSummaryFeatureAllowSet() {{
                    const table=document.getElementById("mars-summary-table");
                    if(!table) return null;
                    const globalMatcher=marsBuildMatcher(marsState.globalQuery);
                    if(!globalMatcher.ok) return null;
                    const parsed=marsParseSummaryExpression(marsState.appliedSummaryExpression);
                    if(!parsed.ok) return marsState.summaryAllowedFeatures;
                    const features=new Set();
                    table.querySelectorAll("tbody tr[data-feature]").forEach((row)=>{{
                        const feature=row.dataset.feature||"";
                        if(feature && marsSummaryRowAllowedWithoutLocal(row, globalMatcher, parsed)) features.add(feature);
                    }});
                    return features;
                }}
                function marsFeatureAllowed(feature) {{ if(!(marsState.summaryAllowedFeatures instanceof Set)) return true; return marsState.summaryAllowedFeatures.has(feature||""); }}
                function marsApplyTableFilter(tableId) {{
                    const table=document.getElementById(tableId);
                    if(!table) return;
                    const globalMatcher=marsBuildMatcher(marsState.globalQuery);
                    if(!globalMatcher.ok) {{ marsSetError("mars-global-error", `Invalid regex: ${{globalMatcher.error}}`); return; }}
                    marsSetError("mars-global-error", "");
                    const localMatcher=marsBuildMatcher(marsState.localQueries[tableId]||"");
                    if(!localMatcher.ok) {{ marsSetError(`${{tableId}}-error`, `Invalid regex: ${{localMatcher.error}}`); return; }}
                    marsSetError(`${{tableId}}-error`, "");
                    const summaryParsed=marsParseSummaryExpression(marsState.appliedSummaryExpression);
                    const isSummary=table.dataset.tableKind==="summary";
                    table.querySelectorAll("tbody tr").forEach((row)=>{{
                        const role=row.dataset.role||"data";
                        if(role!=="data") {{ row.style.display="none"; return; }}
                        const source=row.dataset.dataSource||"UNMAPPED";
                        const feature=row.dataset.feature||"";
                        const text=row.dataset.searchText||row.textContent||"";
                        const globalVisible=marsSourceSelected(source)&&globalMatcher.match(text);
                        if(!globalVisible) {{ row.style.display="none"; return; }}
                        const summaryVisible=isSummary
                            ? (summaryParsed.ok ? marsSummaryRowAllowedWithoutLocal(row, globalMatcher, summaryParsed) : true)
                            : marsFeatureAllowed(feature);
                        const visible=summaryVisible&&localMatcher.match(text);
                        row.style.display=visible?"":"none";
                    }});
                    marsUpdateTableSpecialRows(table);
                }}
                function marsSortTable(tableId, trigger) {{ const table=document.getElementById(tableId); if(!table) return; const th=trigger.closest("th"); const colIndex=Number(th.dataset.colIndex||Array.from(th.parentNode.children).indexOf(th)); if(colIndex<0) return; const tbody=table.querySelector("tbody"); const rows=Array.from(tbody.querySelectorAll("tr")).filter((row)=>(row.dataset.role||"data")==="data"); let nextDir="asc"; if(table.dataset.sortCol===String(colIndex)) nextDir=table.dataset.sortDir==="asc"?"desc":"asc"; const sortType=th.dataset.sortType||"text"; rows.sort((a,b)=>{{ const va=a.children[colIndex]?.dataset.sortValue||""; const vb=b.children[colIndex]?.dataset.sortValue||""; if(sortType==="number") {{ const na=Number(va), nb=Number(vb); const sa=Number.isFinite(na)?na:(nextDir==="asc"?Infinity:-Infinity); const sb=Number.isFinite(nb)?nb:(nextDir==="asc"?Infinity:-Infinity); return nextDir==="asc"?sa-sb:sb-sa; }} return nextDir==="asc"?va.localeCompare(vb,undefined,{{numeric:true,sensitivity:"base"}}):vb.localeCompare(va,undefined,{{numeric:true,sensitivity:"base"}}); }}); rows.forEach((row)=>tbody.appendChild(row)); table.dataset.sortCol=String(colIndex); table.dataset.sortDir=nextDir; th.dataset.sortDir=nextDir; marsApplyTableFilter(tableId); }}
                function marsUpdatePivotViews() {{ const targetValue=document.getElementById("mars-pivot-target")?.value||null; document.querySelectorAll(".mars-pivot-view").forEach((view)=>{{ const sameTarget=!targetValue||view.dataset.yValue===targetValue; view.style.display=sameTarget?"":"none"; }}); }}
                function marsUpdateChartViews() {{
                    const targetValue=document.getElementById("mars-chart-target")?.value||null;
                    const globalMatcher=marsBuildMatcher(marsState.globalQuery);
                    const localMatcher=marsBuildMatcher(marsState.localQueries["mars-chart-cards"]||"");
                    if(!globalMatcher.ok) {{ marsSetError("mars-global-error", `Invalid regex: ${{globalMatcher.error}}`); return; }}
                    marsSetError("mars-global-error", "");
                    if(!localMatcher.ok) {{ marsSetError("mars-chart-cards-error", `Invalid regex: ${{localMatcher.error}}`); return; }}
                    marsSetError("mars-chart-cards-error", "");
                    document.querySelectorAll(".mars-chart-view").forEach((view)=>{{
                        const visibleTarget=!targetValue||view.dataset.yValue===targetValue;
                        view.style.display=visibleTarget?"":"none";
                        if(!visibleTarget) return;
                        view.querySelectorAll(".mars-chart-card").forEach((card)=>{{
                            const source=card.dataset.dataSource||"UNMAPPED";
                            const feature=card.dataset.feature||"";
                            const text=card.dataset.searchText||card.textContent||"";
                            const globalVisible=marsSourceSelected(source)&&globalMatcher.match(text)&&marsFeatureAllowed(feature);
                            const visible=globalVisible&&localMatcher.match(text);
                            card.style.display=visible?"":"none";
                        }});
                    }});
                }}
                function marsBuildExportFeatureMap() {{
                    const table=document.getElementById("mars-summary-table");
                    if(!table) return {{}};
                    const summaryParsed=marsParseSummaryExpression(marsState.appliedSummaryExpression);
                    const featureMap=new Map();
                    const sourceOrder=Array.from(document.querySelectorAll(".mars-source-checkbox")).map((box)=>box.value);
                    table.querySelectorAll("tbody tr[data-feature]").forEach((row)=>{{
                        const source=row.dataset.dataSource||"UNMAPPED";
                        const feature=row.dataset.feature||"";
                        if(!feature || !marsSourceSelected(source)) return;
                        if(summaryParsed.ok && !marsEvaluateSummaryNode(summaryParsed.ast, marsReadRowMetrics(row))) return;
                        if(!featureMap.has(source)) featureMap.set(source, new Set());
                        featureMap.get(source).add(feature);
                    }});
                    const payload={{}};
                    const assignedSources=new Set();
                    sourceOrder.forEach((source)=>{{
                        const values=featureMap.has(source) ? Array.from(featureMap.get(source)).sort((a,b)=>a.localeCompare(b, undefined, {{numeric:true, sensitivity:"base"}})) : [];
                        if(values.length) {{
                            payload[source]=values;
                            assignedSources.add(source);
                        }}
                    }});
                    featureMap.forEach((features, source)=>{{
                        if(assignedSources.has(source)) return;
                        const values=Array.from(features).sort((a,b)=>a.localeCompare(b, undefined, {{numeric:true, sensitivity:"base"}}));
                        if(values.length) payload[source]=values;
                    }});
                    return payload;
                }}
                function marsDownloadTextFile(text, fileName) {{ const blob=new Blob([text], {{type:"text/plain;charset=utf-8"}}); const link=document.createElement("a"); link.href=URL.createObjectURL(blob); link.download=fileName; link.click(); URL.revokeObjectURL(link.href); }}
                function marsExportFeatures() {{ const featureMap=marsBuildExportFeatureMap(); marsDownloadTextFile(JSON.stringify(featureMap, null, 2), "mars_features.txt"); }}
                function marsColumnWidthProperty(columnKey) {{ return columnKey==="feature" ? "--mars-feature-col-width" : columnKey==="secondary" ? "--mars-secondary-col-width" : "--mars-bin-col-width"; }}
                function marsColumnDefaultWidth(columnKey) {{ return columnKey==="feature" ? 220 : columnKey==="secondary" ? 110 : 140; }}
                function marsColumnMinWidth(columnKey) {{ return columnKey==="feature" ? 140 : 90; }}
                function marsApplyColumnWidth(table, columnKey, width) {{
                    if(!table) return;
                    const safeWidth=Math.max(marsColumnMinWidth(columnKey), Number(width)||marsColumnDefaultWidth(columnKey));
                    table.style.setProperty(marsColumnWidthProperty(columnKey), `${{safeWidth}}px`);
                }}
                function marsSyncStickyLayout(table) {{
                    if(!table) return;
                    const featureHeader=table.querySelector("thead .mars-feature-col");
                    if(featureHeader) {{
                        const featureWidth=Math.max(140, Math.ceil(featureHeader.getBoundingClientRect().width || marsColumnDefaultWidth("feature")));
                        marsApplyColumnWidth(table, "feature", featureWidth);
                    }}
                    const secondaryHeader=table.querySelector("thead .mars-secondary-col");
                    if(secondaryHeader) {{
                        const secondaryWidth=Math.max(90, Math.ceil(secondaryHeader.getBoundingClientRect().width || marsColumnDefaultWidth("secondary")));
                        marsApplyColumnWidth(table, "secondary", secondaryWidth);
                    }}
                    const binHeader=table.querySelector("thead .mars-bin-col");
                    if(binHeader) {{
                        const binWidth=Math.max(90, Math.ceil(binHeader.getBoundingClientRect().width || marsColumnDefaultWidth("bin")));
                        marsApplyColumnWidth(table, "bin", binWidth);
                    }}
                }}
                function marsOpenAncestorSections(node) {{
                    let parent=node?.closest("details");
                    while(parent) {{
                        parent.open=true;
                        parent=parent.parentElement?.closest("details");
                    }}
                }}
                function marsFindSummaryFeatureNode(feature, visibleOnly=false) {{
                    const target=marsNormalizeFeatureValue(feature);
                    const nodes=Array.from(document.querySelectorAll("#mars-summary-table tbody tr[data-feature]"));
                    const candidateNodes=visibleOnly ? nodes.filter((node)=>node.style.display!=="none" && node.offsetParent!==null) : nodes;
                    for(const node of candidateNodes) {{
                        if(marsNormalizeFeatureValue(node.dataset.feature)===target) return node;
                    }}
                    for(const node of candidateNodes) {{
                        if(marsNormalizeFeatureValue(node.dataset.feature).includes(target)) return node;
                    }}
                    return null;
                }}
                function marsClearSummaryLocalQuery() {{
                    marsState.localQueries["mars-summary-table"]="";
                    const input=document.getElementById("mars-summary-table-query");
                    if(input) input.value="";
                }}
                function marsFocusSummaryFeature(node) {{
                    if(!node) return;
                    const featureCell=node.querySelector(".mars-feature-col");
                    const scrollBox=node.closest(".mars-table-scroll");
                    marsOpenAncestorSections(node);
                    window.requestAnimationFrame(() => {{
                        node.scrollIntoView({{behavior:"smooth", block:"center", inline:"nearest"}});
                        if(featureCell) featureCell.scrollIntoView({{behavior:"smooth", block:"nearest", inline:"start"}});
                        if(scrollBox) scrollBox.scrollTo({{left:0, behavior:"smooth"}});
                        node.classList.remove("mars-jump-highlight");
                        if(featureCell) featureCell.classList.remove("mars-jump-highlight-cell");
                        window.setTimeout(() => {{
                            node.classList.add("mars-jump-highlight");
                            if(featureCell) featureCell.classList.add("mars-jump-highlight-cell");
                        }}, 10);
                        window.setTimeout(() => {{
                            node.classList.remove("mars-jump-highlight");
                            if(featureCell) featureCell.classList.remove("mars-jump-highlight-cell");
                        }}, 1500);
                    }});
                }}
                function marsJumpToFeature() {{
                    const input=document.getElementById("mars-feature-jump-input");
                    const value=(input?.value||"").trim();
                    if(!value) {{
                        marsSetError("mars-feature-jump-error", "Enter a feature name to jump.");
                        return;
                    }}
                    let node=marsFindSummaryFeatureNode(value, true);
                    if(node) {{
                        marsSetError("mars-feature-jump-error", "");
                        marsFocusSummaryFeature(node);
                        return;
                    }}
                    node=marsFindSummaryFeatureNode(value, false);
                    if(!node) {{
                        marsSetError("mars-feature-jump-error", `Feature "${{value}}" is not shown in Summary.`);
                        return;
                    }}
                    const globalMatcher=marsBuildMatcher(marsState.globalQuery);
                    const summaryParsed=marsParseSummaryExpression(marsState.appliedSummaryExpression);
                    if(marsSummaryRowAllowedWithoutLocal(node, globalMatcher, summaryParsed)) {{
                        marsClearSummaryLocalQuery();
                        marsQueueRefresh();
                        window.requestAnimationFrame(() => {{
                            window.requestAnimationFrame(() => {{
                                const refreshedNode=marsFindSummaryFeatureNode(value, true) || marsFindSummaryFeatureNode(value, false);
                                marsSetError("mars-feature-jump-error", "");
                                marsFocusSummaryFeature(refreshedNode);
                            }});
                        }});
                        return;
                    }}
                    marsSetError("mars-feature-jump-error", `Feature "${{value}}" is hidden by data source, global search, or summary filter.`);
                }}
                function marsStartColumnResize(event, tableId, columnKey) {{
                    event.preventDefault();
                    event.stopPropagation();
                    const table=document.getElementById(tableId);
                    if(!table) return;
                    const property=marsColumnWidthProperty(columnKey);
                    const computed=getComputedStyle(table);
                    const startWidth=parseFloat(computed.getPropertyValue(property)) || marsColumnDefaultWidth(columnKey);
                    marsState.resizeState={{ tableId, columnKey, property, startX:event.clientX, startWidth, pendingWidth:startWidth }};
                    document.body.style.cursor="col-resize";
                    document.body.style.userSelect="none";
                }}
                function marsHandleColumnResize(event) {{
                    if(!marsState.resizeState) return;
                    const {{tableId, startX, startWidth, columnKey}} = marsState.resizeState;
                    const table=document.getElementById(tableId);
                    if(!table) return;
                    const minWidth=marsColumnMinWidth(columnKey);
                    const nextWidth=Math.max(minWidth, startWidth + (event.clientX - startX));
                    marsState.resizeState.pendingWidth=nextWidth;
                    if(marsState.resizeFrameScheduled) return;
                    marsState.resizeFrameScheduled=true;
                    window.requestAnimationFrame(() => {{
                        marsState.resizeFrameScheduled=false;
                        if(!marsState.resizeState) return;
                        const activeTable=document.getElementById(marsState.resizeState.tableId);
                        marsApplyColumnWidth(activeTable, marsState.resizeState.columnKey, marsState.resizeState.pendingWidth);
                    }});
                }}
                function marsStopColumnResize() {{
                    if(!marsState.resizeState) return;
                    const table=document.getElementById(marsState.resizeState.tableId);
                    if(table) marsSyncStickyLayout(table);
                    marsState.resizeState=null;
                    document.body.style.cursor="";
                    document.body.style.userSelect="";
                }}
                function marsRefreshFilters() {{
                    const summaryFeatures=marsGetSummaryFeatureAllowSet();
                    marsState.summaryAllowedFeatures=summaryFeatures instanceof Set ? summaryFeatures : null;
                    marsUpdatePivotViews();
                    marsUpdateChartViews();
                    document.querySelectorAll("table.mars-data-table[id]").forEach((table)=>marsApplyTableFilter(table.id));
                }}
                window.addEventListener("mousemove", marsHandleColumnResize);
                window.addEventListener("mouseup", marsStopColumnResize);
                window.addEventListener("DOMContentLoaded", () => {{
                    document.querySelectorAll("table.mars-data-table[id]").forEach((table)=>marsSyncStickyLayout(table));
                    marsSetDataSources();
                    marsQueueRefresh();
                }});
            </script>
        </body>
        </html>
        """

        sections_html = "".join(html_parts)
        global_tools_html = self._build_global_tools_html(
            feature_jump_html=feature_jump_html,
            source_options=source_options,
        )
        body_html = f"""
            <div id="mars-page-top" aria-hidden="true"></div>
            <div id="mars-floating-header-host" class="mars-floating-header-host" hidden>
                <div id="mars-floating-header-scroll" class="mars-floating-header-scroll"></div>
            </div>
            <div class="mars-page">
                <div class="mars-hero">
                    <h1>{html.escape(safe_report_name)}</h1>
                    <p>Interactive monitoring report with source-aware tables, Excel-like color scales, grouped pivot views, and shared trend charts.</p>
                    <div class="mars-meta">
                        <div class="mars-pill">Features: {n_features}</div>
                        <div class="mars-pill">Trend Metrics: {len(trend_pd_map)}</div>
                        <div class="mars-pill">Group By: {html.escape(str(group_label))}</div>
                    </div>
                    {global_tools_html}
                    <div id="mars-global-error" class="mars-search-error"></div>
                </div>
                <div class="mars-nav">{nav_html}</div>
                {sections_html}
                <div class="mars-footnote">HTML export is self-contained. detail_table remains available in Python and Excel workflows.</div>
            </div>
            <button id="mars-back-to-top" class="mars-back-to-top" type="button" aria-label="Back to top" onclick="marsBackToTop()">Top</button>
        """
        page_html = self._build_html_document(
            report_name=safe_report_name,
            styles=self._build_html_styles(),
            body_html=body_html,
            runtime_script=self._build_html_runtime_script(summary_filter_columns=summary_filter_columns),
        )

        with open(path, "w", encoding="utf-8") as f:
            f.write(page_html)

        logger.info("Exported evaluation report to HTML: %s", path)

    def show_summary(self, 
                     features: Optional[Union[str, List[str]]] = None
                     ) -> "pd.io.formats.style.Styler":
        """
        展示特征汇总评分表。

        Parameters
        ----------
        features : str or List[str], optional
            需要展示的特征名称。若为 ``None``，展示全部特征。

        Returns
        -------
        pd.io.formats.style.Styler
            样式化后的特征汇总表。
        """
        df: pd.DataFrame = _as_pandas_frame(self.summary_table).copy()
        
        # 特征筛选逻辑
        if features is not None:
            if isinstance(features, str):
                features = [features]
            df = df[df["feature"].isin(features)]
        
        # 多目标模式下，将 target 列提前，便于快速按目标查看结果。
        for t_col in ["target", "target_col", "y"]:
            if t_col in df.columns:
                cols = [t_col] + [c for c in df.columns if c != t_col]
                df = df[cols]
                break

        styler = df.style.set_caption("<b>Feature Performance Summary</b>").hide(axis="index")
        
        # 异常熔断：如果筛选后为空，直接返回表框架，避免底图渲染报错
        if df.empty:
            return styler
        
        if "psi_max" in df.columns:
            styler = styler.background_gradient(cmap="RdYlGn_r", subset=["psi_max"], vmin=0, vmax=0.25)
            
        if "iv" in df.columns:
            styler = styler.background_gradient(cmap="RdYlGn", subset=["iv"], vmin=0.02, vmax=0.2)
        if "auc" in df.columns:
            styler = styler.background_gradient(cmap="RdYlGn", subset=["auc"], vmin=0.5, vmax=0.65)
        if "ks" in df.columns:
            styler = styler.background_gradient(cmap="RdYlGn", subset=["ks"], vmin=5, vmax=20)

        if "rc_min" in df.columns:
            styler = styler.background_gradient(cmap="RdYlGn", subset=["rc_min"], vmin=0.5, vmax=1.0)
            
        if "mono" in df.columns:
            # coolwarm 色带: -1 为深蓝(单调递减)，0 为灰白(无单调性)，1 为深红(单调递增)
            styler = styler.background_gradient(cmap="coolwarm", subset=["mono"], vmin=-1, vmax=1)

        return styler.format("{:.4f}", subset=df.select_dtypes("number").columns)

    def show_trend(self, 
                   metric: str, 
                   features: Optional[Union[str, List[str]]] = None,
                   group_ascending: bool = True, 
                   sort_by: Union[str, List[str]] = "Total", 
                   sort_ascending: bool = False) -> "pd.io.formats.style.Styler":
        """
        展示指定指标的时间趋势热力图。

        渲染并返回一个带条件格式 (Conditional Formatting) 的 Pandas Styler 对象，
        用于直观分析特征在不同时间切片（或客群分组）下的指标波动趋势。内置了针对
        风控业务语义优化的专属色盘 (Colormap)。

        Parameters
        ----------
        metric : str
            需要展示的指标名称。支持的选项可通过 `self.trend_tables.keys()` 查看
            (通常包含 'psi', 'auc', 'ks', 'iv', 'bad_rate', 'risk_corr')。
        features : str or List[str], optional
            需要展示的特征名列表。若为 None，则展示所有特征。
        group_ascending : bool, default True
            分组/时间切片列的排序方向 (横向)。True 表示正序（从左到右由旧到新 / 由小到大）。
        sort_by : str or List[str], default "Total"
            特征行的排序依据列。默认按照全局表现 (Total) 排序。
        sort_ascending : bool, default False
            特征行的排序方向 (纵向)。默认降序 (False)，即把表现最差/最好的特征排在最上面。

        Returns
        -------
        pd.io.formats.style.Styler
            渲染完成的热力图对象。在 Jupyter Notebook 环境下会自动渲染为精美表格。

        Raises
        ------
        ValueError
            当 ``metric`` 不在当前报告支持的趋势指标集合中时抛出。
        """
        if metric not in self.trend_tables:
            raise ValueError(f"Unknown metric: {metric}. Options: {list(self.trend_tables.keys())}")
        
        # 转换为 Pandas 副本进行安全的样式处理
        df: pd.DataFrame = _as_pandas_frame(self.trend_tables[metric]).copy()
        
        # 特征筛选逻辑
        if features is not None:
            if isinstance(features, str):
                features = [features]
            df = df[df["feature"].isin(features)]

        # 行排序：紧跟 sort_by 和 sort_ascending 语义
        if sort_by in df.columns or (isinstance(sort_by, list) and all(c in df.columns for c in sort_by)):
            df = df.sort_values(by=sort_by, ascending=sort_ascending)
        
        # 识别列类型并重排时间切片列
        meta_cols = ["feature", "dtype"]
        special_cols = ["Total"]
        time_cols = [c for c in df.columns if c not in meta_cols + special_cols]
        
        # 列排序：受 group_ascending 控制
        time_cols_sorted = sorted(time_cols, reverse=not group_ascending)

        # 组装最终的列顺序：元数据 -> 时间切片 -> 汇总列
        final_cols = [c for c in meta_cols if c in df.columns] + \
                     time_cols_sorted + \
                     [c for c in special_cols if c in df.columns]
        df = df[final_cols]

        # 基础表格样式初始化
        styler = df.style.set_caption(f"<b>Trend Analysis: {metric.upper()}</b>").hide(axis="index")
        styler = styler.set_properties(subset=["feature"], **{'text-align': 'left', 'font-weight': 'bold'})

        if df.empty:
            return styler # 如果筛选后为空，直接返回空表格框架，避免报错

        # 根据不同业务指标的阈值与方向，映射专属渐变色盘
        if metric == "psi":
            styler = styler.background_gradient(
                cmap="RdYlGn_r", subset=time_cols_sorted, vmin=0, vmax=0.25, axis=None
            )
        elif metric in ["auc", "ks", "iv"]:
            styler = styler.background_gradient(
                cmap="RdYlGn", subset=time_cols_sorted, axis=None
            )
        elif metric == "bad_rate":
            styler = styler.background_gradient(
                cmap="Blues", subset=time_cols_sorted, axis=None
            )
        elif metric == "risk_corr":
            styler = styler.background_gradient(
                cmap="RdYlGn", subset=time_cols_sorted, vmin=0.5, vmax=1.0, axis=None
            )

        # 统一数值精度
        format_cols = [c for c in df.select_dtypes(include=[np.number]).columns]
        return styler.format("{:.4f}", subset=format_cols)


    def write_excel(self, path: str = "mars_bin_report.xlsx", engine: str = "openpyxl") -> None:
        """
        导出评估 Excel 报告。

        Parameters
        ----------
        path : str
            导出的 Excel 文件路径。
        engine : str, default="openpyxl"
            写入 Excel 的底层引擎。
            - "auto": 自动检测，Win/Mac 下优先尝试 xlwings，若失败或在 Linux 下则回退至 openpyxl。
            - "xlwings": 强制使用 xlwings 引擎 (依赖本地安装的 Excel 应用程序，格式保留最完美)。
            - "openpyxl": 强制使用 openpyxl 引擎 (无需安装 Excel，跨平台兼容性好)。

        Raises
        ------
        ValueError
            当 ``engine`` 不在支持列表中时抛出。
        """
        valid_engines = ["auto", "xlwings", "openpyxl"]
        if engine not in valid_engines:
            raise ValueError(f"不支持的 engine: '{engine}'，请从 {valid_engines} 中选择。")

        # 智能定位模板路径
        package_name = "mars.analysis" 
        template_name_xlwings = "mars_bin_report_win_mac.xlsx"
        template_name_openpyxl = "mars_bin_report_linux.xlsx"
        
        def get_template_path(fname):
            """解析 Excel 模板文件的物理路径。"""
            try:
                import importlib.resources as resources
                with resources.as_file(resources.files(package_name).joinpath(fname)) as p:
                    return str(p)
            except Exception:
                return os.path.join(os.path.dirname(os.path.abspath(__file__)), fname)

        # 配置内部参数
        START_WRITE_ROW = 4
        STYLE_SOURCE_ROW = 2
        FONT_NAME = "Microsoft YaHei"
        FONT_SIZE = 8
        SHEET_NAME = "分组明细"

        # 引擎解析与初始化
        use_xlwings = False

        if engine == "xlwings":
            use_xlwings = True
        elif engine == "openpyxl":
            use_xlwings = False
        else:  # "auto"
            # 自动探测环境
            is_gui_env = sys.platform.startswith("win") or sys.platform.startswith("darwin")
            use_xlwings = is_gui_env

        # 验证 xlwings 可用性
        if use_xlwings:
            try:
                import xlwings as xw
                # 测试 Excel 应用程序是否真正可用
                xw.App(visible=False, add_book=False).quit()
                template_path = get_template_path(template_name_xlwings)
            except Exception as e:
                if engine == "xlwings":
                    # 用户强制要求但失败，直接抛错
                    raise RuntimeError(f"强制使用 xlwings 引擎失败，请确认系统已正确安装 Excel 及 xlwings 库。\n报错详情: {e}")
                else:
                    # auto 模式下失败，降级处理
                    logger.warning("xlwings 启动失败，将降级使用 openpyxl 引擎: %s", e)
                    use_xlwings = False

        # 若未使用 xlwings，则准备 openpyxl 依赖
        if not use_xlwings:
            import openpyxl
            from openpyxl.utils import get_column_letter
            from openpyxl.worksheet.table import Table, TableStyleInfo
            template_path = get_template_path(template_name_openpyxl)

        if not os.path.exists(template_path):
            raise FileNotFoundError(f"找不到模板文件: {template_path}")

        # 准备数据
        df_pd = _as_pandas_frame(self.detail_table)
        total_cols = len(df_pd.columns)

        # ================= 路径 A: xlwings 写入 (Win/Mac 跨平台兼容) =================
        if use_xlwings:
            app = None
            try:
                app = xw.App(visible=False, add_book=False)
                app.display_alerts = False
                app.screen_updating = False
                
                wb = app.books.open(template_path)
                ws = wb.sheets[SHEET_NAME]
                
                # 防止 Excel 将长数字字符串转为科学计数法
                if 'mars_group' in df_pd.columns:
                    df_pd['mars_group'] = "'" + df_pd['mars_group'].astype(str)
                
                # 写入数据
                ws.range((START_WRITE_ROW, 1)).value = df_pd.values
                final_row = START_WRITE_ROW + len(df_pd) - 1
                
                # 样式格式刷 (跨平台原生写法)
                if final_row >= START_WRITE_ROW:
                    src_row = int(STYLE_SOURCE_ROW)
                    start_row = int(START_WRITE_ROW)
                    end_row = int(final_row)
                    max_col = int(total_cols)

                    source_range = ws.range((src_row, 1), (src_row, max_col))
                    data_range = ws.range((start_row, 1), (end_row, max_col))
                    
                    source_range.copy()
                    data_range.paste(paste='formats') 
                
                # 统一字体 (跨平台原生写法)
                full_range = ws.range((1, 1), (final_row, total_cols))
                full_range.font.name = FONT_NAME
                full_range.font.size = FONT_SIZE
                
                # 更新超级表 ListObject (跨平台原生写法)
                if len(ws.tables) > 0:
                    table = ws.tables[0]
                    new_ref_range = ws.range((1, 1), (final_row, total_cols))
                    table.resize(new_ref_range)
                
                # 清理旧数据 (跨平台原生写法)
                last_used_row = ws.used_range.last_cell.row
                if last_used_row > final_row:
                    ws.range(f"{final_row + 1}:{last_used_row}").delete()

                wb.save(path)
                logger.info("Exported evaluation report via xlwings: %s", path)

            except Exception as e:
                logger.exception("xlwings 导出过程出错。")
                raise RuntimeError(f"xlwings 导出过程出错: {e}")
            finally:
                if 'wb' in locals() and wb: wb.close()
                if app: app.quit() 

        # ================= 路径 B: openpyxl 写入 (Linux 等无界面的兜底方案) =================
        else:
            wb = openpyxl.load_workbook(template_path)
            ws = wb[SHEET_NAME]
            
            mars_group_idx = -1
            if "mars_group" in df_pd.columns:
                mars_group_idx = list(df_pd.columns).index("mars_group") + 1
            
            # 提取并缓存样式模板
            style_map = {}
            for c in range(1, total_cols + 1):
                cell = ws.cell(row=STYLE_SOURCE_ROW, column=c)
                style_map[c] = {
                    "font": copy(cell.font),
                    "border": copy(cell.border),
                    "fill": copy(cell.fill),
                    "alignment": copy(cell.alignment),
                    "number_format": cell.number_format
                }

            # 写入数据
            rows = df_pd.values.tolist()
            for r_offset, row_data in enumerate(rows):
                current_row = START_WRITE_ROW + r_offset
                for c_offset, value in enumerate(row_data):
                    c_idx = c_offset + 1
                    cell = ws.cell(row=current_row, column=c_idx, value=value)
                    
                    # 应用样式
                    if c_idx in style_map:
                        s = style_map[c_idx]
                        cell.font = s["font"]
                        cell.border = s["border"]
                        cell.fill = s["fill"]
                        cell.alignment = s["alignment"]
                        cell.number_format = s["number_format"]
                    
                    # 日期列单独处理
                    if c_idx == mars_group_idx:
                        cell.number_format = "yyyy-mm-dd"

            final_row = START_WRITE_ROW + len(rows) - 1

            # 更新超级表范围并容错
            if hasattr(ws, 'tables') and ws.tables:
                new_ref = f"A1:{get_column_letter(total_cols)}{final_row}"
                for tbl_name in list(ws.tables.keys()):
                    tbl_obj = ws.tables[tbl_name]
                    
                    if hasattr(tbl_obj, 'ref'):
                        tbl_obj.ref = new_ref
                    else:
                        # 容错：处理 openpyxl 偶尔将 Table 解析为纯字符串的问题
                        del ws.tables[tbl_name]
                        new_tbl = Table(displayName=tbl_name, ref=new_ref)
                        style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
                        new_tbl.tableStyleInfo = style
                        ws.add_table(new_tbl)

            # 删除多余行
            if ws.max_row > final_row:
                ws.delete_rows(final_row + 1, ws.max_row - final_row)

            wb.save(path)
            logger.info("Exported evaluation report via openpyxl: %s", path)
