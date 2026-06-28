"""分箱评估 Excel 导出实现。"""

from __future__ import annotations

import importlib.util
import sys
from pathlib import Path
from typing import Any, List, Tuple

import pandas as pd

from mars.compute import to_pandas_frame
from mars.utils.logger import logger


class _BinningExcelWriter:
    """分箱报告 Excel 导出能力。"""

    def __init__(self, report: Any) -> None:
        self._report = report

    def __getattr__(self, name: str) -> Any:
        """将只读数据访问委托给 report 容器。"""
        return getattr(self._report, name)

    @classmethod
    def _resolve_excel_template_path(
        cls: type[_BinningExcelWriter],
        file_name: str,
    ) -> Path:
        """解析 `mars.reporting.template` 下的 Excel 模板路径。"""
        reporting_spec = importlib.util.find_spec("mars.reporting")
        if reporting_spec is None or reporting_spec.origin is None:
            raise FileNotFoundError("无法定位 `mars.reporting` 模块，因此无法读取 Excel 模板。")

        reporting_dir = Path(reporting_spec.origin).resolve().parent
        template_path = reporting_dir / "template" / file_name
        if not template_path.exists():
            raise FileNotFoundError(f"找不到 Excel 模板文件: {template_path}")
        return template_path

    @classmethod
    def _read_detail_template_schema(
        cls: type[_BinningExcelWriter],
        template_path: Path,
        sheet_name: str,
    ) -> Tuple[List[str], int, int]:
        """读取明细 sheet 的模板列与表格列范围。"""
        import openpyxl
        from openpyxl.utils.cell import range_boundaries

        workbook = openpyxl.load_workbook(template_path)
        try:
            worksheet = workbook[sheet_name]
            if not worksheet.tables:
                raise ValueError(f"sheet `{sheet_name}` 缺少 Excel table，无法按模板导出明细表。")

            table = next(iter(worksheet.tables.values()))
            table_ref = getattr(table, "ref", None)
            if not isinstance(table_ref, str) or not table_ref:
                raise ValueError(
                    f"sheet `{sheet_name}` 的 Excel table 缺少 `ref`，无法读取模板列结构。"
                )

            first_col, _, last_col, _ = range_boundaries(table_ref)
            headers: List[str] = []
            for col_idx in range(first_col, last_col + 1):
                header_value = worksheet.cell(row=1, column=col_idx).value
                if header_value is None or not str(header_value).strip():
                    raise ValueError(
                        f"sheet `{sheet_name}` 第 1 行第 {col_idx} 列表头为空，无法按模板导出。"
                    )
                headers.append(str(header_value))

            return headers, first_col, last_col
        finally:
            workbook.close()

    @classmethod
    def _build_excel_detail_export_frame(
        cls: type[_BinningExcelWriter],
        detail_df: pd.DataFrame,
        template_headers: List[str],
    ) -> pd.DataFrame:
        """按模板列顺序构建 Excel 明细导出数据。"""
        missing_columns = [
            column_name
            for column_name in template_headers
            if column_name not in detail_df.columns
        ]
        if missing_columns:
            missing_display = ", ".join(missing_columns)
            raise ValueError(
                "Excel 模板列在 `detail_table` 中缺失："
                f" {missing_display}"
            )

        return detail_df.loc[:, template_headers].copy()
    def write_excel(self: Any, path: str = "mars_bin_report.xlsx", engine: str = "openpyxl") -> None:
        """
        导出评估 Excel 报告。

        Parameters
        ----------
        path : str
            导出的 Excel 文件路径。
        engine : str
            写入 Excel 的底层引擎。
            - "auto": 自动检测，Win/Mac 下优先尝试 xlwings，若失败或在 Linux 下则回退至 openpyxl。
            - "xlwings": 强制使用 xlwings 引擎 (依赖本地安装的 Excel 应用程序，格式保留最完美)。
            - "openpyxl": 强制使用 openpyxl 引擎 (无需安装 Excel，跨平台兼容性好)。

        Raises
        ------
        ValueError
            当 ``engine`` 不在支持列表中时抛出。
        RuntimeError
            当底层 Excel 导出流程失败时抛出。

        Examples
        --------
        >>> import polars as pl
        >>> from pathlib import Path
        >>> from tempfile import TemporaryDirectory
        >>> summary = pl.DataFrame({"feature": ["age"], "iv": [0.12], "ks": [18.0]})
        >>> detail = pl.DataFrame({"feature": ["age"], "bin_index": [0], "count": [100]})
        >>> report = MarsBinningReport(summary, {}, detail)
        >>> with TemporaryDirectory() as tmp:
        ...     report.write_excel(str(Path(tmp) / "evaluation.xlsx"), engine="openpyxl") is None
        True
        """
        valid_engines = ["auto", "xlwings", "openpyxl"]
        if engine not in valid_engines:
            raise ValueError(f"不支持的 engine: '{engine}'，请从 {valid_engines} 中选择。")

        start_write_row = 4
        sheet_name = "分组明细"
        template_name_xlwings = "mars_bin_report_win_mac.xlsx"
        template_name_openpyxl = "mars_bin_report_linux.xlsx"
        is_gui_env = sys.platform.startswith("win") or sys.platform.startswith("darwin")
        use_xlwings = engine == "xlwings" or (engine == "auto" and is_gui_env)

        if use_xlwings:
            try:
                import xlwings as xw

                probe_app = xw.App(visible=False, add_book=False)
                probe_app.quit()
                template_path = self._resolve_excel_template_path(template_name_xlwings)
            except Exception as exc:
                if engine == "xlwings":
                    raise RuntimeError(
                        "强制使用 xlwings 导出失败，请确认本机已安装 Excel 与 xlwings。"
                        f"\n错误详情: {exc}"
                    ) from exc
                logger.warning("xlwings 不可用，自动降级为 openpyxl 导出: %s", exc)
                use_xlwings = False

        if not use_xlwings:
            template_path = self._resolve_excel_template_path(template_name_openpyxl)

        template_headers, first_col, last_col = self._read_detail_template_schema(
            template_path=template_path,
            sheet_name=sheet_name,
        )
        detail_pd: pd.DataFrame = to_pandas_frame(self.detail_table).copy()
        export_pd = self._build_excel_detail_export_frame(
            detail_df=detail_pd,
            template_headers=template_headers,
        )
        rows: List[List[Any]] = export_pd.values.tolist()
        final_row = start_write_row + len(rows) - 1
        table_last_row = max(final_row, start_write_row - 1)

        if use_xlwings:
            app = None
            workbook = None
            try:
                app = xw.App(visible=False, add_book=False)
                app.display_alerts = False
                app.screen_updating = False

                workbook = app.books.open(str(template_path))
                worksheet = workbook.sheets[sheet_name]

                if rows:
                    worksheet.range((start_write_row, first_col)).value = rows

                table = worksheet.tables[0] if worksheet.tables else None
                if table is None:
                    raise ValueError(f"sheet `{sheet_name}` 缺少 Excel table，无法按模板导出明细表。")
                table.resize(
                    worksheet.range(
                        (1, first_col),
                        (table_last_row, last_col),
                    )
                )

                last_used_row = worksheet.used_range.last_cell.row
                if last_used_row > table_last_row:
                    worksheet.range(f"{table_last_row + 1}:{last_used_row}").delete()

                workbook.save(path)
                logger.info("Exported binning report via xlwings: %s", path)
            except Exception as exc:
                raise RuntimeError(f"xlwings 导出过程中发生错误: {exc}") from exc
            finally:
                if workbook is not None:
                    workbook.close()
                if app is not None:
                    app.quit()
            return

        import openpyxl
        from openpyxl.utils import get_column_letter

        workbook = openpyxl.load_workbook(template_path)
        try:
            worksheet = workbook[sheet_name]
            for row_offset, row_data in enumerate(rows):
                current_row = start_write_row + row_offset
                for col_offset, value in enumerate(row_data):
                    worksheet.cell(
                        row=current_row,
                        column=first_col + col_offset,
                        value=value,
                    )

            if not worksheet.tables:
                raise ValueError(f"sheet `{sheet_name}` 缺少 Excel table，无法按模板导出明细表。")

            table = next(iter(worksheet.tables.values()))
            table_ref = (
                f"{get_column_letter(first_col)}1:"
                f"{get_column_letter(last_col)}{table_last_row}"
            )
            if not hasattr(table, "ref"):
                raise ValueError(f"sheet `{sheet_name}` 的 Excel table 缺少 `ref`，无法更新表格范围。")
            table.ref = table_ref

            if worksheet.max_row > table_last_row:
                worksheet.delete_rows(table_last_row + 1, worksheet.max_row - table_last_row)

            workbook.save(path)
            logger.info("Exported binning report via openpyxl: %s", path)
        finally:
            workbook.close()
        return
